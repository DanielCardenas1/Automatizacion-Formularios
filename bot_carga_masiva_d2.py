#!/usr/bin/env python3
"""
Bot Selenium - Carga Masiva de DUITAMA D2
Lee todos los registros de DUITAMA D2 del Excel
y los ingresa completamente en RUB Online
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
from pathlib import Path
from datetime import datetime

print("\n" + "="*70)
print("BOT CARGA MASIVA - DUITAMA D2 (19 registros)")
print("="*70)

# ========== PASO 0: LEER EXCEL Y OBTENER REGISTROS D2 ==========
print("\n[*] Paso 0: Leyendo registros de DUITAMA D2...")

ruta_excel = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"
ruta_fotos_d2 = Path("/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/DUITAMA D2/DUITAMA D2 FOTOS")
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

registros_d2 = []

for row_idx in range(3, ws.max_row + 1):
    nombre_uds = ws.cell(row=row_idx, column=3).value
    
    # Si es DUITAMA D2 (sin D3)
    if nombre_uds and "D2" in str(nombre_uds).upper() and "D3" not in str(nombre_uds).upper():
        datos = {
            'fila': row_idx,
            'documento': str(int(ws.cell(row=row_idx, column=17).value)) if ws.cell(row=row_idx, column=17).value else "",
            'primer_nombre': str(ws.cell(row=row_idx, column=5).value).strip() if ws.cell(row=row_idx, column=5).value else "",
            'segundo_nombre': str(ws.cell(row=row_idx, column=6).value).strip() if ws.cell(row=row_idx, column=6).value else "",
            'primer_apellido': str(ws.cell(row=row_idx, column=7).value).strip() if ws.cell(row=row_idx, column=7).value else "",
            'segundo_apellido': str(ws.cell(row=row_idx, column=8).value).strip() if ws.cell(row=row_idx, column=8).value else "",
            'sexo': str(ws.cell(row=row_idx, column=9).value).strip() if ws.cell(row=row_idx, column=9).value else "",
            'fecha_nac': ws.cell(row=row_idx, column=13).value,
            'fecha_ingreso': ws.cell(row=row_idx, column=4).value,
            'tipo_doc': str(ws.cell(row=row_idx, column=16).value).strip() if ws.cell(row=row_idx, column=16).value else "",
        }
        registros_d2.append(datos)

wb.close()

print(f"[+] Total registros D2 encontrados: {len(registros_d2)}")
for i, reg in enumerate(registros_d2[:3], 1):
    print(f"    {i}. {reg['documento']} - {reg['primer_nombre']} {reg['primer_apellido']}")
print(f"    ... y {len(registros_d2)-3} más\n")

# ========== CONFIGURAR SELENIUM ==========
options = webdriver.ChromeOptions()
options.add_argument('--disable-notifications')
options.add_argument('--disable-popup-blocking')

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 20)

# Variables globales para el proceso
registros_procesados = 0
registros_exitosos = 0
registros_error = []

def navegar_a_formulario():
    """Navega al formulario de nuevo beneficiario"""
    print("\n    [*] Navegando al formulario...")
    
    try:
        # Click RUB online
        time.sleep(2)
        enlacesRUB = driver.find_elements(By.XPATH, "//a[contains(text(), 'Rub online')]")
        if enlacesRUB:
            driver.execute_script("arguments[0].scrollIntoView(true);", enlacesRUB[0])
            time.sleep(0.5)
            ActionChains(driver).move_to_element(enlacesRUB[0]).click().perform()
            time.sleep(3)
        
        # Click primer Beneficiario
        time.sleep(1)
        enlacesBene = driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")
        if enlacesBene:
            driver.execute_script("arguments[0].scrollIntoView(true);", enlacesBene[0])
            time.sleep(0.5)
            ActionChains(driver).move_to_element(enlacesBene[0]).click().perform()
            time.sleep(3)
        
        # Click segundo Beneficiario
        time.sleep(1)
        enlacesBene = driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")
        if len(enlacesBene) > 1:
            driver.execute_script("arguments[0].scrollIntoView(true);", enlacesBene[1])
            time.sleep(0.5)
            ActionChains(driver).move_to_element(enlacesBene[1]).click().perform()
            time.sleep(4)
        
        print("    [+] Navegación completada")
        return True
    except Exception as e:
        print(f"    [-] Error en navegación: {str(e)}")
        return False

def clickear_boton_nuevo():
    """Hace clic en el botón '+' para crear nuevo beneficiario"""
    print("    [*] Ingresando nuevo beneficiario (+)...")
    
    try:
        time.sleep(2)
        iframe = wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
        driver.switch_to.frame(iframe)
        
        boton_nuevo = wait.until(EC.element_to_be_clickable((By.ID, "btnNuevo")))
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_nuevo)
        time.sleep(0.5)
        ActionChains(driver).move_to_element(boton_nuevo).click().perform()
        print("    [+] Botón '+' clickeado")
        time.sleep(3)
        return True
    except Exception as e:
        print(f"    [-] Error: {str(e)}")
        return False

def llenar_filtros_fijos():
    """Llena los filtros que son iguales para todos los registros"""
    print("    [*] Llenando filtros fijos...")
    
    try:
        # Esperar a que el iframe esté listo
        time.sleep(2)
        
        # 1. Uno a uno
        try:
            radioBtns = driver.find_elements(By.XPATH, "//input[@type='radio']")
            for radio in radioBtns:
                if radio.get_attribute('value') and 'Uno a uno' in radio.get_attribute('value'):
                    radio.click()
                    break
        except:
            pass
        time.sleep(1)
        
        # 2. Dirección
        try:
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDireccionesICBF")))
            Select(select_elem).select_by_visible_text("Dirección de Primera Infancia")
        except:
            pass
        time.sleep(1)
        
        # 3. Regional
        try:
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlRegional")))
            Select(select_elem).select_by_visible_text("Boyacá")
        except:
            pass
        time.sleep(1)
        
        # 4. Vigencia
        try:
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdVigencia")))
            Select(select_elem).select_by_visible_text("2026")
        except:
            pass
        time.sleep(1)
        
        # 5. Contrato
        try:
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNumeroContrato")))
            Select(select_elem).select_by_visible_text("OD 15 420272 00015 2026")
        except:
            pass
        time.sleep(1)
        
        # 6. Servicio
        try:
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNombreServicio")))
            Select(select_elem).select_by_visible_text("EDUCACIÓN INICIAL EN EL HOGAR - FAMILIAR Y COMUNITARIA - 420272-2026")
        except:
            pass
        time.sleep(2)
        
        # 7. UDS
        try:
            selects = driver.find_elements(By.TAG_NAME, "select")
            for select_elem in selects:
                options = select_elem.find_elements(By.TAG_NAME, "option")
                for opt in options:
                    if "DUITAMA D2" in opt.text:
                        Select(select_elem).select_by_value(opt.get_attribute("value"))
                        time.sleep(1)
                        break
        except:
            pass
        
        # 8. Tipo de beneficiario
        try:
            driver.execute_script("window.scrollBy(0, 300);")
            time.sleep(1)
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdTipoBeneficiario")))
            options = select_elem.find_elements(By.TAG_NAME, "option")
            for opt in options:
                if "NIÑO O NIÑA ENTRE 6 MESES" in opt.text:
                    Select(select_elem).select_by_value(opt.get_attribute("value"))
                    break
        except:
            pass
        time.sleep(1)
        
        # 9. Tipo Documento
        try:
            driver.execute_script("window.scrollBy(0, 300);")
            time.sleep(1)
            select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdTipoDocumento")))
            Select(select_elem).select_by_visible_text("REGISTRO CIVIL")
        except:
            pass
        time.sleep(1)
        
        print("    [+] Filtros completados")
        return True
    except Exception as e:
        print(f"    [-] Error en filtros: {str(e)}")
        return False

def llenar_datos_beneficiario(datos):
    """Llena los datos específicos del beneficiario"""
    print(f"    [*] Ingresando datos: {datos['primer_nombre']} {datos['primer_apellido']}")
    
    try:
        time.sleep(3)
        
        # Hacer scroll al top del iframe para que aparezca el campo de documento
        print(f"      [*] Haciendo scroll dentro del iframe...")
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)
        
        # PASO 1: Buscar y llenar el documento
        print(f"      [1] Buscando campo de documento...")
        
        # Intentar con varios selectores posibles
        selectores_doc = [
            (By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdentificacion"),
            (By.NAME, "ctl00$cphCont$TabContainer1$tbnDatosB$BeneficiarioVincula$datosBasicosPersona$txtIdentificacion"),
            (By.XPATH, "//input[contains(@name, 'txtIdentificacion')]"),
            (By.XPATH, "//input[contains(@id, 'Identificacion')]"),
        ]
        
        campo_doc = None
        for selector in selectores_doc:
            try:
                campo_doc = driver.find_element(selector[0], selector[1])
                print(f"      [+] Campo encontrado")
                break
            except:
                pass
        
        if campo_doc:
            # Hacer el elemento visible usando JavaScript
            print(f"      [*] Haciendo visible el campo...")
            driver.execute_script("""
                arguments[0].style.display = 'block';
                arguments[0].style.visibility = 'visible';
                arguments[0].style.opacity = '1';
                arguments[0].offsetParent.style.display = 'block';
            """, campo_doc)
            time.sleep(0.5)
            
            # Scroll dentro del iframe
            driver.execute_script("arguments[0].parentElement.scrollIntoView(true);", campo_doc)
            time.sleep(0.5)
            
            # Hacer clic en el campo para habilitarlo
            print(f"      [*] Haciendo clic en el campo...")
            try:
                # Usar JavaScript para hacer click si Selenium falla
                driver.execute_script("arguments[0].click();", campo_doc)
                time.sleep(0.3)
            except Exception as e:
                print(f"      [!] JavaScript click no funcionó, intentando con ActionChains...")
                ActionChains(driver).move_to_element(campo_doc).click().perform()
                time.sleep(0.3)
            
            # Escribir el documento
            print(f"      [*] Escribiendo documento: {datos['documento']}")
            campo_doc.clear()
            campo_doc.send_keys(datos['documento'])
            print(f"      [+] Documento ingresado")
            time.sleep(1)
            
            # PASO 2: Buscar y hacer click en la lupa (junto al campo)
            print(f"      [2] Buscando botón de búsqueda (lupa)...")
            
            lupa_encontrada = False
            
            # Primero: Buscar lupa globalmente por ID
            try:
                lupa = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnBuscar")
                print(f"      [+] Lupa encontrada globalmente por ID")
                driver.execute_script("arguments[0].scrollIntoView(true);", lupa)
                time.sleep(0.3)
                lupa.click()
                print(f"      [+] Click en lupa realizado")
                time.sleep(3)
                lupa_encontrada = True
            except:
                print(f"      [-] Lupa no encontrada por ID global")
            
            # Segundo: Si no se encuentra, buscar en el contenedor padre
            if not lupa_encontrada:
                try:
                    padre = campo_doc.find_element(By.XPATH, "..")
                    selectores_lupa = [
                        (By.XPATH, ".//input[@value='Buscar']"),
                        (By.XPATH, ".//button[contains(text(), 'Buscar')]"),
                        (By.XPATH, ".//img[@title='Buscar']"),
                        (By.XPATH, ".//img[contains(@alt, 'buscar')]"),
                        (By.XPATH, ".//*[contains(@class, 'btnBuscar')]"),
                    ]
                    
                    for selector in selectores_lupa:
                        try:
                            lupa = padre.find_element(selector[0], selector[1])
                            print(f"      [+] Lupa encontrada en el contenedor padre")
                            driver.execute_script("arguments[0].scrollIntoView(true);", lupa)
                            time.sleep(0.3)
                            try:
                                lupa.click()
                            except:
                                ActionChains(driver).move_to_element(lupa).click().perform()
                            print(f"      [+] Click en lupa realizado")
                            time.sleep(3)
                            lupa_encontrada = True
                            break
                        except:
                            pass
                    
                    if not lupa_encontrada:
                        print(f"      [-] Lupa no encontrada en el contenedor")
                except:
                    print(f"      [-] Error buscando en elemento padre")
            
            # Si no se encuentra en el padre, buscar globalmente
            if not lupa_encontrada:
                print(f"      [*] Buscando lupa globalmente...")
                selectores_lupa_global = [
                    (By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnBuscar"),
                    (By.XPATH, "//input[@value='Buscar']"),
                    (By.XPATH, "//input[@type='button' and @value='Buscar']"),
                    (By.XPATH, "//button[contains(text(), 'Buscar')]"),
                    (By.XPATH, "//img[@title='Buscar']"),
                    (By.XPATH, "//img[@alt='buscar']"),
                    (By.XPATH, "//*[contains(@class, 'btnBuscar')]"),
                ]
                
                for selector in selectores_lupa_global:
                    try:
                        elementos = driver.find_elements(selector[0], selector[1])
                        if elementos:
                            lupa = elementos[0]
                            print(f"      [+] Lupa encontrada globalmente")
                            driver.execute_script("arguments[0].scrollIntoView(true);", lupa)
                            time.sleep(0.3)
                            try:
                                lupa.click()
                            except:
                                ActionChains(driver).move_to_element(lupa).click().perform()
                            print(f"      [+] Click en lupa realizado")
                            time.sleep(3)
                            lupa_encontrada = True
                            break
                    except:
                        pass
            
            if not lupa_encontrada:
                print(f"      [-] No se encontró botón de búsqueda")
        else:
            print(f"      [-] Campo de documento no encontrado")
        
        # PASO 3: Llenar resto de datos
        print(f"      [3] Llenando nombres...")
        
        # Intentar llenar con diferentes IDs posibles
        selectores_nombres = {
            'primer_nombre': [
                (By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdPrimerNombre"),
                (By.XPATH, "//input[contains(@name, 'PrimerNombre')]"),
                (By.XPATH, "//input[@placeholder='Primer Nombre']"),
            ],
            'segundo_nombre': [
                (By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdSegundoNombre"),
                (By.XPATH, "//input[contains(@name, 'SegundoNombre')]"),
            ],
            'primer_apellido': [
                (By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdPrimerApellido"),
                (By.XPATH, "//input[contains(@name, 'PrimerApellido')]"),
            ],
            'segundo_apellido': [
                (By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdSegundoApellido"),
                (By.XPATH, "//input[contains(@name, 'SegundoApellido')]"),
            ],
        }
        
        for campo_type, selectores in selectores_nombres.items():
            valor = datos.get(campo_type, "")
            if valor:
                campo_encontrado = False
                for selector in selectores:
                    try:
                        campo = driver.find_element(selector[0], selector[1])
                        driver.execute_script("arguments[0].scrollIntoView(true);", campo)
                        time.sleep(0.2)
                        campo.click()
                        time.sleep(0.2)
                        campo.clear()
                        campo.send_keys(valor)
                        print(f"      [+] {campo_type}: {valor}")
                        campo_encontrado = True
                        time.sleep(0.3)
                        break
                    except:
                        pass
                if not campo_encontrado:
                    print(f"      [-] Campo {campo_type} no encontrado")
        
        # Sexo
        try:
            if datos['sexo'] and datos['sexo'].upper() in ['MASCULINO', 'FEMENINO']:
                select_sexo = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdSexo")
                from selenium.webdriver.support.select import Select
                Select(select_sexo).select_by_visible_text(datos['sexo'].upper())
                print(f"      [+] Sexo: {datos['sexo'].upper()}")
        except:
            pass
        time.sleep(0.5)
        
        # Fechas
        try:
            if datos['fecha_nac']:
                campo_fnac = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdFechaNacimiento")
                driver.execute_script("arguments[0].scrollIntoView(true);", campo_fnac)
                time.sleep(0.2)
                campo_fnac.click()
                time.sleep(0.2)
                campo_fnac.clear()
                fecha_formato = datos['fecha_nac'].strftime("%d/%m/%Y")
                campo_fnac.send_keys(fecha_formato)
                print(f"      [+] Fecha Nac: {fecha_formato}")
        except:
            pass
        time.sleep(0.3)
        
        try:
            if datos['fecha_ingreso']:
                campo_fingreso = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdFechaIngreso")
                driver.execute_script("arguments[0].scrollIntoView(true);", campo_fingreso)
                time.sleep(0.2)
                campo_fingreso.click()
                time.sleep(0.2)
                campo_fingreso.clear()
                fecha_ingreso_formato = datos['fecha_ingreso'].strftime("%d/%m/%Y")
                campo_fingreso.send_keys(fecha_ingreso_formato)
                print(f"      [+] Fecha Ingreso: {fecha_ingreso_formato}")
        except:
            pass
        time.sleep(0.3)

        # Fecha de atencion
        try:
            campo_fatencion = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_cuwFechaAtencion_txtFecha")
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_fatencion)
            time.sleep(0.2)
            fecha_atencion_actual = (campo_fatencion.get_attribute("value") or "").strip()
            fecha_atencion_formato = fecha_atencion_actual
            if not fecha_atencion_formato and datos['fecha_ingreso']:
                fecha_atencion_formato = datos['fecha_ingreso'].strftime("%d/%m/%Y")
                campo_fatencion.click()
                time.sleep(0.2)
                campo_fatencion.clear()
                campo_fatencion.send_keys(fecha_atencion_formato)
            if fecha_atencion_formato:
                print(f"      [+] Fecha Atencion: {fecha_atencion_formato}")
        except:
            pass
        time.sleep(0.3)

        # Presenta discapacidad
        try:
            select_discapacidad = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlPresentaDiscapacidad")
            Select(select_discapacidad).select_by_visible_text("No")
            print("      [+] Discapacidad: No")
        except:
            pass
        time.sleep(0.3)

        # Foto del beneficiario
        if not cargar_foto_beneficiario(datos['documento']):
            print("      [!] No se pudo cargar la foto")
        
        print(f"    [+] Datos ingresados")
        return True
    except Exception as e:
        print(f"    [-] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def buscar_foto_por_documento(documento):
    coincidencias = sorted(ruta_fotos_d2.glob(f"*{documento}*"))
    if coincidencias:
        return str(coincidencias[0])
    return None


def cargar_foto_beneficiario(documento):
    """Carga la foto del beneficiario antes de guardar."""
    print("      [4] Cargando foto...")

    ruta_foto = buscar_foto_por_documento(documento)
    if not ruta_foto:
        print(f"      [-] No se encontró foto para el documento {documento}")
        return False

    try:
        input_foto = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_FileUploadControl")))
        driver.execute_script("arguments[0].scrollIntoView(true);", input_foto)
        time.sleep(0.5)
        input_foto.send_keys(ruta_foto)
        print(f"      [+] Foto seleccionada: {Path(ruta_foto).name}")
        time.sleep(1)

        boton_cargar = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnCargar")))
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_cargar)
        time.sleep(0.5)
        boton_cargar.click()
        print("      [+] Botón 'Cargar foto' pulsado")
        print("      [*] Esperando 8 segundos a que termine la carga de la foto...")
        time.sleep(8)
        return True
    except Exception as e:
        print(f"      [-] Error cargando foto: {str(e)}")
        return False

def guardar_registro():
    """Guarda el registro"""
    print("    [*] Guardando...")
    
    try:
        selectores_guardar = [
            (By.ID, "btnGuardar"),
            (By.XPATH, "//a[@id='btnGuardar']"),
            (By.XPATH, "//img[@title='Guardar']/parent::a"),
            (By.XPATH, "//img[@alt='Guardar']/parent::a"),
        ]

        for selector in selectores_guardar:
            try:
                boton_guardar = WebDriverWait(driver, 60).until(EC.presence_of_element_located((selector[0], selector[1])))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_guardar)
                print("    [*] Esperando que el disquete quede habilitado...")
                time.sleep(2)
                try:
                    boton_guardar.click()
                except Exception:
                    try:
                        ActionChains(driver).move_to_element(boton_guardar).click().perform()
                    except Exception:
                        driver.execute_script("arguments[0].click();", boton_guardar)
                print("    [+] Guardado")
                print("    [*] Esperando 10 segundos para verificar el guardado...")
                time.sleep(10)
                return True
            except Exception:
                pass

        print("    [-] No se encontró el botón guardar (disquete)")
        return False
    except Exception as e:
        print(f"    [-] Error: {str(e)}")
        return False

# ========== MAIN PROCESS ==========
try:
    # LOGIN INICIAL
    print("\n[*] Paso 1: Accediendo y haciendo login...")
    driver.get("https://rubonline.icbf.gov.co/DefaultF.aspx")
    time.sleep(3)
    
    campo_usuario = wait.until(EC.presence_of_element_located((By.ID, "UserName")))
    campo_usuario.clear()
    campo_usuario.send_keys("Usuario")
    
    campo_password = driver.find_element(By.ID, "Password")
    campo_password.clear()
    campo_password.send_keys("Contraseña")
    
    boton_login = driver.find_element(By.ID, "LoginButton")
    boton_login.click()
    time.sleep(8)
    print("[+] Login completado\n")
    
    # PROCESAR SOLO EL PRIMER REGISTRO
    for idx, datos in enumerate(registros_d2[:1], 1):  # Solo el primero
        print(f"\n{'='*70}")
        print(f"PROCESANDO REGISTRO {idx}/{len(registros_d2)} (SOLO PRUEBA)")
        print(f"{'='*70}")
        print(f"Documento: {datos['documento']}")
        print(f"Nombre: {datos['primer_nombre']} {datos['segundo_nombre']} {datos['primer_apellido']} {datos['segundo_apellido']}")
        
        try:
            # Navegar
            if not navegar_a_formulario():
                registros_error.append({'idx': idx, 'error': 'Navegación fallida'})
                continue
            
            # Clickear +
            if not clickear_boton_nuevo():
                registros_error.append({'idx': idx, 'error': 'No pudo hacer click en +'})
                continue
            
            # Llenar filtros
            if not llenar_filtros_fijos():
                registros_error.append({'idx': idx, 'error': 'Error en filtros fijos'})
                continue
            
            # Llenar datos
            if not llenar_datos_beneficiario(datos):
                registros_error.append({'idx': idx, 'error': 'Error llenando datos'})
                continue
            
            # Guardar
            if not guardar_registro():
                registros_error.append({'idx': idx, 'error': 'Error al guardar'})
                continue
            
            registros_exitosos += 1
            print(f"\n[✓] REGISTRO {idx} COMPLETADO\n")
            
            # Pausa entre registros
            if idx < len(registros_d2):
                time.sleep(2)
        
        except Exception as e:
            print(f"\n[-] Error procesando registro {idx}: {str(e)}")
            registros_error.append({'idx': idx, 'error': str(e)})
        
        finally:
            registros_procesados = idx
            try:
                driver.switch_to.default_content()
            except:
                pass
    
    # RESUMEN FINAL
    print(f"\n\n{'='*70}")
    print("RESUMEN DE PROCESAMIENTO")
    print(f"{'='*70}")
    print(f"Total registros a procesar: {len(registros_d2)}")
    print(f"Registros exitosos: {registros_exitosos}")
    print(f"Registros con error: {len(registros_error)}")
    
    if registros_error:
        print("\nRegistros con problemas:")
        for err in registros_error[:5]:
            print(f"  - Registro {err['idx']}: {err['error']}")
        if len(registros_error) > 5:
            print(f"  ... y {len(registros_error)-5} más")
    
    print(f"\n{'='*70}\n")
    
    print("Navegador abierto - Presiona Ctrl+C para cerrar")
    while True:
        time.sleep(1)

except KeyboardInterrupt:
    print("\n\n[!] Cerrando navegador...")
    driver.quit()
    print("[+] Cerrado")

except Exception as e:
    print(f"\n[-] Error: {str(e)}")
    import traceback
    traceback.print_exc()
    driver.quit()
