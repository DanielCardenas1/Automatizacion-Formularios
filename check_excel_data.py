#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path('/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA')
RUTA_EXCEL = BASE_DIR / "DUITAMA E/CARGUE MASIVO DUITAMA E 2026.xlsx"

print(f"Excel file: {RUTA_EXCEL}\n")

if RUTA_EXCEL.exists():
    wb = load_workbook(RUTA_EXCEL, data_only=True)
    ws = wb.active
    
    print(f"Total rows: {ws.max_row}")
    print(f"\nFirst row headers (cols 1-15):")
    for col in range(1, 16):
        header = ws.cell(row=1, column=col).value
        print(f"  Col {col}: {header}")
    
    print(f"\nSearching all rows for doc 1052848446...")
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, 20):
            cell_val = str(ws.cell(row=row_idx, column=col).value or "").strip()
            if '1052848446' in cell_val:
                print(f"\nFOUND at row {row_idx}, col {col}")
                print(f"Headers and values for this row:")
                for c in range(1, 15):
                    header = ws.cell(row=1, column=c).value or f"Col{c}"
                    val = ws.cell(row=row_idx, column=c).value
                    print(f"  {header}: {val}")
                exit(0)
    
    print("Document not found in any column")
else:
    print("Excel file not found!")
