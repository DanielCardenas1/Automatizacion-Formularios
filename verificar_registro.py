#!/usr/bin/env python3
import openpyxl

# Ruta del archivo
ruta = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"

print("\n[*] Buscando documento: 1145330865 en TODAS las filas...\n")

# Abrir el workbook
wb = openpyxl.load_workbook(ruta)
ws = wb.active

documento_buscado = "1145330865"
encontrado = False
coincidencias = []

# Buscar en TODOS los datos (columna 17 es NUMERO DE DOCUMENTO)
for row_idx in range(3, ws.max_row + 1):
    valor = ws.cell(row=row_idx, column=17).value
    if valor and str(valor).strip() == documento_buscado:
        coincidencias.append(row_idx)

print(f"[+] Encontradas {len(coincidencias)} coincidencias\n")

if coincidencias:
    for fila in coincidencias[:3]:  # Mostrar primeras 3
        print(f"\n{'='*70}")
        print(f"REGISTRO ENCONTRADO EN FILA {fila}")
        print(f"{'='*70}\n")
        
        # Mostrar encabezados cortos
        campos_importante = {
            1: "CONTRATO",
            2: "MUNICIPIO", 
            3: "NOMBRE UDS",
            4: "FECHA INGRESO",
            5: "PRIMER NOMBRE",
            6: "SEGUNDO NOMBRE",
            7: "PRIMER APELLIDO",
            8: "SEGUNDO APELLIDO",
            9: "SEXO",
            13: "FECHA NACIMIENTO",
            16: "TIPO DOC",
            17: "NUMERO DOC",
            19: "EPS"
        }
        
        print("[DATOS DEL EXCEL]:\n")
        for col, etiqueta in campos_importante.items():
            valor = ws.cell(row=fila, column=col).value
            print(f"  {etiqueta}: {valor}")
        
        print("\n" + "="*70)
        print("VERIFICACIÓN vs FORMULARIO")
        print("="*70 + "\n")
        
        print("[✓/✗] COMPARACIÓN:\n")
        
        # Datos del formulario (de las imágenes)
        datos_form = {
            "NUMERO DOC": ("1145330865", ws.cell(row=fila, column=17).value),
            "PRIMER NOMBRE": ("ALAN", ws.cell(row=fila, column=5).value),
            "SEGUNDO NOMBRE": ("", ws.cell(row=fila, column=6).value),
            "PRIMER APELLIDO": ("GARCIA", ws.cell(row=fila, column=7).value),
            "SEGUNDO APELLIDO": ("RIVERA", ws.cell(row=fila, column=8).value),
            "SEXO": ("Hombre", ws.cell(row=fila, column=9).value),
            "FECHA NACIMIENTO": ("05/08/2025", ws.cell(row=fila, column=13).value),
            "FECHA INGRESO": ("09/02/2026", ws.cell(row=fila, column=4).value),
            "TIPO DOC": ("REGISTRO CIVIL", ws.cell(row=fila, column=16).value),
        }
        
        for campo, (valor_form, valor_excel) in datos_form.items():
            # Verificar coincidencia
            if valor_form.lower() == str(valor_excel).lower() or valor_form == "" or valor_form == "Hombre" and str(valor_excel).lower() == "m":
                marca = "✓"
            else:
                marca = "✗"
            
            print(f"{marca} {campo}:")
            print(f"    FORMULARIO: {valor_form}")
            print(f"    EXCEL:      {valor_excel}")
            print()

else:
    print("[-] Documento NO encontrado en todo el archivo")

wb.close()
