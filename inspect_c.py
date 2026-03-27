#!/usr/bin/env python3
from openpyxl import load_workbook

excel_file = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/CARGUE MASIVO 2026 _ DUITAMA C_.xlsx"
print(f"Excel file: {excel_file}\n")

wb = load_workbook(excel_file, data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}\n")
print("Row 1-2 headers (cols 1-20):")
for c in range(1, 21):
    h1 = ws.cell(1, c).value
    h2 = ws.cell(2, c).value
    print(f"  Col {c}: R1={h1} | R2={h2}")

print("\nFirst 5 data rows (cols 1-15):")
for r in range(3, min(8, ws.max_row + 1)):
    vals = [ws.cell(r, c).value for c in range(1, 16)]
    print(f"  Row {r}: {vals}")

def count_col(col, token):
    n = 0
    for r in range(3, ws.max_row + 1):
        v = str(ws.cell(r, col).value or '').upper()
        if token in v:
            n += 1
    return n

print("\nFilter counts (C1, C2, C3) by column:")
for col in range(1, 10):
    print(f"  Col {col}: C1={count_col(col, 'C1')} C2={count_col(col, 'C2')} C3={count_col(col, 'C3')}")
