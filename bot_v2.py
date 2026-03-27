#!/usr/bin/env python3
"""
Bot Selenium - Carga Masiva de DUITAMA D2 (VERSIÓN MEJORADA)
Lee todos los registros de DUITAMA D2 del Excel
Hace clic correctamente en el campo de documento, luego en la lupa
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
from datetime import datetime

print("\n" + "="*70)
print("BOT CARGA MASIVA - DUITAMA D2 (VERSIÓN MEJORADA)")
print("="*70)

# ========== LEER EXCEL ==========
print("\n[*] Paso 0: Leyendo registros de DUITAMA D2...")

ruta_excel = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

registros_d2 = []

for row_idx in range(3, ws.max_row + 1):
    nombre_uds = ws.cell(row=row_idx, column=3).value
    
    if nombre_uds and "D2" in str(nombre_uds).upper() and "D3" not in str(nombre_uds).upper():
        datos = {
            'fila': row_idx,
            'documento': str(int(ws.cell(row=row_idx, column=17).value)) if ws.cell(row=row_idx, column=17).value else "",
            'primer_nombre': str(ws.cell(row=row_idx, column=5).value).strip() if ws.cell(row=row_idx, column=5).value else "",
            'segundo_nombre': str(ws.cell(row=row_idx, column=6).value).strip() if ws.cell(row=row_idx, column=6).value else "",
            'primer_apellido': str(ws.cell(row=row_idx, column=7).value).strip() if ws.cell(row=row_idx, column=7).value else "",
            'segundo_apellido': str(ws.cell(row=row_idx, column=8).value).strip() if ws.cell(row=row_idx, column=8).value else "",
            'sexo': str(ws.cell(row=row_idx, column=9).value).strip() if ws.cell(row=row_idx, column=9).value else "",
            'fecha_nac': ws.cell(row=row_idx, column=13).value,
            'fecha_ingreso': ws.cell(row=row_idx, column=4).value,
            'tipo_doc': str(ws.cell(row=row_idx, column=16).value).strip() if ws.cell(row=row_idx, column=16).value else "",
        }
        registros_d2.append(datos)

wb.close()

print(f"[+] Total registros D2 encontrados: {len(registros_d2)}")
for i, reg in enumerate(registros_d2[:1], 1):
    print(f"    {i}. {reg['documento']} - {reg['primer_nombre']} {reg['primer_apellido']}")

# ========== CONFIGURAR SELENIUM ==========
options = webdriver.ChromeOptions()
options.add_argument('--disable-notifications')
options.add_argument('--disable-popup-blocking')

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 20)

registros_exitosos = 0
registros_error = []

try:
    # LOGIN
    print("\n[*] Paso 1: Login...")
    driver.get("https://rubonline.icbf.gov.co/DefaultF.aspx")
    time.sleep(3)
    
    campo_usuario = wait.until(EC.presence_of_element_located((By.ID, "UserName")))
    campo_usuario.send_keys("angie.cardenas")
    driver.find_element(By.ID, "Password").send_keys("Celeste1020*")
    driver.find_element(By.ID, "LoginButton").click()
    time.sleep(8)
    print("[+] Login completado")
    
    # PROCESAR SOLO PRIMER REGISTRO
    for idx, datos in enumerate(registros_d2[:1], 1):
        print(f"\n{'='*70}")
        print(f"PROCESANDO REGISTRO {idx}/1")
        print(f"{'='*70}")
        print(f"Documento: {datos['documento']}")
        print(f"Nombre: {datos['primer_nombre']} {datos['primer_apellido']}")
        
        try:
            # Navegar
            print("\n  [NAVEGACIÓN]")
            time.sleep(2)
            driver.find_elements(By.XPATH, "//a[contains(text(), 'Rub online')]")[0].click()
            time.sleep(3)
            
            driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")[0].click()
            time.sleep(3)
            
            driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")[1].click()
            time.sleep(4)
            
            print("  [+] Navegación completada")
            
            # Click en botón +
            print("  [NUEVO BENEFICIARIO]")
            time.sleep(2)
            iframe = wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
            driver.switch_to.frame(iframe)
            
            boton_nuevo = wait.until(EC.element_to_be_clickable((By.ID, "btnNuevo")))
            ActionChains(driver).move_to_element(boton_nuevo).click().perform()
            print("  [+] Botón '+' clickeado")
            time.sleep(3)
            
            # Llenar filtros
            print("  [FILTROS]")
            try:
                # Uno a uno
                radioBtns = driver.find_elements(By.XPATH, "//input[@type='radio']")
                for radio in radioBtns:
                    if 'Uno a uno' in radio.get_attribute('value'):
                        radio.click()
                        break
                time.sleep(1)
                
                # Dirección
                Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDireccionesICBF")).select_by_visible_text("Dirección de Primera Infancia")
                time.sleep(1)
                
                # Regional
                Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlRegional")).select_by_visible_text("Boyacá")
                time.sleep(1)
                
                # Vigencia
                Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdVigencia")).select_by_visible_text("2026")
                time.sleep(1)
                
                # Contrato
                Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNumeroContrato")).select_by_visible_text("OD 15 420272 00015 2026")
                time.sleep(1)
                
                # Servicio
                Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNombreServicio")).select_by_visible_text("EDUCACIÓN INICIAL EN EL HOGAR - FAMILIAR Y COMUNITARIA - 420272-2026")
                time.sleep(2)
                
                # UDS
                selects = driver.find_elements(By.TAG_NAME, "select")
                for select_elem in selects:
                    for opt in select_elem.find_elements(By.TAG_NAME, "option"):
                        if "DUITAMA D2" in opt.text:
                            Select(select_elem).select_by_value(opt.get_attribute("value"))
                            break
                
                print("  [+] Filtros completados")
            except Exception as e:
                print(f"  [-] Error: {str(e)}")
            
            # AQUÍ INGRESARÁ LOS DATOS DEL BENEFICIARIO
            print("  [DATOS]")
            time.sleep(3)
            
            # Buscar campo de documento
            print(f"    [1] Buscando campo de documento...")
            campo_doc = None
            try:
                # Scroll al inicio
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.5)
                
                # Buscar el campo
                campo_doc = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdentificacion")
                print(f"    [+] Campo encontrado")
                
                # Hacer visible con JavaScript
                driver.execute_script("""
                    var elem = arguments[0];
                    elem.style.display = 'block';
                    elem.style.visibility = 'visible';
                    elem.style.opacity = '1';
                    elem.scrollIntoView({behavior: 'smooth', block: 'center'});
                """, campo_doc)
                time.sleep(1)
                
                # Hacer clic
                print(f"    [*] Haciendo clic...")
                driver.execute_script("arguments[0].focus();", campo_doc)
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", campo_doc)
                time.sleep(0.5)
                
                # Escribir
                print(f"    [*] Escribiendo: {datos['documento']}")
                campo_doc.send_keys(datos['documento'])
                print(f"    [+] Documento ingresado")
                time.sleep(1)
                
                # Buscar lupa
                print(f"    [2] Buscando lupa...")
                lupa = None
                
                # Buscar primero por ID
                try:
                    lupa = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnBuscar")
                    driver.execute_script("arguments[0].scrollIntoView(true);", lupa)
                    time.sleep(0.3)
                    lupa.click()
                    print(f"    [+] Click en lupa realizado")
                    time.sleep(3)
                except:
                    # Buscar TODO tipo de botones cercanos
                    todos_botones = driver.find_elements(By.XPATH, "//input[@type='button'] | //button")
                    for btn in todos_botones:
                        if 'buscar' in btn.get_attribute('value').lower() if btn.get_attribute('value') else False or 'buscar' in btn.text.lower():
                            lupa = btn
                            break
                except:
                    pass
                
                if lupa:
                    print(f"    [+] Lupa encontrada")
                    driver.execute_script("""
                        var elem = arguments[0];
                        elem.style.display = 'block';
                        elem.style.visibility = 'visible';
                        elem.style.opacity = '1';
                        elem.scrollIntoView({behavior: 'smooth', block: 'center'});
                    """, lupa)
                    time.sleep(0.5)
                    driver.execute_script("arguments[0].click();", lupa)
                    print(f"    [+] Click en lupa realizado")
                    time.sleep(5)
                else:
                    print(f"    [-] Lupa no encontrada")
                    
            except Exception as e:
                print(f"    [-] Error: {str(e)}")
                import traceback
                traceback.print_exc()
            
            registros_exitosos += 1
            print(f"\n[✓] REGISTRO PROCESADO")
            
        except Exception as e:
            print(f"\n[-] Error: {str(e)}")
            registros_error.append({'idx': idx, 'error': str(e)})
        
        finally:
            try:
                driver.switch_to.default_content()
            except:
                pass
    
    # RESUMEN
    print(f"\n\n{'='*70}")
    print("RESUMEN")
    print(f"{'='*70}")
    print(f"Exitosos: {registros_exitosos}")
    print(f"Errores: {len(registros_error)}")
    
    print("\nNavegador abierto - Presiona Ctrl+C para cerrar")
    while True:
        time.sleep(1)

except KeyboardInterrupt:
    print("\n[!] Cerrando...")
    driver.quit()

except Exception as e:
    print(f"\n[-] Error: {str(e)}")
    driver.quit()
