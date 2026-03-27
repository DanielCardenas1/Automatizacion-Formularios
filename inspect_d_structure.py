#!/usr/bin/env python3

import openpyxl
from pathlib import Path

BASE_DIR = Path("/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA")
# Find D Excel file
import glob
d_files = glob.glob(str(BASE_DIR / "DUITAMA D" / "*DUITAMA D*.xlsx"))
if d_files:
    excel_path = d_files[0]
else:
    print("No D Excel file found")
    exit(1)

print(f"D Excel: {excel_path}")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Get first row D1 data
print("HEADERS (Row 1-2):")
for col in range(1, 30):
    h1 = ws.cell(row=1, column=col).value
    h2 = ws.cell(row=2, column=col).value
    header = h2 or h1
    print(f"  Col {col}: {header}")

print("\n" + "="*80)
print("FIRST D RECORD (Row 3):")
for col in range(1, 30):
    cell_value = ws.cell(row=3, column=col).value
    print(f"  Col {col}: {cell_value}")
