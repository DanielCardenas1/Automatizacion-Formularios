import openpyxl
wb = openpyxl.load_workbook('DUITAMA F/CARGUE MASIVO 2026_DUITAMA F_ACTUALIZADO.xlsx')
sheet = wb.active
f2_rows = []
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        val = str(sheet.cell(row, col).value or '').strip().upper()
        if val == 'F2':
            f2_rows.append(row)
            break
print(f'Total filas con F2: {len(f2_rows)}')
if f2_rows:
    # Encabezados
    headers = [str(sheet.cell(1, col).value or '').strip() for col in range(1, sheet.max_column + 1)]
    print('Encabezados relevantes:')
    for i, h in enumerate(headers):
        if h and ('DOCUMENTO' in h.upper() or 'NOMBRE' in h.upper() or 'UDS' in h.upper()):
            print(f'  Col {i+1}: {h}')
    # Primera fila F2
    row = f2_rows[0]
    values = [str(sheet.cell(row, col).value or '').strip() for col in range(1, sheet.max_column + 1)]
    print(f'Primera fila F2 (fila {row}):')
    for i, (h, v) in enumerate(zip(headers, values)):
        if h and v:
            print(f'  {h}: {v}')