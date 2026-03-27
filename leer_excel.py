#!/usr/bin/env python3
import openpyxl
import os

# Ruta del archivo
ruta = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"

print("\n" + "="*70)
print("INFORMACIÓN DEL ARCHIVO EXCEL")
print("="*70)

# Verificar que existe
if not os.path.exists(ruta):
    print(f"[-] Archivo no encontrado: {ruta}")
    exit(1)

print(f"\n[+] Archivo encontrado")
print(f"[+] Tamaño: {os.path.getsize(ruta)} bytes\n")

# Abrir el workbook
try:
    wb = openpyxl.load_workbook(ruta)
    print(f"[+] Workbook cargado exitosamente")
    
    # Listar hojas
    print(f"\n[*] Hojas en el workbook: {len(wb.sheetnames)}")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. '{sheet_name}' - {ws.max_row} filas x {ws.max_column} columnas")
    
    # Leer la primera hoja
    ws = wb.active
    print(f"\n[*] Leyendo hoja activa: '{ws.title}'")
    print(f"[+] Filas: {ws.max_row}, Columnas: {ws.max_column}\n")
    
    # Mostrar encabezados
    print("="*70)
    print("ENCABEZADOS (Primera fila)")
    print("="*70)
    encabezados = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        encabezados.append(header)
        print(f"  Col {col}: {header}")
    
    # Mostrar primeras filas de datos
    print("\n" + "="*70)
    print("PRIMEROS 10 REGISTROS")
    print("="*70 + "\n")
    
    for row_idx in range(2, min(12, ws.max_row + 1)):
        print(f"\n[Fila {row_idx}]")
        for col_idx, header in enumerate(encabezados, 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor is not None:
                print(f"  {header}: {valor}")
    
    # Resumen de datos
    print("\n" + "="*70)
    print("RESUMEN")
    print("="*70)
    print(f"[+] Total de registros: {ws.max_row - 1}")
    print(f"[+] Total de campos: {ws.max_column}")
    
    wb.close()
    
except Exception as e:
    print(f"[-] Error al abrir el archivo: {str(e)}")
    import traceback
    traceback.print_exc()
