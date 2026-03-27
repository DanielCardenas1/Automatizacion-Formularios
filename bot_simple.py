#!/usr/bin/env python3
"""
Bot SIMPLIFICADO - Solo llena documento y datos del beneficiario
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time

print("\n" + "="*70)
print("BOT SIMPLIFICADO - ALAN GARCIA RIVERA")
print("="*70 + "\n")

# Leer Excel
ruta_excel = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

# Fila 18 = ALAN GARCIA RIVERA
fila = 18
datos = {
    'documento': str(int(ws.cell(row=fila, column=17).value)),
    'primer_nombre': str(ws.cell(row=fila, column=5).value).strip(),
    'segundo_nombre': str(ws.cell(row=fila, column=6).value).strip(),
    'primer_apellido': str(ws.cell(row=fila, column=7).value).strip(),
    'segundo_apellido': str(ws.cell(row=fila, column=8).value).strip(),
    'sexo': str(ws.cell(row=fila, column=9).value).strip(),
    'fecha_nac': ws.cell(row=fila, column=13).value,
    'fecha_ingreso': ws.cell(row=fila, column=4).value,
}
wb.close()

print(f"Documento: {datos['documento']}")
print(f"Nombre: {datos['primer_nombre']} {datos['segundo_nombre']} {datos['primer_apellido']} {datos['segundo_apellido']}")
print(f"Sexo: {datos['sexo']}\n")

options = webdriver.ChromeOptions()
options.add_argument('--disable-notifications')
options.add_argument('--disable-popup-blocking')

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 20)

try:
    # LOGIN
    print("[1] Login...")
    driver.get("https://rubonline.icbf.gov.co/DefaultF.aspx")
    time.sleep(3)
    
    driver.find_element(By.ID, "UserName").send_keys("angie.cardenas")
    driver.find_element(By.ID, "Password").send_keys("Celeste1020*")
    driver.find_element(By.ID, "LoginButton").click()
    time.sleep(8)
    print("[+] Login completado\n")
    
    # NAVEGAR
    print("[2] Navegación...")
    time.sleep(2)
    driver.find_elements(By.XPATH, "//a[contains(text(), 'Rub online')]")[0].click()
    time.sleep(3)
    driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")[0].click()
    time.sleep(3)
    driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")[1].click()
    time.sleep(4)
    print("[+] Navegación completada\n")
    
    # CLICK EN +
    print("[3] Click en botón +...")
    iframe = wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
    driver.switch_to.frame(iframe)
    
    boton_nuevo = wait.until(EC.element_to_be_clickable((By.ID, "btnNuevo")))
    ActionChains(driver).move_to_element(boton_nuevo).click().perform()
    print("[+] Botón + clickeado\n")
    time.sleep(4)
    
    # LLENAR FILTROS (SIN ITERAR PARA EVITAR STALE ELEMENTS)
    print("[4] Llenando filtros...")
    try:
        # Uno a uno - buscar el radio directamente sin iterar
        driver.find_element(By.XPATH, "//input[@type='radio' and contains(@value, 'Uno a uno')]").click()
        print("  [+] Uno a uno")
        time.sleep(1)
    except:
        print("  [-] No se pudo seleccionar 'Uno a uno'")
    
    try:
        Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDireccionesICBF")).select_by_visible_text("Dirección de Primera Infancia")
        print("  [+] Dirección")
        time.sleep(1)
    except Exception as e:
        print(f"  [-] Dirección: {str(e)}")
    
    try:
        Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlRegional")).select_by_visible_text("Boyacá")
        print("  [+] Regional")
        time.sleep(1)
    except Exception as e:
        print(f"  [-] Regional: {str(e)}")
    
    try:
        Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdVigencia")).select_by_visible_text("2026")
        print("  [+] Vigencia")
        time.sleep(1)
    except Exception as e:
        print(f"  [-] Vigencia: {str(e)}")
    
    try:
        Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNumeroContrato")).select_by_visible_text("OD 15 420272 00015 2026")
        print("  [+] Contrato")
        time.sleep(1)
    except Exception as e:
        print(f"  [-] Contrato: {str(e)}")
    
    try:
        Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNombreServicio")).select_by_visible_text("EDUCACIÓN INICIAL EN EL HOGAR - FAMILIAR Y COMUNITARIA - 420272-2026")
        print("  [+] Servicio")
        time.sleep(2)
    except Exception as e:
        print(f"  [-] Servicio: {str(e)}")
    
    # UDS - buscar DUITAMA D2
    try:
        all_options = driver.find_elements(By.XPATH, "//option[contains(text(), 'DUITAMA D2')]")
        if all_options:
            option = all_options[0]
            valor = option.get_attribute("value")
            select_elem = option.find_element(By.XPATH, "..")
            Select(select_elem).select_by_value(valor)
            print("  [+] UDS")
            time.sleep(1)
    except Exception as e:
        print(f"  [-] UDS: {str(e)}")
    
    # Tipo de beneficiario - NIÑO O NIÑA ENTRE 6 MESES
    try:
        driver.execute_script("window.scrollBy(0, 300);")
        time.sleep(0.5)
        select_beneficiario = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdTipoBeneficiario")
        opciones = select_beneficiario.find_elements(By.TAG_NAME, "option")
        for opt in opciones:
            if "NIÑO O NIÑA ENTRE 6 MESES" in opt.text:
                Select(select_beneficiario).select_by_value(opt.get_attribute("value"))
                break
        print("  [+] Tipo beneficiario")
        time.sleep(1)
    except Exception as e:
        print(f"  [-] Tipo beneficiario: {str(e)}")
    
    # Tipo de Documento - REGISTRO CIVIL
    try:
        driver.execute_script("window.scrollBy(0, 100);")
        time.sleep(0.5)
        select_tipo_doc = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdTipoDocumento")
        Select(select_tipo_doc).select_by_visible_text("REGISTRO CIVIL")
        print("  [+] Tipo documento")
        time.sleep(1)
    except Exception as e:
        print(f"  [-] Tipo documento: {str(e)}")
    
    print()
    
    # LLENAR DATOS DEL BENEFICIARIO
    print("[5] Datos beneficiario...")
    time.sleep(2)
    
    # Scroll al inicio
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)
    
    # Documento - BUSCAR Y HACER CLICK DIRECTO
    print("\n  [DOCUMENTO]")
    try:
        # Usar XPath más flexible
        campo_doc = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdentificacion")
        print(f"    [+] Campo encontrado")
        driver.execute_script("arguments[0].scrollIntoView(true);", campo_doc)
        time.sleep(0.5)
        campo_doc.click()
        time.sleep(0.3)
        campo_doc.send_keys(datos['documento'])
        print(f"    [+] Documento ingresado: {datos['documento']}")
        time.sleep(0.5)
        
        # BUSCAR LUPA
        print(f"    [LUPA]")
        try:
            lupa = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnBuscar")
            lupa.click()
            print(f"    [+] Click en lupa realizado")
            time.sleep(3)
        except:
            botones_all = driver.find_elements(By.TAG_NAME, "button")
            print(f"    Total botones: {len(botones_all)}")
        
        # Búsqueda simple de la lupa
        for i, btn in enumerate(botones_all):
            btn_text = btn.text.strip().lower()
            btn_class = btn.get_attribute("class") or ""
            print(f"    Botón {i}: '{btn_text}' class='{btn_class}'")
            
            if 'buscar' in btn_text or 'search' in btn_class.lower():
                print(f"    [+] Lupa encontrada en botón {i}")
                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                time.sleep(0.5)
                btn.click()
                print(f"    [+] Click en lupa realizado")
                time.sleep(4)
                break
        
    except Exception as e:
        print(f"    [-] Error: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print("\n[✓] PROCESO COMPLETADO")
    print("\nNavegador abierto - Presiona Ctrl+C para cerrar")
    while True:
        time.sleep(1)

except KeyboardInterrupt:
    print("\n[!] Cerrando...")
    driver.quit()

except Exception as e:
    print(f"\n[-] Error: {str(e)}")
    import traceback
    traceback.print_exc()
    driver.quit()
