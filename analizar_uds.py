#!/usr/bin/env python3
import openpyxl
from collections import Counter

ruta = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"
wb = openpyxl.load_workbook(ruta)
ws = wb.active

print("\n[*] Analizando distribución de UDS en el archivo...\n")

# Columna 3 = NOMBRE UDS
uds_list = []
d2_registros = []

for row_idx in range(3, ws.max_row + 1):
    nombre_uds = ws.cell(row=row_idx, column=3).value
    if nombre_uds:
        uds_list.append(str(nombre_uds).strip())
        
        # Si es D2 (pueden ser DUITAMA D2, D2, etc)
        if "D2" in str(nombre_uds).upper() and "D3" not in str(nombre_uds).upper():
            doc = ws.cell(row=row_idx, column=17).value
            nombre = ws.cell(row=row_idx, column=5).value
            apellido = ws.cell(row=row_idx, column=7).value
            d2_registros.append({
                'fila': row_idx,
                'uds': nombre_uds,
                'documento': doc,
                'nombre': nombre,
                'apellido': apellido
            })

# Contar UDS
contador_uds = Counter(uds_list)

print("[+] UDS encontradas en el archivo:")
for uds, cantidad in sorted(contador_uds.items(), key=lambda x: x[1], reverse=True):
    print(f"    {uds}: {cantidad} registros")

print(f"\n[+] Registros de DUITAMA D2 (sin D3): {len(d2_registros)}")
if d2_registros:
    print("\n[PRIMEROS 10 REGISTROS D2]:")
    for i, reg in enumerate(d2_registros[:10], 1):
        print(f"  {i}. Fila {reg['fila']}: {reg['documento']} - {reg['nombre']} {reg['apellido']} ({reg['uds']})")
    
    if len(d2_registros) > 10:
        print(f"\n  ... y {len(d2_registros) - 10} más")

wb.close()
