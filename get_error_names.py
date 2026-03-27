#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA G/CARGUE MASIVO_DUITAMA_G1_G2_G3_2026.xlsm', data_only=True)
ws = wb.active

docs_error = {
    '1052419916': 'G1',
    '1052420240': 'G2', 
    '1052420246': 'G3'
}

print("BENEFICIARIOS CON ERROR:\n")
for row in range(5, ws.max_row + 1):
    doc_val = ws.cell(row, 19).value
    if doc_val:
        doc = str(doc_val).strip().split('.')[0]  # Remove decimal if float
        if doc in docs_error:
            primer_nombre = ws.cell(row, 7).value or ''
            segundo_nombre = ws.cell(row, 8).value or ''
            apellido1 = ws.cell(row, 9).value or ''
            apellido2 = ws.cell(row, 10).value or ''
            full_name = f"{primer_nombre} {segundo_nombre} {apellido1} {apellido2}".replace('  ', ' ').strip()
            print(f"{docs_error[doc]} | {doc} | {full_name}")
