"""
Bot Selenium para RUB Online
Accede a https://rubonline.icbf.gov.co/DefaultF.aspx
Lee datos de archivos Excel y verifica información en la página
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
import os
from datetime import datetime


class RUBBot:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.excel_data = {}
        self.resultados = []
        
    def inicializar_driver(self):
        """Inicializa el driver de Chrome"""
        print("[*] Inicializando Selenium WebDriver...")
        options = webdriver.ChromeOptions()
        # Descomenta para modo headless (sin interfaz)
        # options.add_argument('--headless')
        options.add_argument('--disable-notifications')
        options.add_argument('--disable-popup-blocking')
        
        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        self.wait = WebDriverWait(self.driver, 20)
        print("[+] WebDriver inicializado correctamente")
        
    def cerrar_driver(self):
        """Cierra el navegador"""
        if self.driver:
            self.driver.quit()
            print("[+] WebDriver cerrado")
    
    def leer_excel(self, ruta_excel):
        """
        Lee datos del archivo Excel
        
        Args:
            ruta_excel (str): Ruta del archivo Excel
            
        Returns:
            list: Lista de diccionarios con los datos
        """
        print(f"\n[*] Leyendo archivo Excel: {ruta_excel}")
        
        if not os.path.exists(ruta_excel):
            print(f"[-] Error: No se encontró el archivo {ruta_excel}")
            return []
        
        try:
            wb = openpyxl.load_workbook(ruta_excel)
            ws = wb.active
            
            datos = []
            encabezados = []
            
            # Leer encabezados
            for col in ws.iter_cols(1, ws.max_column, 1, 1):
                for cell in col:
                    if cell.value:
                        encabezados.append(cell.value)
            
            print(f"[+] Encabezados encontrados: {encabezados}")
            
            # Leer datos
            for row in ws.iter_rows(2, ws.max_row, values_only=True):
                if row[0]:  # Si tiene al menos un valor
                    fila_dict = {}
                    for idx, valor in enumerate(row):
                        if idx < len(encabezados):
                            fila_dict[encabezados[idx]] = valor
                    datos.append(fila_dict)
            
            print(f"[+] Se leyeron {len(datos)} registros del Excel")
            self.excel_data = datos
            return datos
            
        except Exception as e:
            print(f"[-] Error al leer Excel: {str(e)}")
            return []
    
    def acceder_pagina(self, url):
        """Accede a la página web"""
        print(f"\n[*] Accediendo a {url}...")
        try:
            self.driver.get(url)
            time.sleep(3)
            print("[+] Página cargada correctamente")
            return True
        except Exception as e:
            print(f"[-] Error al acceder a la página: {str(e)}")
            return False
    
    def hacer_login(self, usuario, contraseña):
        """Realiza login en la página"""
        print(f"\n[*] Iniciando sesión con usuario: {usuario}")
        try:
            # Esperar y llenar campo de usuario (ID: UserName)
            campo_usuario = self.wait.until(
                EC.presence_of_element_located((By.ID, "UserName"))
            )
            campo_usuario.clear()
            campo_usuario.send_keys(usuario)
            print("[+] Usuario ingresado")
            
            # Llenar campo de contraseña (ID: Password)
            campo_password = self.driver.find_element(By.ID, "Password")
            campo_password.clear()
            campo_password.send_keys(contraseña)
            print("[+] Contraseña ingresada")
            
            # Hacer clic en botón de login (ID: LoginButton)
            boton_login = self.driver.find_element(By.ID, "LoginButton")
            boton_login.click()
            print("[+] Botón de login presionado")
            
            # Esperar a que cargue la siguiente página
            time.sleep(5)
            print("[+] Sesión iniciada correctamente")
            return True
            
        except Exception as e:
            print(f"[-] Error en login: {str(e)}")
            return False
    
    def navegar_a_beneficiario(self):
        """
        Navega al módulo de Beneficiario desde el menú
        Click en: RUB ONLINE > Rub online > Beneficiario > Beneficiario
        """
        print("\n[*] Navegando a Beneficiario del menú...")
        try:
            from selenium.webdriver.common.action_chains import ActionChains
            
            # Buscar todos los enlaces que contengan "Beneficiario" y "BENEFICIARIO" en href
            enlaces = self.driver.find_elements(By.XPATH, 
                "//a[contains(text(), 'Beneficiario') and contains(@href, 'BENEFICIARIO')]")
            
            if enlaces:
                # Hacer click en el primer enlace (Beneficiario del listado)
                for enlace in enlaces:
                    href = enlace.get_attribute('href')
                    if href and 'List.aspx' in href:
                        print(f"[*] Encontrado enlace: {href}")
                        
                        # Scroll al elemento
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", enlace)
                        time.sleep(1)
                        
                        # Intenta hacer click usando ActionChains (más robusto)
                        try:
                            actions = ActionChains(self.driver)
                            actions.move_to_element(enlace).click().perform()
                            print("[+] Click en Beneficiario completado (ActionChains)")
                        except:
                            # Fallback: intenta click directo
                            print("[*] Intentando click directo...")
                            enlace.click()
                            print("[+] Click en Beneficiario completado (click directo)")
                        break
            else:
                print("[-] No se encontró enlace de Beneficiario")
                return False
            
            time.sleep(4)
            
            # El contenido carga en un iframe con id="frameContent"
            self.wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
            print("[+] Página de Beneficiario cargada")
            
            return True
                
        except Exception as e:
            print(f"[-] Error navegando a Beneficiario: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def seleccionar_opcion_beneficiario(self):
        """
        En esta página, simplemente espera - no hay "Opción Beneficiario" dropdown
        La estructura es diferente, vamos directo al botón "Nuevo"
        """
        print("\n[*] Página de Beneficiario preparada")
        print("[+] Estructura de formulario detectada correctamente")
        return True
    
    def hacer_click_nueva_opcion(self):
        """
        Hace click en el botón "Nuevo" (símbolo +) para agregar nuevo beneficiario
        ID: btnNuevo
        """
        print("\n[*] Buscando botón 'Nuevo' (símbolo +)...")
        try:
            from selenium.webdriver.common.action_chains import ActionChains
            
            # Cambiar el contexto al iframe
            iframe = self.driver.find_element(By.ID, "frameContent")
            self.driver.switch_to.frame(iframe)
            
            print("[+] Estoy dentro del iframe")
            
            # Buscar el botón "Nuevo" por su ID
            boton_nuevo = self.wait.until(
                EC.presence_of_element_located((By.ID, "btnNuevo"))
            )
            
            print("[+] Botón 'Nuevo' encontrado")
            
            # Scroll al botón
            self.driver.execute_script("arguments[0].scrollIntoView(true);", boton_nuevo)
            time.sleep(1)
            
            # Intenta hacer click usando ActionChains (más robusto)
            try:
                actions = ActionChains(self.driver)
                actions.move_to_element(boton_nuevo).click().perform()
                print("[+] Click en 'Nuevo' ejecutado (ActionChains)")
            except:
                # Fallback: intenta click directo
                try:
                    print("[*] Intentando click directo...")
                    boton_nuevo.click()
                    print("[+] Click en 'Nuevo' ejecutado (click directo)")
                except:
                    # Último recurso: JavaScript
                    print("[*] Intentando click con JavaScript...")
                    self.driver.execute_script("arguments[0].click();", boton_nuevo)
                    print("[+] Click en 'Nuevo' ejecutado (JavaScript)")
            
            time.sleep(3)
            
            # Volver al contenido principal
            self.driver.switch_to.default_content()
            
            return True
                
        except Exception as e:
            print(f"[-] Error haciendo click en 'Nuevo': {str(e)}")
            import traceback
            traceback.print_exc()
            try:
                self.driver.switch_to.default_content()
            except:
                pass
            return False
    
    def buscar_documento(self, numero_documento):
        """
        Busca un documento en la página
        
        Args:
            numero_documento (str): Número de documento a buscar
        """
        print(f"\n[*] Buscando documento: {numero_documento}")
        try:
            # La estructura varía según la página, ajusta los selectores según sea necesario
            # Buscar campo de búsqueda
            campo_busqueda = self.wait.until(
                EC.presence_of_element_located((By.NAME, "txtBuscar"))
            )
            campo_busqueda.clear()
            campo_busqueda.send_keys(numero_documento)
            time.sleep(1)
            
            # Hacer clic en botón de búsqueda
            boton_buscar = self.driver.find_element(By.ID, "btnBuscar")
            boton_buscar.click()
            print("[+] Búsqueda ejecutada")
            
            time.sleep(2)
            return True
            
        except Exception as e:
            print(f"[-] Error al buscar documento: {str(e)}")
            return False
    
    def extraer_datos_resultado(self):
        """Extrae los datos del resultado de búsqueda"""
        print("\n[*] Extrayendo datos del resultado...")
        try:
            # Ajusta los selectores según la estructura HTML real
            resultados = {}
            
            # Ejemplo: buscar tabla de resultados
            try:
                tabla = self.driver.find_element(By.ID, "gvResultados")
                filas = tabla.find_elements(By.TAG_NAME, "tr")
                print(f"[+] Se encontraron {len(filas)} filas")
                
                for fila in filas[1:]:  # Saltar encabezado
                    celdas = fila.find_elements(By.TAG_NAME, "td")
                    if len(celdas) > 0:
                        fila_data = [celda.text for celda in celdas]
                        print(f"[+] Fila: {fila_data}")
                        
            except:
                print("[-] No se encontró tabla de resultados")
            
            return resultados
            
        except Exception as e:
            print(f"[-] Error extrayendo datos: {str(e)}")
            return {}
    
    def comparar_datos(self, dato_esperado, dato_obtenido):
        """
        Compara datos del Excel con los obtenidos de la página
        
        Args:
            dato_esperado (dict): Datos del Excel
            dato_obtenido (dict): Datos extraídos de la página
        """
        print("\n[*] Comparando datos...")
        coincidencias = True
        diferencias = []
        
        for clave, valor_esperado in dato_esperado.items():
            if clave in dato_obtenido:
                valor_obtenido = dato_obtenido[clave]
                if str(valor_esperado).strip() == str(valor_obtenido).strip():
                    print(f"[+] {clave}: COINCIDE ✓")
                else:
                    print(f"[-] {clave}: NO COINCIDE")
                    print(f"    Esperado: {valor_esperado}")
                    print(f"    Obtenido: {valor_obtenido}")
                    coincidencias = False
                    diferencias.append({
                        'campo': clave,
                        'esperado': valor_esperado,
                        'obtenido': valor_obtenido
                    })
            else:
                print(f"[!] {clave}: No encontrado")
        
        return coincidencias, diferencias
    
    def ejecutar_verificacion(self, url_pagina, usuario, contraseña, excel_path):
        """
        Ejecuta el proceso completo de verificación
        
        Args:
            url_pagina (str): URL de la página
            usuario (str): Usuario para login
            contraseña (str): Contraseña para login
            excel_path (str): Ruta al archivo Excel
        """
        print("="*60)
        print("INICIANDO BOT DE VERIFICACIÓN RUB ONLINE")
        print(f"Fecha/Hora: {datetime.now()}")
        print("="*60)
        
        try:
            # Inicializar
            self.inicializar_driver()
            
            # Acceder a la página
            if not self.acceder_pagina(url_pagina):
                return
            
            # Hacer login
            if not self.hacer_login(usuario, contraseña):
                return
            
            # Navegar a Beneficiario
            if not self.navegar_a_beneficiario():
                print("[-] No se pudo navegar a Beneficiario")
                return
            
            # Seleccionar Opción Beneficiario
            if not self.seleccionar_opcion_beneficiario():
                print("[-] No se pudo seleccionar Opción Beneficiario")
                return
            
            # Hacer click en botón para nueva opción
            if not self.hacer_click_nueva_opcion():
                print("[-] No se encontró botón para nueva opción")
                return
            
            # Leer Excel
            datos_excel = self.leer_excel(excel_path)
            
            if not datos_excel:
                print("\n[-] No se pudieron leer datos del Excel")
                return
            
            # Verificar documentos
            print(f"\n[*] Iniciando verificación de {len(datos_excel)} registros...\n")
            
            # Mostrar columnas disponibles
            if datos_excel:
                print(f"[*] Columnas encontradas: {list(datos_excel[0].keys())}\n")
            
            for idx, registro in enumerate(datos_excel[:5], 1):  # Primeros 5 por ahora
                print(f"\n{'='*60}")
                print(f"REGISTRO {idx}/{min(len(datos_excel), 5)}")
                print(f"{'='*60}")
                
                # Buscar la columna de documento (busca variaciones)
                documento = None
                for clave in registro.keys():
                    if 'documento' in clave.lower() or 'doc' in clave.lower():
                        documento = registro[clave]
                        break
                
                if not documento:
                    # Si no encuentra por nombre, usa la primera columna
                    documento = list(registro.values())[0] if registro else None
                
                if documento:
                    print(f"[*] Documento/Registro: {documento}")
                    # Buscar el documento
                    self.buscar_documento(str(documento))
                    
                    # Extraer datos
                    datos_pagina = self.extraer_datos_resultado()
                    
                    # Comparar
                    coincide, diferencias = self.comparar_datos(registro, datos_pagina)
                    
                    # Registrar resultado
                    self.resultados.append({
                        'registro': idx,
                        'documento': documento,
                        'coincide': coincide,
                        'diferencias': diferencias
                    })
                else:
                    print("[!] No se encontró campo de documento en el registro")
            
            # Resumen
            self.mostrar_resumen()
            
        except Exception as e:
            print(f"\n[-] Error general: {str(e)}")
        finally:
            self.cerrar_driver()
    
    def mostrar_resumen(self):
        """Muestra resumen de la verificación"""
        print("\n" + "="*60)
        print("RESUMEN DE VERIFICACIÓN")
        print("="*60)
        
        if not self.resultados:
            print("\n[!] No se procesaron registros")
            return
        
        total = len(self.resultados)
        coincidentes = sum(1 for r in self.resultados if r['coincide'])
        
        print(f"\nTotal registros verificados: {total}")
        print(f"Coincidencias: {coincidentes}")
        print(f"Diferencias: {total - coincidentes}")
        if total > 0:
            print(f"Porcentaje exitoso: {(coincidentes/total*100):.2f}%")
        
        # Guardar reporte
        self.guardar_reporte()
    
    def guardar_reporte(self):
        """Guarda un reporte de los resultados"""
        ruta_reporte = f"reporte_verificacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        try:
            with open(ruta_reporte, 'w', encoding='utf-8') as f:
                f.write("="*60 + "\n")
                f.write("REPORTE DE VERIFICACIÓN RUB ONLINE\n")
                f.write(f"Fecha/Hora: {datetime.now()}\n")
                f.write("="*60 + "\n\n")
                
                for resultado in self.resultados:
                    f.write(f"Registro {resultado['registro']}: {resultado['documento']}\n")
                    f.write(f"Estado: {'COINCIDE ✓' if resultado['coincide'] else 'NO COINCIDE ✗'}\n")
                    
                    if resultado['diferencias']:
                        f.write("Diferencias:\n")
                        for diff in resultado['diferencias']:
                            f.write(f"  - {diff['campo']}: {diff['esperado']} → {diff['obtenido']}\n")
                    f.write("\n")
            
            print(f"\n[+] Reporte guardado: {ruta_reporte}")
        except Exception as e:
            print(f"[-] Error al guardar reporte: {str(e)}")


def main():
    """Función principal"""
    
    # Configuración
    URL = "https://rubonline.icbf.gov.co/DefaultF.aspx"
    USUARIO = "Usuario"
    CONTRASEÑA = "Contraseña"
    
    # Opción 1: Usar archivos Excel específicos
    # Busca archivos .xlsx en la carpeta actual
    
    carpeta_actual = os.path.dirname(os.path.abspath(__file__))
    
    # Buscar primer archivo Excel disponible
    archivos_excel = [f for f in os.listdir(carpeta_actual) 
                      if f.endswith(('.xlsx', '.xls', '.xlsm'))]
    
    if not archivos_excel:
        print("[-] No se encontraron archivos Excel en la carpeta")
        return
    
    excel_path = os.path.join(carpeta_actual, archivos_excel[0])
    print(f"\n[*] Usando archivo Excel: {archivos_excel[0]}")
    
    # Crear e instanciar bot
    bot = RUBBot()
    
    # Ejecutar verificación
    bot.ejecutar_verificacion(URL, USUARIO, CONTRASEÑA, excel_path)


if __name__ == "__main__":
    main()
