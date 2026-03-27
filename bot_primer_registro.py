#!/usr/bin/env python3
"""
Bot Selenium - Primer Ejercicio Completo
Lee el registro de ALAN GARCIA RIVERA (1145330865) del Excel
y lo ingresa completamente en RUB Online
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
from datetime import datetime

print("\n" + "="*70)
print("BOT PRIMER EJERCICIO - ALAN GARCIA RIVERA (1145330865)")
print("="*70)

# ========== PASO 0: LEER EXCEL ==========
print("\n[*] Paso 0: Leyendo datos del Excel...")

ruta_excel = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

# Leer datos de la fila 18 (ALAN GARCIA RIVERA)
fila = 18
datos = {
    'documento': str(int(ws.cell(row=fila, column=17).value)),
    'primer_nombre': str(ws.cell(row=fila, column=5).value).strip(),
    'segundo_nombre': str(ws.cell(row=fila, column=6).value).strip(),
    'primer_apellido': str(ws.cell(row=fila, column=7).value).strip(),
    'segundo_apellido': str(ws.cell(row=fila, column=8).value).strip(),
    'sexo': str(ws.cell(row=fila, column=9).value).strip(),
    'fecha_nac': ws.cell(row=fila, column=13).value,
    'fecha_ingreso': ws.cell(row=fila, column=4).value,
    'tipo_doc': str(ws.cell(row=fila, column=16).value).strip(),
}

print(f"\n[+] Datos leídos del Excel (Fila {fila}):")
print(f"    Documento: {datos['documento']}")
print(f"    Nombre: {datos['primer_nombre']} {datos['segundo_nombre']} {datos['primer_apellido']} {datos['segundo_apellido']}")
print(f"    Sexo: {datos['sexo']}")
print(f"    Fecha Nac: {datos['fecha_nac']}")
print(f"    Fecha Ingreso: {datos['fecha_ingreso']}")
print(f"    Tipo Doc: {datos['tipo_doc']}")

wb.close()

# ========== CONFIGURAR SELENIUM ==========
options = webdriver.ChromeOptions()
options.add_argument('--disable-notifications')
options.add_argument('--disable-popup-blocking')

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

wait = WebDriverWait(driver, 20)

try:
    # PASO 1: LOGIN
    print("\n[*] Paso 1: Accediendo y haciendo login...")
    driver.get("https://rubonline.icbf.gov.co/DefaultF.aspx")
    time.sleep(3)
    
    campo_usuario = wait.until(EC.presence_of_element_located((By.ID, "UserName")))
    campo_usuario.clear()
    campo_usuario.send_keys("Usuario")
    
    campo_password = driver.find_element(By.ID, "Password")
    campo_password.clear()
    campo_password.send_keys("Contraseña")
    
    boton_login = driver.find_element(By.ID, "LoginButton")
    boton_login.click()
    time.sleep(8)
    print("[+] Login completado")
    
    # PASO 2: NAVIGATE TO BENEFICIARIO
    print("\n[*] Paso 2: Navegando al formulario...")
    
    # Click RUB online
    time.sleep(2)
    enlacesRUB = driver.find_elements(By.XPATH, "//a[contains(text(), 'Rub online')]")
    if enlacesRUB:
        driver.execute_script("arguments[0].scrollIntoView(true);", enlacesRUB[0])
        time.sleep(0.5)
        ActionChains(driver).move_to_element(enlacesRUB[0]).click().perform()
        time.sleep(3)
    
    # Click primer Beneficiario
    time.sleep(1)
    enlacesBene = driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")
    if enlacesBene:
        driver.execute_script("arguments[0].scrollIntoView(true);", enlacesBene[0])
        time.sleep(0.5)
        ActionChains(driver).move_to_element(enlacesBene[0]).click().perform()
        time.sleep(3)
    
    # Click segundo Beneficiario
    time.sleep(1)
    enlacesBene = driver.find_elements(By.XPATH, "//a[contains(text(), 'Beneficiario')]")
    if len(enlacesBene) > 1:
        driver.execute_script("arguments[0].scrollIntoView(true);", enlacesBene[1])
        time.sleep(0.5)
        ActionChains(driver).move_to_element(enlacesBene[1]).click().perform()
        time.sleep(4)
    
    print("[+] Navegación completada")
    
    # PASO 3: HACER CLIC EN BOTÓN "+"
    print("\n[*] Paso 3: Ingresando nuevo beneficiario (+)...")
    
    time.sleep(2)
    try:
        iframe = wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
        driver.switch_to.frame(iframe)
        print("[+] Iframe encontrado")
        
        boton_nuevo = wait.until(EC.element_to_be_clickable((By.ID, "btnNuevo")))
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_nuevo)
        time.sleep(0.5)
        ActionChains(driver).move_to_element(boton_nuevo).click().perform()
        print("[+] Botón '+' clickeado")
        time.sleep(3)
        
    except Exception as e:
        print(f"[-] Error: {str(e)}")
    
    # PASO 4: LLENAR FILTROS FIJOS
    print("\n[*] Paso 4: Llenando filtros fijos...")
    
    try:
        # 1. Uno a uno
        print("  [1] Forma vinculación: Uno a uno")
        radioBtns = driver.find_elements(By.XPATH, "//input[@type='radio']")
        for radio in radioBtns:
            if radio.get_attribute('value') and 'Uno a uno' in radio.get_attribute('value'):
                radio.click()
                print("[+] Seleccionado")
                break
        time.sleep(0.5)
        
        # 2. Dirección
        print("  [2] Dirección: Primera Infancia")
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDireccionesICBF")
        Select(select_elem).select_by_visible_text("Dirección de Primera Infancia")
        print("[+] Seleccionado")
        time.sleep(1)
        
        # 3. Regional
        print("  [3] Regional: Boyacá")
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlRegional")
        Select(select_elem).select_by_visible_text("Boyacá")
        print("[+] Seleccionado")
        time.sleep(1)
        
        # 4. Vigencia
        print("  [4] Vigencia: 2026")
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdVigencia")
        Select(select_elem).select_by_visible_text("2026")
        print("[+] Seleccionado")
        time.sleep(1)
        
        # 5. Contrato
        print("  [5] Contrato: OD 15 420272 00015 2026")
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNumeroContrato")
        Select(select_elem).select_by_visible_text("OD 15 420272 00015 2026")
        print("[+] Seleccionado")
        time.sleep(1)
        
        # 6. Servicio
        print("  [6] Servicio: EDUCACIÓN INICIAL...")
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNombreServicio")
        Select(select_elem).select_by_visible_text("EDUCACIÓN INICIAL EN EL HOGAR - FAMILIAR Y COMUNITARIA - 420272-2026")
        print("[+] Seleccionado")
        time.sleep(2)
        
        # 7. UDS
        print("  [7] UDS: DUITAMA D2 D3")
        selects = driver.find_elements(By.TAG_NAME, "select")
        for select_elem in selects:
            options = select_elem.find_elements(By.TAG_NAME, "option")
            for opt in options:
                if "DUITAMA D2 D3" in opt.text:
                    Select(select_elem).select_by_value(opt.get_attribute("value"))
                    print("[+] Seleccionado")
                    time.sleep(1)
                    break
        
        # 8. Tipo de beneficiario
        print("  [8] Tipo Beneficiario: NIÑO/NIÑA 6+ MESES")
        driver.execute_script("window.scrollBy(0, 300);")
        time.sleep(1)
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdTipoBeneficiario")
        options = select_elem.find_elements(By.TAG_NAME, "option")
        for opt in options:
            if "NIÑO O NIÑA ENTRE 6 MESES" in opt.text:
                Select(select_elem).select_by_value(opt.get_attribute("value"))
                print("[+] Seleccionado")
                break
        time.sleep(1)
        
        # 9. Tipo Documento
        print("  [9] Tipo Documento: REGISTRO CIVIL")
        driver.execute_script("window.scrollBy(0, 300);")
        time.sleep(1)
        select_elem = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdTipoDocumento")
        Select(select_elem).select_by_visible_text("REGISTRO CIVIL")
        print("[+] Seleccionado")
        time.sleep(1)
        
    except Exception as e:
        print(f"[-] Error en filtros: {str(e)}")
    
    # PASO 5: LLENAR DATOS DEL BENEFICIARIO
    print("\n[*] Paso 5: Llenando datos del beneficiario (EXCEL)...")
    
    try:
        driver.execute_script("window.scrollBy(0, 300);")
        time.sleep(1)
        
        # Número de Documento
        print(f"  [1] Documento: {datos['documento']}")
        campo_doc = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdentificacion")
        campo_doc.clear()
        campo_doc.send_keys(datos['documento'])
        print("[+] Ingresado")
        time.sleep(0.5)
        
        # Primer Nombre
        print(f"  [2] Primer Nombre: {datos['primer_nombre']}")
        campo_pnombre = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdPrimerNombre")
        campo_pnombre.clear()
        campo_pnombre.send_keys(datos['primer_nombre'])
        print("[+] Ingresado")
        time.sleep(0.5)
        
        # Segundo Nombre
        print(f"  [3] Segundo Nombre: {datos['segundo_nombre']}")
        campo_snombre = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdSegundoNombre")
        campo_snombre.clear()
        campo_snombre.send_keys(datos['segundo_nombre'])
        print("[+] Ingresado")
        time.sleep(0.5)
        
        # Primer Apellido
        print(f"  [4] Primer Apellido: {datos['primer_apellido']}")
        campo_papellido = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdPrimerApellido")
        campo_papellido.clear()
        campo_papellido.send_keys(datos['primer_apellido'])
        print("[+] Ingresado")
        time.sleep(0.5)
        
        # Segundo Apellido
        print(f"  [5] Segundo Apellido: {datos['segundo_apellido']}")
        campo_sapellido = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdSegundoApellido")
        campo_sapellido.clear()
        campo_sapellido.send_keys(datos['segundo_apellido'])
        print("[+] Ingresado")
        time.sleep(0.5)
        
        # Sexo
        print(f"  [6] Sexo: {datos['sexo']}")
        select_sexo = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdSexo")
        Select(select_sexo).select_by_visible_text("MASCULINO")
        print("[+] Seleccionado")
        time.sleep(1)
        
        # Fecha de Nacimiento
        print(f"  [7] Fecha Nacimiento: {datos['fecha_nac']}")
        campo_fnac = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdFechaNacimiento")
        campo_fnac.clear()
        # Formatear fecha como DD/MM/YYYY
        if datos['fecha_nac']:
            fecha_formato = datos['fecha_nac'].strftime("%d/%m/%Y")
            campo_fnac.send_keys(fecha_formato)
        print("[+] Ingresada")
        time.sleep(0.5)
        
        # Fecha de Ingreso (si existe)
        print(f"  [8] Fecha Ingreso: {datos['fecha_ingreso']}")
        try:
            campo_fingreso = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdFechaIngreso")
            campo_fingreso.clear()
            if datos['fecha_ingreso']:
                fecha_ingreso_formato = datos['fecha_ingreso'].strftime("%d/%m/%Y")
                campo_fingreso.send_keys(fecha_ingreso_formato)
            print("[+] Ingresada")
        except:
            print("[!] Campo no encontrado (puede estar deshabilitado)")
        
        time.sleep(0.5)
        
    except Exception as e:
        print(f"[-] Error al llenar datos: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # PASO 6: GUARDAR
    print("\n[*] Paso 6: Guardando registro...")
    
    try:
        driver.execute_script("window.scrollBy(0, 300);")
        time.sleep(1)
        
        # Buscar botón Guardar
        botones = driver.find_elements(By.TAG_NAME, "button")
        for btn in botones:
            if "Guardar" in btn.text or "GUARDAR" in btn.text or "guardar" in btn.text:
                print(f"[+] Botón encontrado: {btn.text}")
                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                time.sleep(0.5)
                ActionChains(driver).move_to_element(btn).click().perform()
                print("[+] Guardado clickeado")
                time.sleep(5)
                break
    except Exception as e:
        print(f"[-] Error al guardar: {str(e)}")
    
    print("\n" + "="*70)
    print("[✓] PRIMER EJERCICIO COMPLETADO - ALAN GARCIA RIVERA")
    print("="*70)
    
    print("\nNavegador abierto - Presiona Ctrl+C para cerrar")
    while True:
        time.sleep(1)
    
except KeyboardInterrupt:
    print("\n\n[!] Cerrando navegador...")
    driver.quit()
    print("[+] Cerrado")
    
except Exception as e:
    print(f"\n[-] Error: {str(e)}")
    import traceback
    traceback.print_exc()
    driver.quit()
