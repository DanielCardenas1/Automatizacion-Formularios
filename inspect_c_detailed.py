#!/usr/bin/env python3

import openpyxl
from pathlib import Path

BASE_DIR = Path("/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA")
excel_path = BASE_DIR / "CARGUE MASIVO 2026 _ DUITAMA C_.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Get first row C1 data
print("HEADERS (Row 2):")
for col in range(1, 30):
    cell_value = ws.cell(row=2, column=col).value
    print(f"  Col {col}: {cell_value}")

print("\n" + "="*80)
print("FIRST C1 RECORD (Row 5):")
for col in range(1, 30):
    cell_value = ws.cell(row=5, column=col).value
    print(f"  Col {col}: {cell_value}")

print("\n" + "="*80)
print("MAPPING VERIFICATION:")
print("Looking for key fields...")
for row in range(2, 4):
    for col in range(1, 30):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str):
            if 'nombre' in val.lower() or 'documento' in val.lower() or 'apellido' in val.lower() or 'sexo' in val.lower() or 'nacimiento' in val.lower():
                print(f"  Row {row}, Col {col}: {val}")
