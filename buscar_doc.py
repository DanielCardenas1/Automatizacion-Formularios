#!/usr/bin/env python3
import openpyxl

ruta = "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA/DUITAMA D/CARQUE MASIVO 2026_DUITAMA D.xlsx"
wb = openpyxl.load_workbook(ruta)
ws = wb.active

print(f"\n[*] Buscando '1145330865' en TODA la hoja...")
print(f"[*] Total de filas: {ws.max_row}\n")

# Búsqueda simple en todas las celdas
encontrado = False
contador = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    for col_idx, cell in enumerate(row, start=1):
        if cell.value and "1145330865" in str(cell.value):
            print(f"[+] ENCONTRADO en: Fila {row_idx}, Columna {col_idx}")
            print(f"    Valor: {cell.value}\n")
            
            fila = row_idx
            print(f"[DATOS ENCONTRADOS EN FILA {fila}]:\n")
            
            campos = {
                1: "CONTRATO", 2: "MUNICIPIO", 3: "NOMBRE UDS", 4: "FECHA INGRESO",
                5: "PRIMER NOMBRE", 6: "SEGUNDO NOMBRE", 7: "PRIMER APELLIDO",
                8: "SEGUNDO APELLIDO", 9: "SEXO", 13: "FECHA NAC", 16: "TIPO DOC", 17: "NUM DOC",
            }
            
            for col, etiq in campos.items():
                val = ws.cell(row=fila, column=col).value
                if val:
                    print(f"  {etiq}: {val}")
            
            encontrado = True
            break
    if encontrado:
        break

if not encontrado:
    print("[-] Documento NO encontrado en el archivo")
    # Mostrar algunos documentos de ejemplo
    print("\n[*] Primeros documentos en el archivo:")
    for i in range(3, min(10, ws.max_row + 1)):
        doc = ws.cell(row=i, column=17).value
        nombre = ws.cell(row=i, column=5).value
        print(f"  Fila {i}: {doc} - {nombre}")

wb.close()
