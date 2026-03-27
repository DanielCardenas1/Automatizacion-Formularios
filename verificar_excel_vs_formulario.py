#!/usr/bin/env python3
"""
Script de Verificación: Compara datos del Excel con lo que trae el formulario al buscar
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import time
import os
from pathlib import Path
from datetime import datetime
import unicodedata
import re
import sys
from urllib.parse import urljoin

BASE_DIR = Path("/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA")

try:
    sys.stdout.reconfigure(line_buffering=True, write_through=True)
    sys.stderr.reconfigure(line_buffering=True, write_through=True)
except Exception:
    pass

EQUIVALENCIAS_TIPO_VIA = {
    'AC': ['AC', 'AVENIDA CALLE'],
    'AK': ['AK', 'AVENIDA CARRERA'],
    'AU': ['AU', 'AUTOPISTA'],
    'AV': ['AV', 'AVENIDA'],
    'BL': ['BL', 'BULEVAR'],
    'CL': ['CL', 'CALLE'],
    'KR': ['KR', 'CARRERA'],
    'CT': ['CT', 'CARRETERA'],
    'CQ': ['CQ', 'CIRCULAR'],
    'DG': ['DG', 'DIAGONAL'],
    'PJ': ['PJ', 'PASAJE'],
    'PS': ['PS', 'PASEO'],
    'PT': ['PT', 'PEATONAL'],
    'TV': ['TV', 'TRANSVERSAL'],
    'TC': ['TC', 'TRONCAL'],
    'VT': ['VT', 'VARIANTE'],
    'VI': ['VI', 'VIA', 'VIA PRINCIPAL'],
}

VEREDAS_DUITAMA = [
    'EL CARMEN',
    'SANTA ANA',
    'SIRATA',
    'SAN ANTONIO SUR',
    'HIGUERAS',
    'CAJON',
    'LA PARROQUIA',
    'QUEBRADA DE BECERRAS',
    'LA PRADERA',
    'SAN LORENZO DE ARRIBA',
    'SURBA Y BONZA',
    'SAN ANTONIO NORTE',
    'AVENDANOS',
    'SAN LORENZO DE ABAJO',
    'SANTA BARBARA',
    'SANTA HELENA',
    'SAN LUIS',
    'AGUATENDIDA',
    'TOCOGUA',
    'OTRO',
]

CATALOGO_UDS = {
    "DUITAMA C1": "152381148301",
    "DUITAMA F3": "152381130836",
    "DUITAMA A1": "152381147353",
    "DUITAMA A2 A3": "1523800124756",
    "DUITAMA B1": "152381130854",
    "DUITAMA B2 B3": "152381148302",
    "DUITAMA C2 C3": "152381150770",
    "DUITAMA D1": "152381130829",
    "DUITAMA D2 D3": "1523800124748",
    "DUITAMA E1": "152381139987",
    "DUITAMA E2 E3": "152381139989",
    "DUITAMA F1": "152381130850",
    "DUITAMA F2": "152381130852",
    "DUITAMA G1 G2": "152381154299",
    "DUITAMA G3": "152381145301",
}

CATALOGO_FOTOS = {
    "A1": BASE_DIR / "DUITAMA A" / "FOTOS" / "FOTOS A1",
    "A2A3": BASE_DIR / "DUITAMA A" / "FOTOS" / "FOTOS A2 Y A3",
    "C1": BASE_DIR / "DUITAMA C" / "DUITAMA C1" / "C1",
    "C2": BASE_DIR / "DUITAMA C" / "DUITAMA C2" / "FOTOS",
    "C3": BASE_DIR / "DUITAMA C" / "DUITAMA C3" / "FOTOS",
    "D1": BASE_DIR / "DUITAMA D" / "DUITAMA D1" / "DUITAMA D1 FOTOS",
    "D2": BASE_DIR / "DUITAMA D" / "DUITAMA D2" / "DUITAMA D2 FOTOS",
    "D3": BASE_DIR / "DUITAMA D" / "DUITAMA D3" / "DUITAMA D3 FOTOS",
    "F1": BASE_DIR / "DUITAMA F" / "DUITAMA F1" / "FOTOS" / "DUITAMA F1",
    "F2": BASE_DIR / "DUITAMA F" / "DUITAMA F2" / "FOTOS",
    "F3": BASE_DIR / "DUITAMA F" / "DUITAMA F3" / "FOTOS",
    "E1": BASE_DIR / "DUITAMA E" / "DUITAMA E1" / "FOTOS",
    "E2": BASE_DIR / "DUITAMA E" / "DUITAMA E2" / "FOTOS",
    "E3": BASE_DIR / "DUITAMA E" / "DUITAMA E3" / "FOTOS",
    "G1": BASE_DIR / "DUITAMA G" / "DUITAMA G1" / "FOTOS",
    "G2": BASE_DIR / "DUITAMA G" / "DUITAMA G2" / "FOTOS",
    "G3": BASE_DIR / "DUITAMA G" / "DUITAMA G3" / "FOTOS",
}

RUTA_EXCEL_DUITAMA_E = Path(
    os.getenv(
        "RUTA_EXCEL_DUITAMA_E",
        "/Users/stevenruiz/Downloads/CARGUE MASIVO EIH DUITAMA - F1 BACKUP/DUITAMA E/CARGUE MASIVO DUITAMA E 2026.xlsx",
    )
)

CONFIG_PRESETS = {
    "D3": {
        "descripcion": "DUITAMA D3",
        "filtro_excel": ["D3"],
        "uds_selector": "DUITAMA D2 D3",
        "uds_id": CATALOGO_UDS["DUITAMA D2 D3"],
        "ruta_excel": BASE_DIR / "DUITAMA D" / "CARQUE MASIVO 2026_DUITAMA D.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["D3"]],
        "prefijo_reporte": "d3",
        "indice_inicio": 0,
        "excel_column_offset": 0,
    },
    "A2A3": {
        "descripcion": "DUITAMA A2 Y A3",
        "filtro_excel": ["A2", "A3"],
        "uds_selector": "DUITAMA A2 A3",
        "uds_id": CATALOGO_UDS["DUITAMA A2 A3"],
        "ruta_excel": BASE_DIR / "DUITAMA A" / "CARGUE MASIVO_DUITAMA A_ICBF_2026.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["A2A3"]],
        "prefijo_reporte": "a2a3",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "F1": {
        "descripcion": "DUITAMA F1",
        "filtro_excel": ["F1"],
        "uds_selector": "DUITAMA F1",
        "uds_id": CATALOGO_UDS["DUITAMA F1"],
        "ruta_excel": BASE_DIR / "DUITAMA F" / "CARGUE MASIVO 2026_DUITAMA F_ACTUALIZADO.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["F1"]],
        "prefijo_reporte": "f1",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "F2": {
        "descripcion": "DUITAMA F2",
        "filtro_excel": ["F2"],
        "uds_selector": "DUITAMA F2",
        "uds_id": CATALOGO_UDS["DUITAMA F2"],
        "ruta_excel": BASE_DIR / "DUITAMA F" / "CARGUE MASIVO 2026_DUITAMA F_ACTUALIZADO.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["F2"]],
        "prefijo_reporte": "f2",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "F3": {
        "descripcion": "DUITAMA F3",
        "filtro_excel": ["F3"],
        "uds_selector": "DUITAMA F3",
        "uds_id": CATALOGO_UDS["DUITAMA F3"],
        "ruta_excel": BASE_DIR / "DUITAMA F" / "CARGUE MASIVO 2026_DUITAMA F_ACTUALIZADO.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["F3"]],
        "prefijo_reporte": "f3",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "E1": {
        "descripcion": "DUITAMA E1",
        "filtro_excel": ["E1"],
        "uds_selector": "DUITAMA E1",
        "uds_id": CATALOGO_UDS["DUITAMA E1"],
        "ruta_excel": RUTA_EXCEL_DUITAMA_E,
        "rutas_fotos": [CATALOGO_FOTOS["E1"]],
        "prefijo_reporte": "e1",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "E2": {
        "descripcion": "DUITAMA E2",
        "filtro_excel": ["E2"],
        "uds_selector": "DUITAMA E2 E3",
        "uds_id": CATALOGO_UDS["DUITAMA E2 E3"],
        "ruta_excel": RUTA_EXCEL_DUITAMA_E,
        "rutas_fotos": [CATALOGO_FOTOS["E2"]],
        "prefijo_reporte": "e2",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "E3": {
        "descripcion": "DUITAMA E3",
        "filtro_excel": ["E3"],
        "uds_selector": "DUITAMA E2 E3",
        "uds_id": CATALOGO_UDS["DUITAMA E2 E3"],
        "ruta_excel": RUTA_EXCEL_DUITAMA_E,
        "rutas_fotos": [CATALOGO_FOTOS["E3"]],
        "prefijo_reporte": "e3",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "G1": {
        "descripcion": "DUITAMA G1",
        "filtro_excel": ["G1"],
        "uds_selector": "DUITAMA G1 G2",
        "uds_id": CATALOGO_UDS["DUITAMA G1 G2"],
        "ruta_excel": BASE_DIR / "DUITAMA G" / "CARGUE MASIVO_DUITAMA_G1_G2_G3_2026.xlsm",
        "rutas_fotos": [CATALOGO_FOTOS["G1"]],
        "prefijo_reporte": "g1",
        "indice_inicio": 0,
        "excel_column_offset": 1,
        "excel_column_offset_from": 3,
        "excel_column_offset_extra": 1,
        "excel_column_offset_extra_from": 5,
    },
    "G2": {
        "descripcion": "DUITAMA G2",
        "filtro_excel": ["G2"],
        "uds_selector": "DUITAMA G1 G2",
        "uds_id": CATALOGO_UDS["DUITAMA G1 G2"],
        "ruta_excel": BASE_DIR / "DUITAMA G" / "CARGUE MASIVO_DUITAMA_G1_G2_G3_2026.xlsm",
        "rutas_fotos": [CATALOGO_FOTOS["G2"]],
        "prefijo_reporte": "g2",
        "indice_inicio": 0,
        "excel_column_offset": 1,
        "excel_column_offset_from": 3,
        "excel_column_offset_extra": 1,
        "excel_column_offset_extra_from": 5,
    },
    "G3": {
        "descripcion": "DUITAMA G3",
        "filtro_excel": ["G3"],
        "uds_selector": "DUITAMA G3",
        "uds_id": CATALOGO_UDS["DUITAMA G3"],
        "ruta_excel": BASE_DIR / "DUITAMA G" / "CARGUE MASIVO_DUITAMA_G1_G2_G3_2026.xlsm",
        "rutas_fotos": [CATALOGO_FOTOS["G3"]],
        "prefijo_reporte": "g3",
        "indice_inicio": 0,
        "excel_column_offset": 1,
        "excel_column_offset_from": 3,
        "excel_column_offset_extra": 1,
        "excel_column_offset_extra_from": 5,
    },
    "C1": {
        "descripcion": "DUITAMA C1",
        "filtro_excel": ["C1"],
        "uds_selector": "DUITAMA C1",
        "uds_id": CATALOGO_UDS["DUITAMA C1"],
        "ruta_excel": BASE_DIR / "CARGUE MASIVO 2026 _ DUITAMA C_.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["C1"]],
        "prefijo_reporte": "c1",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "C2": {
        "descripcion": "DUITAMA C2",
        "filtro_excel": ["C2"],
        "uds_selector": "DUITAMA C2 C3",
        "uds_id": CATALOGO_UDS["DUITAMA C2 C3"],
        "ruta_excel": BASE_DIR / "CARGUE MASIVO 2026 _ DUITAMA C_.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["C2"]],
        "prefijo_reporte": "c2",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
    "C3": {
        "descripcion": "DUITAMA C3",
        "filtro_excel": ["C3"],
        "uds_selector": "DUITAMA C2 C3",
        "uds_id": CATALOGO_UDS["DUITAMA C2 C3"],
        "ruta_excel": BASE_DIR / "CARGUE MASIVO 2026 _ DUITAMA C_.xlsx",
        "rutas_fotos": [CATALOGO_FOTOS["C3"]],
        "prefijo_reporte": "c3",
        "indice_inicio": 0,
        "excel_column_offset": 1,
    },
}

CONFIG_PRESET = (os.getenv("CONFIG_PRESET") or "D3").strip().upper()
CONFIG_EJECUCION = dict(CONFIG_PRESETS.get(CONFIG_PRESET, CONFIG_PRESETS["D3"]))

print("\n" + "="*80)
print(f"VERIFICACIÓN: EXCEL vs FORMULARIO ({CONFIG_EJECUCION['descripcion']})")
print("="*80)

FECHA_CAPTURA_UBICACION = "06/02/2026"
HORA_CAPTURA_UBICACION = "09:30:23"
DEBUG_CAMPOS_CON_VALOR = False
ESPERA_BUSQUEDA = 2
ESPERA_CARGA_FOTO = 3
ESPERA_HABILITAR_GUARDAR = 0.5
ESPERA_POST_GUARDADO = 3
ESPERA_NUEVO_REGISTRO = 1
ESPERA_MODO_AGREGAR_PERSONA = 0.35
ESPERA_GUARDAR_PERSONA = 0.6
ESPERA_ACTUALIZAR_PERSONA = 0.6
ESTANDARIZAR_CUNDINAMARCA_BOGOTA_DC = True
ESPERA_POST_CLICK_MENU = 0.15
ESPERA_POST_FILTRO = 0.35
DIAGNOSTICO_SESION = False
DIAGNOSTICO_TIEMPOS_POST_LOGIN = True
LIMITE_REGISTROS_DIAGNOSTICO = None
TAB_ACTIVA = ""


def texto_excel(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def normalizar_filtro_excel(filtro_excel):
    if filtro_excel is None:
        return []
    if isinstance(filtro_excel, str):
        filtro_excel = [filtro_excel]

    filtros = []
    for valor in filtro_excel:
        texto = texto_excel(valor).upper()
        if texto:
            filtros.append(texto)
    return filtros


def pertenece_a_segmento_excel(nombre_uds, filtro_excel):
    texto = texto_excel(nombre_uds).upper()
    filtros = normalizar_filtro_excel(filtro_excel)
    if not texto:
        return False
    if not filtros or "TODOS" in filtros:
        return True

    for segmento_objetivo in filtros:
        if re.search(rf"(?<![A-Z0-9]){re.escape(segmento_objetivo)}(?![A-Z0-9])", texto):
            return True
    return False


def descripcion_filtro_excel(filtro_excel):
    filtros = normalizar_filtro_excel(filtro_excel)
    if not filtros or "TODOS" in filtros:
        return "TODOS"
    return " Y ".join(filtros)


def inferir_vereda_desde_texto(texto):
    normalizado = normalizar_texto(texto)
    if not normalizado:
        return ""

    for vereda in VEREDAS_DUITAMA:
        if vereda in normalizado:
            return vereda
    return ""


def nombre_completo(*partes):
    return " ".join(texto_excel(parte) for parte in partes if texto_excel(parte))


def fecha_excel(valor):
    if not valor:
        return ""
    if hasattr(valor, 'strftime'):
        return valor.strftime("%d/%m/%Y")
    return texto_excel(valor)


def fecha_comparable(valor):
    if not valor:
        return None
    if hasattr(valor, 'date'):
        try:
            return valor.date()
        except Exception:
            pass

    texto = texto_excel(valor)
    if not texto:
        return None

    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except Exception:
            continue
    return None


def normalizar_telefono_excel(valor):
    texto = texto_excel(valor)
    if not texto:
        return ""

    segmentos = [re.sub(r"\D", "", segmento) for segmento in re.split(r"[^\d]+", texto) if re.sub(r"\D", "", segmento)]
    candidatos = [segmento for segmento in segmentos if len(segmento) in {7, 10}]
    if candidatos:
        return candidatos[0]

    solo_digitos = re.sub(r"\D", "", texto)
    if len(solo_digitos) in {7, 10}:
        return solo_digitos
    if len(solo_digitos) > 10:
        return solo_digitos[:10]
    return solo_digitos


def debe_sobrescribir_ultimo_reporte():
    valor = (os.getenv("SOBRESCRIBIR_ULTIMO_REPORTE") or "").strip().upper()
    return valor in {"1", "SI", "TRUE", "YES"}


def obtener_ruta_reporte_txt():
    patron = f"reporte_inconsistencias_{CONFIG_EJECUCION['prefijo_reporte']}_*.txt"
    if debe_sobrescribir_ultimo_reporte():
        existentes = sorted(BASE_DIR.glob(patron), key=lambda path: path.stat().st_mtime, reverse=True)
        if existentes:
            return existentes[0]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return BASE_DIR / f"reporte_inconsistencias_{CONFIG_EJECUCION['prefijo_reporte']}_{timestamp}.txt"


def columna_excel(numero_columna):
    desplazamiento = CONFIG_EJECUCION.get('excel_column_offset', 0)
    desde_columna = CONFIG_EJECUCION.get('excel_column_offset_from', 5)
    desplazamiento_extra = CONFIG_EJECUCION.get('excel_column_offset_extra', 0)
    desde_columna_extra = CONFIG_EJECUCION.get('excel_column_offset_extra_from', 0)

    columna = numero_columna
    if numero_columna >= desde_columna:
        columna += desplazamiento
    if desplazamiento_extra and desde_columna_extra and numero_columna >= desde_columna_extra:
        columna += desplazamiento_extra
    return columna


def parsear_georreferencia(valor, tipo=''):
    texto = texto_excel(valor)
    def normalizar_componentes(grados_valor, minutos_valor, segundos_valor):
        try:
            grados_num = int(float(grados_valor or 0))
            minutos_num = float(minutos_valor or 0)
            segundos_num = float(segundos_valor or 0)

            if minutos_num % 1:
                segundos_num += (minutos_num % 1) * 60
                minutos_num = int(minutos_num)
            else:
                minutos_num = int(minutos_num)

            segundos_num = int(round(segundos_num))
            if segundos_num >= 60:
                minutos_num += segundos_num // 60
                segundos_num = segundos_num % 60

            if minutos_num >= 60:
                grados_num += minutos_num // 60
                minutos_num = minutos_num % 60

            return [str(grados_num), str(minutos_num), str(segundos_num)]
        except Exception:
            return [texto_excel(grados_valor), texto_excel(minutos_valor), texto_excel(segundos_valor)]

    coincidencia = re.search(r"([NSEOW])?\s*(\d+)°\s*(\d+(?:\.\d+)?)'\s*(\d+(?:\.\d+)?)\"?", texto, re.IGNORECASE)
    if coincidencia:
        return normalizar_componentes(coincidencia.group(2), coincidencia.group(3), coincidencia.group(4))

    coincidencia_minutos = re.search(r"([NSEOW])?\s*(\d+)°\s*(\d+(?:\.\d+)?)'?\s*\"?[NSEOW]?", texto, re.IGNORECASE)
    if coincidencia_minutos:
        return normalizar_componentes(coincidencia_minutos.group(2), coincidencia_minutos.group(3), 0)

    numeros = re.findall(r"\d+(?:\.\d+)?", texto)
    if not numeros:
        return ["", "", ""]

    max_grados = 13 if tipo == 'latitud' else 180 if tipo == 'longitud' else None
    tiene_grado_explicito = '°' in texto
    primer_numero = numeros[0]

    if (not tiene_grado_explicito) and max_grados is not None and primer_numero.isdigit() and len(primer_numero) >= 3:
        grados_candidato = int(primer_numero[:-2])
        minutos_candidato = int(primer_numero[-2:])
        if 0 <= grados_candidato <= max_grados and 0 <= minutos_candidato < 60:
            segundos_candidato = numeros[1] if len(numeros) > 1 else '0'
            return [
                str(grados_candidato),
                str(minutos_candidato),
                str(int(float(segundos_candidato))),
            ]

    grados = numeros[0] if len(numeros) > 0 else ""
    minutos = numeros[1] if len(numeros) > 1 else ""
    segundos = numeros[2] if len(numeros) > 2 else 0
    return normalizar_componentes(grados, minutos, segundos)


def obtener_direccion_georreferencia(valor, tipo):
    texto = texto_excel(valor).upper()
    coincidencia = re.search(r"([NSEOW])", texto)
    if coincidencia:
        direccion = coincidencia.group(1)
        if direccion == 'W':
            return 'O'
        return direccion

    if tipo == 'latitud':
        return 'N'
    if tipo == 'longitud':
        return 'O'
    return ''


def expandir_prefijo_direccion_compacto(texto):
    base = texto_excel(texto).strip().upper()
    reemplazos = [
        (r'^C(?=\d)', 'CALLE '),
        (r'^CL(?=\d)', 'CALLE '),
        (r'^K(?=\d)', 'CARRERA '),
        (r'^KR(?=\d)', 'CARRERA '),
        (r'^CRA(?=\d)', 'CARRERA '),
        (r'^DG(?=\d)', 'DIAGONAL '),
        (r'^TV(?=\d)', 'TRANSVERSAL '),
    ]
    for patron, reemplazo in reemplazos:
        if re.match(patron, base):
            return re.sub(patron, reemplazo, base, count=1)
    return base


def descomponer_direccion_residencia(direccion):
    texto = expandir_prefijo_direccion_compacto(direccion)
    if not texto:
        return {}

    patrones_via = [
        (r'^(AVENIDA\s+CALLE|AV\.?\s*CALLE|AC)\b', 'AC', 'Calle', 'Avenida Calle'),
        (r'^(AVENIDA\s+CARRERA|AV\.?\s*CARRERA|AK)\b', 'AK', 'Carrera', 'Avenida Carrera'),
        (r'^(AUTOPISTA|AU)\b', 'AU', 'Autopista', 'Autopista'),
        (r'^(AVENIDA|AV)\b', 'AV', 'Avenida', 'Avenida'),
        (r'^(BULEVAR|BL)\b', 'BL', 'Bulevar', 'Bulevar'),
        (r'^(CALLE|CL)\b', 'CL', 'Calle', 'Calle'),
        (r'^(CARRERA|KR)\b', 'KR', 'Carrera', 'Carrera'),
        (r'^(CARRETERA|CT)\b', 'CT', 'Carretera', 'Carretera'),
        (r'^(CIRCULAR|CQ)\b', 'CQ', 'Circular', 'Circular'),
        (r'^(DIAGONAL|DG)\b', 'DG', 'Diagonal', 'Diagonal'),
        (r'^(PASAJE|PJ)\b', 'PJ', 'Pasaje', 'Pasaje'),
        (r'^(PASEO|PS)\b', 'PS', 'Paseo', 'Paseo'),
        (r'^(PEATONAL|PT)\b', 'PT', 'Peatonal', 'Peatonal'),
        (r'^(TRANSVERSAL|TV)\b', 'TV', 'Transversal', 'Transversal'),
        (r'^(TRONCAL|TC)\b', 'TC', 'Troncal', 'Troncal'),
        (r'^(VARIANTE|VT)\b', 'VT', 'Variante', 'Variante'),
        (r'^(VIA|VI)\b', 'VI', 'Via', 'Vía'),
    ]

    via_codigo = ''
    texto_restante = texto.strip()
    for patron, codigo_iso, nombre_corto, valor_select in patrones_via:
        coincidencia = re.match(patron, texto_restante, flags=re.IGNORECASE)
        if coincidencia:
            via_codigo = codigo_iso
            texto_restante = texto_restante[coincidencia.end():].strip()
            break

    texto_restante = re.sub(r'\s+', ' ', texto_restante.upper()).strip()
    texto_parseo = re.sub(r'[#-]', ' ', texto_restante)
    texto_parseo = re.sub(r'\s+', ' ', texto_parseo).strip()
    patron_urbano = re.compile(
        r'^(?P<nombre>\d+)'
        r'\s*(?P<letra>[A-Z]{1,2})?'
        r'\s*(?P<bis>BIS)?'
        r'\s*(?P<sentido>NORTE|SUR|N|S)?'
        r'\s+(?P<numero>\d+)'
        r'\s*(?P<letra3>[A-Z]{1,2})?'
        r'\s*(?P<bis2>BIS)?'
        r'\s*[-# ]\s*(?P<placa>\d+)'
        r'\s*(?P<sentido2>ESTE|OESTE|E|O)?'
        r'\s*(?P<complemento>.*)$'
    )
    coincidencia = patron_urbano.match(texto_parseo)
    if coincidencia:
        datos = coincidencia.groupdict()
        return {
            'via': via_codigo,
            'nombre_via': texto_excel(datos.get('nombre')),
            'letra': texto_excel(datos.get('letra')),
            'bis': 'BIS' if texto_excel(datos.get('bis')) else 'N/A',
            'sentido': {'N': 'NORTE', 'S': 'SUR'}.get(texto_excel(datos.get('sentido')), texto_excel(datos.get('sentido'))),
            'numero': texto_excel(datos.get('numero')),
            'letra3': texto_excel(datos.get('letra3')),
            'bis2': 'BIS' if texto_excel(datos.get('bis2')) else 'N/A',
            'placa': texto_excel(datos.get('placa')),
            'sentido2': {'E': 'ESTE', 'O': 'OESTE'}.get(texto_excel(datos.get('sentido2')), texto_excel(datos.get('sentido2'))),
            'complemento': texto_excel(datos.get('complemento')),
        }

    patron_urbano_sin_placa = re.compile(
        r'^(?P<nombre>\d+)'
        r'\s*(?P<letra>[A-Z]{1,2})?'
        r'\s*(?P<bis>BIS)?'
        r'\s*(?P<sentido>NORTE|SUR|N|S)?'
        r'\s+(?P<numero>\d+)'
        r'\s*(?P<letra3>[A-Z]{1,2})?'
        r'\s*(?P<bis2>BIS)?'
        r'\s*(?P<complemento>.*)$'
    )
    coincidencia = patron_urbano_sin_placa.match(texto_parseo)
    if coincidencia:
        datos = coincidencia.groupdict()
        return {
            'via': via_codigo,
            'nombre_via': texto_excel(datos.get('nombre')),
            'letra': texto_excel(datos.get('letra')),
            'bis': 'BIS' if texto_excel(datos.get('bis')) else 'N/A',
            'sentido': {'N': 'NORTE', 'S': 'SUR'}.get(texto_excel(datos.get('sentido')), texto_excel(datos.get('sentido'))),
            'numero': texto_excel(datos.get('numero')),
            'letra3': texto_excel(datos.get('letra3')),
            'bis2': 'BIS' if texto_excel(datos.get('bis2')) else 'N/A',
            'placa': '',
            'sentido2': '',
            'complemento': texto_excel(datos.get('complemento')),
        }

    return {
        'via': via_codigo,
        'nombre_via': texto_restante,
        'letra': '',
        'bis': 'N/A',
        'sentido': '',
        'numero': '',
        'letra3': '',
        'bis2': 'N/A',
        'placa': '',
        'sentido2': '',
        'complemento': '',
    }


def obtener_controles_direccion_residencia():
    prefijo = "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_"
    controles = []
    for elemento in driver.find_elements(By.CSS_SELECTOR, f"[id^='{prefijo}']"):
        try:
            if not elemento.is_displayed():
                continue
            element_id = elemento.get_attribute("id") or ""
            tag_name = (elemento.tag_name or "").lower()
            input_type = (elemento.get_attribute("type") or "").lower()
            if tag_name == "select":
                controles.append((element_id, "select"))
            elif tag_name == "input" and input_type in {"text", "search", "tel", ""}:
                controles.append((element_id, "input"))
        except Exception:
            continue
    return controles


def select_tiene_opcion(element_id, texto_objetivo):
    objetivo = normalizar_texto(texto_objetivo)
    try:
        elemento = driver.find_element(By.ID, element_id)
        for opcion in elemento.find_elements(By.TAG_NAME, "option"):
            texto_opcion = normalizar_texto(opcion.text)
            if texto_opcion == objetivo or objetivo in texto_opcion or texto_opcion in objetivo:
                return True
    except Exception:
        return False
    return False


def normalizar_direccion_residencia_texto(texto):
    base = texto_excel(texto).upper()
    base = expandir_prefijo_direccion_compacto(base)
    base = re.sub(r'\bAPTO\b', 'APT', base)
    base = re.sub(r'\bAPARTAMENTO\b', 'APT', base)
    base = re.sub(r'[#\-.,]', ' ', base)
    base = re.sub(r'\bN/A\b', ' ', base)
    base = re.sub(r'\s+', ' ', base)
    return base.strip()


def leer_partes_direccion_residencia_actual():
    partes = {
        'via': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlVia"),
        'letra': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlLetra"),
        'bis': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlBis"),
        'sentido': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlSentido"),
        'letra3': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlLetra3"),
        'bis2': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlBis2"),
        'sentido2': obtener_texto_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlSentido2"),
        'nombre_via': obtener_valor_por_xpath("//*[@id='cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtNombreVia']"),
        'numero': obtener_valor_por_xpath("//*[@id='cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtNumero']"),
        'placa': obtener_valor_por_xpath("//*[@id='cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtPlaca']"),
        'complemento': obtener_valor_por_xpath("//*[@id='cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtComplemento']"),
    }
    return partes


def partes_direccion_coinciden(partes_excel, partes_actuales):
    campos = ['via', 'nombre_via', 'letra', 'bis', 'sentido', 'numero', 'letra3', 'bis2', 'placa', 'sentido2', 'complemento']
    for campo in campos:
        esperado = normalizar_direccion_residencia_texto(partes_excel.get(campo, ''))
        actual = normalizar_direccion_residencia_texto(partes_actuales.get(campo, ''))
        if esperado in {'', 'N A'}:
            esperado = ''
        if actual in {'', 'N A'}:
            actual = ''
        if esperado != actual:
            return False
    return True


def direccion_requiere_modo_libre(partes):
    if not partes:
        return True

    placa = texto_excel(partes.get('placa'))
    complemento = normalizar_direccion_residencia_texto(partes.get('complemento', ''))
    letra3 = normalizar_direccion_residencia_texto(partes.get('letra3', ''))

    if placa:
        return False

    tokens_complemento = {'APT', 'APTO', 'APARTAMENTO', 'TORRE', 'T', 'INTERIOR', 'INT', 'BLOQUE', 'BL', 'CASA'}
    if any(token in complemento.split() for token in tokens_complemento):
        return True
    if letra3 in {'AP', 'APT', 'APTO', 'T'}:
        return True

    return not bool(texto_excel(partes.get('via')) and texto_excel(partes.get('nombre_via')) and texto_excel(partes.get('numero')) and placa)


def direccion_residencia_coincide(direccion_excel, direccion_formulario):
    if es_valor_ausente(direccion_excel):
        return True

    return normalizar_direccion_residencia_texto(direccion_excel) == normalizar_direccion_residencia_texto(direccion_formulario)


def completar_direccion_residencia_libre(direccion_excel):
    direccion_objetivo = texto_excel(direccion_excel)
    element_id = "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtDireccion"

    asignar_radio_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_rbtnDetalleZonaGeo_0",
        "Detalle zona: No",
        timeout=8.0,
    )
    esperar_postback_finalizado(2)
    if asignar_valor_input_verificando_por_id(
        element_id,
        direccion_objetivo,
        "Dirección residencia libre",
        timeout=8.0,
        reintentos=3,
    ):
        return True

    print("  [~] Dirección residencia libre bloqueada: usando cuadro de texto de zona como respaldo")
    return completar_direccion_residencia_respaldo_zona(direccion_objetivo)


def completar_direccion_residencia_respaldo_zona(direccion_excel):
    direccion_objetivo = texto_excel(direccion_excel)
    if not direccion_objetivo:
        return False

    campos_respaldo = [
        ("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtBarrioVereda", "Barrio/Vereda (respaldo dirección)"),
        ("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtObservacionAdicionalUbicacion", "Observación adicional ubicación (respaldo dirección)"),
        ("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtNombreZonaResto", "Nombre Zona Resto (respaldo dirección)"),
    ]

    for element_id, descripcion in campos_respaldo:
        if asignar_valor_input_verificando_por_id(
            element_id,
            direccion_objetivo,
            descripcion,
            timeout=8.0,
            reintentos=3,
        ):
            return True

    print("  [!] No fue posible registrar la dirección en ningún campo de respaldo de zona")
    return False


def leer_direccion_residencia_compuesta():
    try:
        direccion_libre = (
            driver.find_element(
                By.ID,
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtDireccion",
            ).get_attribute("value")
            or ""
        ).strip()
        if direccion_libre:
            return direccion_libre
    except Exception:
        pass

    partes = []
    campos_select = [
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlVia",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlLetra",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlBis",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlSentido",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlLetra3",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlBis2",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlSentido2",
    ]
    campos_input = [
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtNombreVia",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtNumero",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtPlaca",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtComplemento",
    ]

    for element_id in campos_select:
        try:
            valor = Select(driver.find_element(By.ID, element_id)).first_selected_option.text.strip()
            if valor and normalizar_texto(valor) not in {"", "N/A", "NO APLICA"}:
                partes.append(valor)
        except Exception:
            continue

    for element_id in campos_input:
        try:
            valor = (driver.find_element(By.ID, element_id).get_attribute("value") or "").strip()
            if valor:
                partes.append(valor)
        except Exception:
            continue

    direccion_compuesta = ' '.join(partes).strip()
    if direccion_compuesta:
        return direccion_compuesta

    campos_respaldo = [
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtBarrioVereda",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtObservacionAdicionalUbicacion",
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtNombreZonaResto",
    ]
    for element_id in campos_respaldo:
        try:
            valor = (driver.find_element(By.ID, element_id).get_attribute("value") or "").strip()
            if valor:
                return valor
        except Exception:
            continue

    return ''


def completar_direccion_residencia(direccion_excel):
    direccion_objetivo = texto_excel(direccion_excel)
    partes = descomponer_direccion_residencia(direccion_excel)
    partes_actuales = leer_partes_direccion_residencia_actual()
    if partes_direccion_coinciden(partes, partes_actuales):
        print("  [=] Dirección residencia: ya está correcta y completa (sin cambios)")
        return True

    if direccion_requiere_modo_libre(partes):
        direccion_actual = leer_direccion_residencia_compuesta()
        if direccion_residencia_coincide(direccion_excel, direccion_actual):
            print("  [=] Dirección residencia: ya coincide como texto completo (sin cambios)")
            return True
        print("  [~] Dirección sin placa deducible o con anexo compacto: se usará dirección libre completa")
        return completar_direccion_residencia_libre(direccion_excel)

    via_valor = texto_excel(partes.get('via'))

    detalle_urbana_ok = asignar_radio_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_rbtnDetalleZonaGeo_1",
        "Detalle zona: Urbana",
        timeout=8.0,
    )
    if not detalle_urbana_ok:
        print("  [~] Detalle zona Urbana bloqueado; se usará campo de dirección principal como último recurso")
        if asignar_valor_input_verificando_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtDireccion",
            direccion_objetivo,
            "Dirección residencia (último recurso)",
            timeout=8.0,
            reintentos=3,
        ):
            return True
        return completar_direccion_residencia_respaldo_zona(direccion_objetivo)
    esperar_postback_finalizado(2)

    if via_valor:
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlVia",
            via_valor,
            "Tipo de vía",
        )
        esperar_postback_finalizado(2)

    asignar_valor_input_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtNombreVia",
        texto_excel(partes.get('nombre_via')),
        "Dirección nombre vía",
        timeout=8.0,
    )
    asignar_select_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlLetra",
        texto_excel(partes.get('letra')) or ' ',
        "Dirección letra",
    )
    asignar_select_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlBis",
        texto_excel(partes.get('bis')) or 'N/A',
        "Dirección bis",
    )
    if texto_excel(partes.get('sentido')):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlSentido",
            texto_excel(partes.get('sentido')),
            "Dirección sentido",
        )
    asignar_valor_input_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtNumero",
        texto_excel(partes.get('numero')),
        "Dirección número",
        timeout=8.0,
    )
    if texto_excel(partes.get('letra3')):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlLetra3",
            texto_excel(partes.get('letra3')),
            "Dirección letra secundaria",
        )
    asignar_select_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlBis2",
        texto_excel(partes.get('bis2')) or 'N/A',
        "Dirección bis secundaria",
    )
    asignar_valor_input_por_id(
        "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtPlaca",
        texto_excel(partes.get('placa')),
        "Dirección placa",
        timeout=8.0,
    )
    if texto_excel(partes.get('sentido2')):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlSentido2",
            texto_excel(partes.get('sentido2')),
            "Dirección sentido secundario",
        )
    if texto_excel(partes.get('complemento')):
        asignar_valor_input_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_txtComplemento",
            texto_excel(partes.get('complemento')),
            "Dirección complemento",
            timeout=8.0,
        )

    direccion_final = leer_direccion_residencia_compuesta()
    if direccion_objetivo and not texto_excel(direccion_final):
        print("  [~] Dirección quedó vacía en bloque principal; usando cuadro de texto bajo zona")
        return completar_direccion_residencia_respaldo_zona(direccion_objetivo)

    return True


def asignar_radio_por_id(element_id, descripcion, timeout=4.0):
    try:
        radio = esperar_control_habilitado(element_id, timeout=timeout)
        if radio is None:
            print(f"  [~] {descripcion}: control no disponible, se omite")
            return False
        if not control_esta_habilitado(radio):
            print(f"  [~] {descripcion}: control bloqueado, se omite")
            return False
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", radio)
        if not radio.is_selected():
            try:
                radio.click()
            except Exception:
                try:
                    ActionChains(driver).move_to_element(radio).click().perform()
                except Exception:
                    driver.execute_script("arguments[0].click();", radio)
            esperar_postback_finalizado(4)
            radio = driver.find_element(By.ID, element_id)
            if not radio.is_selected():
                raise RuntimeError("el radio no quedó seleccionado")
        print(f"  [+] {descripcion}")
        return True
    except Exception as e:
        print(f"  [!] No fue posible ajustar {descripcion}: {e}")
        return False


def asignar_checkbox_por_id(element_id, marcado, descripcion, timeout=4.0):
    try:
        checkbox = esperar_control_habilitado(element_id, timeout=timeout)
        if checkbox is None:
            print(f"  [~] {descripcion}: control no disponible, se omite")
            return False
        if not control_esta_habilitado(checkbox):
            print(f"  [~] {descripcion}: control bloqueado, se omite")
            return False
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", checkbox)
        actual = checkbox.is_selected()
        if actual == marcado:
            print(f"  [=] {descripcion}: ya tiene '{'SI' if marcado else 'NO'}' (sin cambios)")
            return True
        try:
            checkbox.click()
        except Exception:
            try:
                ActionChains(driver).move_to_element(checkbox).click().perform()
            except Exception:
                driver.execute_script("arguments[0].click();", checkbox)
        esperar_postback_finalizado(4)
        checkbox = driver.find_element(By.ID, element_id)
        if checkbox.is_selected() != marcado:
            raise RuntimeError("el checkbox no quedó en el estado esperado")
        print(f"  [+] {descripcion}: {'SI' if marcado else 'NO'}")
        return True
    except Exception as e:
        print(f"  [!] No fue posible ajustar {descripcion}: {e}")
        return False


def excel_tiene_georreferencia(datos_excel):
    latitud = texto_excel(datos_excel.get('latitud_excel'))
    longitud = texto_excel(datos_excel.get('longitud_excel'))
    return not es_valor_ausente(latitud) and not es_valor_ausente(longitud)


def completar_controles_adicionales_ubicacion(datos_excel):
    zona = normalizar_zona_excel(datos_excel.get('zona_residencia', ''))
    centro_poblado = texto_excel(datos_excel.get('centro_poblado') or datos_excel.get('municipio_residencia'))
    tipo_cabecera = texto_excel(datos_excel.get('tipo_cabecera') or 'COMUNA')
    comuna = texto_excel(datos_excel.get('comuna'))
    barrio = texto_excel(datos_excel.get('barrio'))
    barrio_selector = texto_excel(datos_excel.get('barrio_selector') or 'OTRO')
    direccion = texto_excel(datos_excel.get('direccion_residencia'))

    if normalizar_texto(zona) == 'CABECERA':
        if centro_poblado:
            asignar_select_por_id(
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdCentroPoblado",
                centro_poblado,
                "Centro poblado",
            )
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdTipoCabecera",
            tipo_cabecera,
            "Tipo cabecera",
        )
        comuna_asignada = False
        candidatos_comuna = [valor for valor in [comuna, 'SIN INFORMACION'] if valor]
        for candidato_comuna in candidatos_comuna:
            if asignar_select_por_id(
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdComuna",
                candidato_comuna,
                "Comuna/Localidad",
            ):
                comuna_asignada = True
                break
        if not comuna_asignada:
            asignar_select_por_etiquetas_con_candidatos(
                ['Comuna/Localidad', 'Comuna', 'Localidad'],
                candidatos_comuna,
                'Comuna/Localidad',
            )
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdBarrio",
            barrio_selector,
            "Barrio",
        )
        if barrio:
            asignar_valor_input_por_id(
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtObservacionAdicionalUbicacion",
                barrio,
                "Observación adicional ubicación",
            )
        return

    if normalizar_texto(zona) == 'RESTO':
        asignar_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlZonaUbicacion",
            "RESTO",
            "Zona residencia",
            reintentos=3,
        )
        esperar_postback_finalizado(1.2)

        texto_vereda = direccion or barrio
        nombre_zona_resto = texto_vereda if texto_vereda else "OTRO"
        
        # Verificar si ya está correcto antes de asignar
        valor_actual_zona_resto = ""
        try:
            elemento_zona_resto = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtNombreZonaResto")
            valor_actual_zona_resto = (elemento_zona_resto.get_attribute('value') or '').strip()
        except Exception:
            pass
        
        if normalizar_texto(valor_actual_zona_resto) != normalizar_texto(nombre_zona_resto):
            asignar_valor_input_por_id(
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtNombreZonaResto",
                nombre_zona_resto,
                "Nombre Zona Resto",
            ) or asignar_valor_input_por_etiqueta(
                "Nombre Zona Resto",
                nombre_zona_resto,
                "Nombre Zona Resto",
            )
        else:
            print(f"  [=] Nombre Zona Resto: ya tiene '{valor_actual_zona_resto}' (sin cambios)")

        vereda_objetivo = inferir_vereda_desde_texto(texto_vereda)
        vereda_actual = obtener_texto_select_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdVereda"
        )

        if vereda_objetivo:
            if (
                not vereda_actual
                or texto_select_coincide(vereda_actual, 'Seleccione')
                or texto_select_coincide(vereda_actual, 'OTRO')
            ):
                asignar_select_por_id(
                    "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdVereda",
                    vereda_objetivo,
                    "Vereda",
                )
            else:
                print(f"  [=] Vereda: se conserva '{vereda_actual}'")
        else:
            # Si no hay vereda específica, seleccionar "OTRO" para evitar "Seleccione"
            if (
                not vereda_actual
                or texto_select_coincide(vereda_actual, 'Seleccione')
            ):
                asignar_select_por_id(
                    "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdVereda",
                    "OTRO",
                    "Vereda",
                )
            else:
                print(f"  [=] Vereda: se conserva '{vereda_actual}'")

        if texto_vereda:
            asignar_valor_input_por_id(
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtBarrioVereda",
                texto_vereda,
                "Barrio/Vereda",
            ) or asignar_valor_input_por_etiqueta(
                "Barrio/Vereda",
                texto_vereda,
                "Barrio/Vereda",
            ) or asignar_valor_input_por_etiqueta(
                "Barrio",
                texto_vereda,
                "Barrio/Vereda",
            )

            asignar_valor_input_por_id(
                "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtObservacionAdicionalUbicacion",
                texto_vereda,
                "Observación adicional ubicación",
            )


def imprimir_estado_sesion(etiqueta, restaurar_iframe=False):
    if not DIAGNOSTICO_SESION:
        return

    print(f"  [DIAG] {etiqueta}")
    try:
        driver.switch_to.default_content()
        url_actual = driver.current_url
        titulo_actual = driver.title
        login_visible = bool(driver.find_elements(By.ID, "UserName"))
        frame_visible = bool(driver.find_elements(By.ID, "frameContent"))
        print(f"    URL: {url_actual}")
        print(f"    Título: {titulo_actual}")
        print(f"    Login visible: {'SI' if login_visible else 'NO'}")
        print(f"    frameContent visible: {'SI' if frame_visible else 'NO'}")

        if restaurar_iframe and frame_visible:
            iframe = driver.find_element(By.ID, "frameContent")
            driver.switch_to.frame(iframe)
            print("    Contexto restaurado a frameContent")
    except Exception as e:
        print(f"    [DIAG] No fue posible leer el estado de sesión: {e}")


def obtener_primer_elemento_interactivo(xpath):
    for elemento in driver.find_elements(By.XPATH, xpath):
        try:
            if not elemento.is_displayed():
                continue
            rect = elemento.rect or {}
            if rect.get('width', 0) <= 0 or rect.get('height', 0) <= 0:
                continue
            return elemento
        except Exception:
            continue
    return None


def formulario_beneficiario_disponible():
    try:
        driver.switch_to.default_content()
        iframe = driver.find_element(By.ID, "frameContent")
        driver.switch_to.frame(iframe)
        wait.until(EC.presence_of_element_located((By.ID, "btnNuevo")))
        return True
    except Exception:
        return False
    finally:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass


def esperar_hasta(timeout, condicion, descripcion="condición"):
    return WebDriverWait(driver, timeout, poll_frequency=0.2).until(lambda _driver: condicion())


def esperar_documento_listo(timeout=15):
    try:
        WebDriverWait(driver, timeout, poll_frequency=0.2).until(
            lambda _driver: driver.execute_script("return document.readyState") == "complete"
        )
        return True
    except Exception:
        return False


def imprimir_tiempo_etapa(etiqueta, inicio):
    if not DIAGNOSTICO_TIEMPOS_POST_LOGIN:
        return
    duracion = time.perf_counter() - inicio
    print(f"  [TIEMPO] {etiqueta}: {duracion:.2f}s")


def esperar_postback_finalizado(timeout=8.0):
    script_estado = """
    const documentoListo = document.readyState === 'complete';
    const jqueryListo = !window.jQuery || window.jQuery.active === 0;
    let aspnetListo = true;
    try {
        if (window.Sys && Sys.WebForms && Sys.WebForms.PageRequestManager) {
            const manager = Sys.WebForms.PageRequestManager.getInstance();
            aspnetListo = !manager || !manager.get_isInAsyncPostBack();
        }
    } catch (error) {
        aspnetListo = true;
    }
    return documentoListo && jqueryListo && aspnetListo;
    """
    try:
        WebDriverWait(driver, timeout, poll_frequency=0.1).until(
            lambda _driver: driver.execute_script(script_estado)
        )
        return True
    except Exception:
        return False


def esperar_valor_input(element_id, valor_esperado, timeout=4):
    try:
        WebDriverWait(driver, timeout, poll_frequency=0.1).until(
            lambda _driver: (driver.find_element(By.ID, element_id).get_attribute("value") or "").strip() == valor_esperado
        )
        return True
    except Exception:
        return False


def seleccionar_select_si_hace_falta(select_id, texto_objetivo, descripcion, timeout=8, reintentos=3):
    ultimo_error = None
    for intento in range(reintentos):
        try:
            select_elem = wait.until(EC.presence_of_element_located((By.ID, select_id)))
            actual = Select(select_elem).first_selected_option.text.strip()
            if texto_select_coincide(actual, texto_objetivo):
                print(f"  [=] {descripcion}: ya estaba seleccionado")
                return True

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_elem)
            if not seleccionar_por_texto_normalizado(select_elem, texto_objetivo):
                print(f"  [!] No se encontró opción para {descripcion}: {texto_objetivo}")
                return False

            select_refrescado = driver.find_element(By.ID, select_id)
            driver.execute_script(
                """
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                """,
                select_refrescado,
            )
            esperar_postback_finalizado(timeout)
            WebDriverWait(driver, timeout, poll_frequency=0.1).until(
                lambda _driver: texto_select_coincide(obtener_texto_select_por_id(select_id), texto_objetivo)
            )
            print(f"  [+] {descripcion}: {texto_objetivo}")
            return True
        except StaleElementReferenceException as e:
            ultimo_error = e
            esperar_postback_finalizado(2)
            continue
        except Exception as e:
            ultimo_error = e
            break

    if ultimo_error:
        print(f"  [!] Error seleccionando {descripcion}: {ultimo_error}")
    return False


def seleccionar_uds_si_hace_falta(timeout=8, reintentos=3):
    uds_texto_objetivo = CONFIG_EJECUCION['uds_selector'].upper()
    uds_id_objetivo = CONFIG_EJECUCION['uds_id']

    ultimo_error = None
    for intento in range(reintentos):
        try:
            selects = driver.find_elements(By.TAG_NAME, "select")
            for select_elem in selects:
                select_id = select_elem.get_attribute("id") or ""
                if not select_id:
                    continue

                select_actual = driver.find_element(By.ID, select_id)
                options = select_actual.find_elements(By.TAG_NAME, "option")
                coincidencia_texto = ""
                valor_coincidencia = ""
                for opt in options:
                    texto_opcion = (opt.text or "").upper()
                    if uds_id_objetivo in texto_opcion or uds_texto_objetivo in texto_opcion:
                        coincidencia_texto = opt.text.strip()
                        valor_coincidencia = opt.get_attribute("value") or ""
                        break

                if not coincidencia_texto:
                    continue

                actual = Select(select_actual).first_selected_option.text.strip()
                if texto_select_coincide(actual, coincidencia_texto):
                    print("  [=] UDS: ya estaba seleccionada")
                    return True

                if not valor_coincidencia:
                    print("  [!] La opción de UDS no tiene value utilizable")
                    return False

                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_actual)
                Select(select_actual).select_by_value(valor_coincidencia)
                select_refrescado = driver.find_element(By.ID, select_id)
                driver.execute_script(
                    """
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                    """,
                    select_refrescado,
                )
                esperar_postback_finalizado(timeout)
                WebDriverWait(driver, timeout, poll_frequency=0.1).until(
                    lambda _driver: texto_select_coincide(
                        Select(driver.find_element(By.ID, select_id)).first_selected_option.text.strip(),
                        coincidencia_texto,
                    )
                )
                print(f"  [+] UDS seleccionada: {coincidencia_texto}")
                return True
        except StaleElementReferenceException as e:
            ultimo_error = e
            esperar_postback_finalizado(2)
            continue
        except Exception as e:
            ultimo_error = e
            break

    if ultimo_error:
        print(f"  [!] No fue posible seleccionar la UDS: {ultimo_error}")
        return False

    print("  [!] No se encontró la UDS objetivo en el formulario")
    return False


def esperar_elemento_interactivo(xpath, timeout=12):
    try:
        return esperar_hasta(timeout, lambda: obtener_primer_elemento_interactivo(xpath), descripcion=xpath)
    except Exception:
        return None


def click_elemento_interactivo(elemento, espera_post=0.4):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
    try:
        ActionChains(driver).move_to_element(elemento).click().perform()
    except Exception:
        driver.execute_script("arguments[0].click();", elemento)
    if espera_post:
        esperar_postback_finalizado(max(1.0, espera_post))


def tab_esta_activa(elemento_tab):
    try:
        return driver.execute_script(
            """
            const tab = arguments[0];
            if (!tab) return false;
            const aria = (tab.getAttribute('aria-selected') || '').toLowerCase() === 'true';
            const cls = (tab.className || '').toLowerCase();
            const parentCls = ((tab.parentElement && tab.parentElement.className) || '').toLowerCase();
            return aria || cls.includes('active') || parentCls.includes('active');
            """,
            elemento_tab,
        )
    except Exception:
        return False


def tab_pertenencia_etnica_lista(timeout=2.0):
    fin = time.time() + timeout
    while time.time() < fin:
        esperar_postback_finalizado(0.8)
        try:
            elementos = driver.find_elements(
                By.XPATH,
                "//div[contains(@id, 'tbnPertenenciaEtnica')]//select | "
                "//select[contains(@id, 'ddlIdGrupoEtnico') or contains(@id, 'PertenenciaEtnica')] | "
                "//label[contains(translate(normalize-space(.), 'ÁÉÍÓÚ', 'AEIOU'), 'Grupo etnico')] | "
                "//*[contains(translate(normalize-space(.), 'ÁÉÍÓÚ', 'AEIOU'), 'Grupo etnico al que pertenece el beneficiario')]"
            )
            if any(elemento.is_displayed() for elemento in elementos):
                return True
        except Exception:
            pass
        time.sleep(0.1)
    return False


def tab_ubicacion_lista(timeout=2.0):
    fin = time.time() + timeout
    while time.time() < fin:
        esperar_postback_finalizado(0.8)
        try:
            elementos = driver.find_elements(By.XPATH, "//input[contains(@id, 'txtTelefono') or contains(@id, 'txtDireccion') or contains(@id, 'txtBarrioVereda')]")
            if any(elemento.is_displayed() for elemento in elementos):
                return True
        except Exception:
            pass
        time.sleep(0.1)
    return False


def abrir_tab_con_espera(nombre_tab, selectores_tab, funcion_listo, cache_key, timeout_postback=1.5, timeout_listo=2.0):
    global TAB_ACTIVA
    print(f"  [*] Abriendo pestaña {nombre_tab}...")

    if TAB_ACTIVA == cache_key and funcion_listo(timeout=min(timeout_listo, 1.0)):
        print(f"  [=] Pestaña {nombre_tab}: ya estaba lista")
        return True

    for selector in selectores_tab:
        try:
            tab = wait.until(EC.presence_of_element_located((selector[0], selector[1])))
            if tab_esta_activa(tab) and funcion_listo(timeout=min(timeout_listo, 1.0)):
                TAB_ACTIVA = cache_key
                print(f"  [=] Pestaña {nombre_tab}: ya estaba activa")
                return True

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
            try:
                tab.click()
            except Exception:
                driver.execute_script("arguments[0].click();", tab)

            esperar_postback_finalizado(timeout_postback)
            if funcion_listo(timeout_listo):
                TAB_ACTIVA = cache_key
                print(f"  [+] Pestaña {nombre_tab} abierta")
                return True
        except Exception:
            pass

    print(f"  [!] No se pudo abrir la pestaña {nombre_tab}")
    return False


def obtener_url_directa_beneficiario():
    url_por_defecto = "https://rubonline.icbf.gov.co/Page/RUBONLINE/BENEFICIARIO/List.aspx"
    try:
        driver.switch_to.default_content()
        enlaces = driver.find_elements(By.XPATH, "//a[contains(@href, '/Page/RUBONLINE/BENEFICIARIO/List.aspx')]")
        enlace = enlaces[0] if enlaces else None
        if not enlace:
            return url_por_defecto
        href = (enlace.get_attribute("href") or "").strip()
        if not href:
            return url_por_defecto
        return urljoin(driver.current_url, href)
    except Exception:
        return url_por_defecto


def cargar_beneficiario_directo_en_iframe():
    url_beneficiario = obtener_url_directa_beneficiario()
    if not url_beneficiario:
        return False

    try:
        driver.switch_to.default_content()
        iframe = wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
        driver.execute_script("arguments[0].src = arguments[1];", iframe, url_beneficiario)
        WebDriverWait(driver, 4, poll_frequency=0.2).until(
            lambda _driver: "BENEFICIARIO/List.aspx" in ((iframe.get_attribute("src") or "").upper())
        )
        driver.switch_to.frame(iframe)
        WebDriverWait(driver, 8, poll_frequency=0.2).until(
            EC.presence_of_element_located((By.ID, "btnNuevo"))
        )
        driver.switch_to.default_content()
        print(f"  [+] Formulario Beneficiario cargado directo en iframe: {url_beneficiario}")
        return True
    except Exception:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        return False


def login_o_menu_rub_disponible():
    try:
        if formulario_beneficiario_disponible():
            return True
    except Exception:
        pass

    try:
        driver.switch_to.default_content()
        if obtener_primer_elemento_interactivo("//a[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'rub online')]"):
            return True
    except Exception:
        return False
    return False

# ========== PASO 0: LEER EXCEL ==========
print("\n[*] Leyendo datos del Excel...")

ruta_excel = str(CONFIG_EJECUCION['ruta_excel'])
rutas_fotos_unidad = [Path(ruta) for ruta in CONFIG_EJECUCION.get('rutas_fotos', [])]
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

registros_unidad = []
registros_omitidos_fecha = []

for row_idx in range(3, ws.max_row + 1):
    nombre_uds = ws.cell(row=row_idx, column=columna_excel(3)).value
    
    if pertenece_a_segmento_excel(nombre_uds, CONFIG_EJECUCION['filtro_excel']):
        fecha_nacimiento_raw = ws.cell(row=row_idx, column=columna_excel(13)).value
        fecha_ingreso_raw = ws.cell(row=row_idx, column=columna_excel(4)).value
        datos = {
            'fila': row_idx,
            'uds_excel': texto_excel(nombre_uds),
            'tipo_documento_beneficiario': texto_excel(ws.cell(row=row_idx, column=columna_excel(16)).value),
            'documento': texto_excel(ws.cell(row=row_idx, column=columna_excel(17)).value),
            'primer_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(5)).value),
            'segundo_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(6)).value),
            'primer_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(7)).value),
            'segundo_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(8)).value),
            'sexo': texto_excel(ws.cell(row=row_idx, column=columna_excel(9)).value),
            'nacionalidad_beneficiario': texto_excel(ws.cell(row=row_idx, column=columna_excel(10)).value),
            'pais_nacimiento_beneficiario': texto_excel(ws.cell(row=row_idx, column=columna_excel(10)).value),
            'departamento_nacimiento_beneficiario': texto_excel(ws.cell(row=row_idx, column=columna_excel(11)).value),
            'municipio_nacimiento_beneficiario': texto_excel(ws.cell(row=row_idx, column=columna_excel(12)).value),
            'fecha_nacimiento_beneficiario': fecha_excel(fecha_nacimiento_raw),
            'fecha_ingreso': fecha_ingreso_raw,
            'pais_residencia': texto_excel(ws.cell(row=row_idx, column=columna_excel(20)).value),
            'departamento_residencia': texto_excel(ws.cell(row=row_idx, column=columna_excel(21)).value),
            'municipio_residencia': texto_excel(ws.cell(row=row_idx, column=columna_excel(22)).value),
            'zona_residencia': texto_excel(ws.cell(row=row_idx, column=columna_excel(23)).value),
            'barrio': texto_excel(ws.cell(row=row_idx, column=columna_excel(24)).value),
            'direccion_residencia': texto_excel(ws.cell(row=row_idx, column=columna_excel(25)).value),
            'latitud_excel': texto_excel(ws.cell(row=row_idx, column=columna_excel(26)).value),
            'longitud_excel': texto_excel(ws.cell(row=row_idx, column=columna_excel(27)).value),
            'telefono_original_excel': texto_excel(ws.cell(row=row_idx, column=columna_excel(28)).value),
            'telefono': normalizar_telefono_excel(ws.cell(row=row_idx, column=columna_excel(28)).value),
            'territorio_etnico': texto_excel(ws.cell(row=row_idx, column=columna_excel(29)).value),
            'discapacidad_excel': texto_excel(ws.cell(row=row_idx, column=columna_excel(30)).value),
            'responsable_parentesco': texto_excel(ws.cell(row=row_idx, column=columna_excel(61)).value),
            'responsable_doc_tipo': texto_excel(ws.cell(row=row_idx, column=columna_excel(62)).value),
            'responsable_documento': texto_excel(ws.cell(row=row_idx, column=columna_excel(63)).value),
            'responsable_nombre': nombre_completo(
                ws.cell(row=row_idx, column=columna_excel(64)).value,
                ws.cell(row=row_idx, column=columna_excel(65)).value,
                ws.cell(row=row_idx, column=columna_excel(66)).value,
                ws.cell(row=row_idx, column=columna_excel(67)).value,
            ),
            'responsable_primer_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(64)).value),
            'responsable_segundo_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(65)).value),
            'responsable_primer_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(66)).value),
            'responsable_segundo_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(67)).value),
            'responsable_fecha_nacimiento': fecha_excel(ws.cell(row=row_idx, column=columna_excel(68)).value),
            'responsable_departamento_nacimiento': texto_excel(ws.cell(row=row_idx, column=columna_excel(69)).value),
            'responsable_municipio_nacimiento': texto_excel(ws.cell(row=row_idx, column=columna_excel(70)).value),
            'padre_doc_tipo': texto_excel(ws.cell(row=row_idx, column=columna_excel(71)).value),
            'padre_documento': texto_excel(ws.cell(row=row_idx, column=columna_excel(72)).value),
            'padre_nombre': nombre_completo(
                ws.cell(row=row_idx, column=columna_excel(74)).value,
                ws.cell(row=row_idx, column=columna_excel(75)).value,
                ws.cell(row=row_idx, column=columna_excel(76)).value,
                ws.cell(row=row_idx, column=columna_excel(77)).value,
            ),
            'padre_primer_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(74)).value),
            'padre_segundo_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(75)).value),
            'padre_primer_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(76)).value),
            'padre_segundo_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(77)).value),
            'padre_fecha_nacimiento': fecha_excel(ws.cell(row=row_idx, column=columna_excel(78)).value),
            'padre_departamento_nacimiento': texto_excel(ws.cell(row=row_idx, column=columna_excel(79)).value),
            'padre_municipio_nacimiento': texto_excel(ws.cell(row=row_idx, column=columna_excel(80)).value),
            'madre_doc_tipo': texto_excel(ws.cell(row=row_idx, column=columna_excel(81)).value),
            'madre_documento': texto_excel(ws.cell(row=row_idx, column=columna_excel(82)).value),
            'madre_nombre': nombre_completo(
                ws.cell(row=row_idx, column=columna_excel(83)).value,
                ws.cell(row=row_idx, column=columna_excel(84)).value,
                ws.cell(row=row_idx, column=columna_excel(85)).value,
                ws.cell(row=row_idx, column=columna_excel(86)).value,
            ),
            'madre_primer_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(83)).value),
            'madre_segundo_nombre': texto_excel(ws.cell(row=row_idx, column=columna_excel(84)).value),
            'madre_primer_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(85)).value),
            'madre_segundo_apellido': texto_excel(ws.cell(row=row_idx, column=columna_excel(86)).value),
            'madre_fecha_nacimiento': fecha_excel(ws.cell(row=row_idx, column=columna_excel(87)).value),
            'madre_departamento_nacimiento': texto_excel(ws.cell(row=row_idx, column=columna_excel(88)).value),
            'madre_municipio_nacimiento': texto_excel(ws.cell(row=row_idx, column=columna_excel(89)).value),
            'jefe_hogar': texto_excel(ws.cell(row=row_idx, column=columna_excel(90)).value),
            'padre_convive': texto_excel(ws.cell(row=row_idx, column=columna_excel(91)).value),
            'madre_convive': texto_excel(ws.cell(row=row_idx, column=columna_excel(92)).value),
        }

        fecha_nacimiento = fecha_comparable(fecha_nacimiento_raw)
        fecha_ingreso = fecha_comparable(fecha_ingreso_raw)
        if fecha_nacimiento and fecha_ingreso and fecha_ingreso < fecha_nacimiento:
            registros_omitidos_fecha.append({
                'documento': datos['documento'],
                'fila': row_idx,
                'fecha_ingreso': fecha_excel(fecha_ingreso_raw),
                'fecha_nacimiento': fecha_excel(fecha_nacimiento_raw),
            })
            print(
                f"[~] Omitiendo fila {row_idx} ({datos['documento']}): fecha de ingreso {fecha_excel(fecha_ingreso_raw)} "
                f"menor a fecha de nacimiento {fecha_excel(fecha_nacimiento_raw)}"
            )
            continue
        registros_unidad.append(datos)

print(f"[+] Total registros encontrados: {len(registros_unidad)}")
if registros_omitidos_fecha:
    print(f"[~] Registros omitidos por fechas inválidas: {len(registros_omitidos_fecha)}")
    for registro_omitido in registros_omitidos_fecha[:10]:
        print(
            f"    - {registro_omitido['documento']} (ingreso {registro_omitido['fecha_ingreso']} < nacimiento {registro_omitido['fecha_nacimiento']})"
        )
print(f"[*] Filtro Excel activo: {descripcion_filtro_excel(CONFIG_EJECUCION['filtro_excel'])}")
for reg in registros_unidad[:5]:
    print(f"    - {reg['documento']}: {reg['primer_nombre']} {reg['primer_apellido']}")

# ========== PASO 1: LOGIN ==========
print("\n[*] Abriendo navegador...")
options = webdriver.ChromeOptions()
options.add_argument('--disable-notifications')
options.add_argument('--disable-popup-blocking')
options.add_experimental_option('detach', True)
options.page_load_strategy = 'eager'

CHROMEDRIVER_PATH = "/Users/stevenruiz/.wdm/drivers/chromedriver/mac64/146.0.7680.165/chromedriver-mac-arm64/chromedriver"
try:
    driver = webdriver.Chrome(
        service=Service(CHROMEDRIVER_PATH),
        options=options
    )
except Exception:
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
wait = WebDriverWait(driver, 20)
driver.get("https://rubonline.icbf.gov.co/DefaultF.aspx")
esperar_documento_listo()

print("[*] Haciendo login...")

# Login con credenciales conocidas
try:
    campo_usuario = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "UserName")))
    campo_usuario.clear()
    campo_usuario.send_keys("angie.cardenas")

    campo_password = driver.find_element(By.ID, "Password")
    campo_password.clear()
    campo_password.send_keys("Celeste1020*")

    boton_login = driver.find_element(By.ID, "LoginButton")
    boton_login.click()
    print("[+] Login completado")
    esperar_documento_listo()
    esperar_hasta(15, login_o_menu_rub_disponible, "menu de RUB o formulario")
    imprimir_estado_sesion("Estado después del login")
except Exception as e:
    print(f"[!] Error al hacer login: {e}")
    driver.quit()
    raise


def navegar_a_formulario():
    """Replica la navegación funcional del bot principal."""
    print("  [*] Navegando al formulario...")

    try:
        imprimir_estado_sesion("Antes de navegar al formulario")
        if formulario_beneficiario_disponible():
            print("  [+] Formulario de Beneficiario ya disponible")
            return True

        if cargar_beneficiario_directo_en_iframe():
            return True

        enlaces_rub = esperar_elemento_interactivo("//a[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'rub online')]", timeout=6)
        if enlaces_rub:
            click_elemento_interactivo(enlaces_rub, espera_post=ESPERA_POST_CLICK_MENU)
            esperar_documento_listo(6)

        if cargar_beneficiario_directo_en_iframe():
            return True

        if formulario_beneficiario_disponible():
            print("  [+] Formulario de Beneficiario ya disponible tras entrar a RUB")
            return True

        enlace_listado = esperar_elemento_interactivo("//a[contains(text(), 'Beneficiario') and contains(@href, 'BENEFICIARIO') and contains(@href, 'List.aspx')]", timeout=4)
        if enlace_listado:
            click_elemento_interactivo(enlace_listado, espera_post=ESPERA_POST_CLICK_MENU)
            esperar_documento_listo(6)
        else:
            enlace_menu = esperar_elemento_interactivo("//a[contains(text(), 'Beneficiario')]", timeout=4)
            if enlace_menu:
                click_elemento_interactivo(enlace_menu, espera_post=ESPERA_POST_CLICK_MENU)
                esperar_documento_listo(6)

            enlace_listado = esperar_elemento_interactivo("//a[contains(text(), 'Beneficiario') and contains(@href, 'BENEFICIARIO') and contains(@href, 'List.aspx')]", timeout=4)
            if not enlace_listado:
                raise RuntimeError("No se encontró el enlace interactivo al listado de Beneficiario")
            click_elemento_interactivo(enlace_listado, espera_post=ESPERA_POST_CLICK_MENU)
            esperar_documento_listo(6)

        esperar_hasta(8, formulario_beneficiario_disponible, "formulario de beneficiario")

        print("  [+] Navegación completada")
        imprimir_estado_sesion("Después de navegar al formulario")
        return True
    except Exception as e:
        print(f"  [!] Error navegando al formulario: {e}")
        imprimir_estado_sesion("Error durante navegación al formulario")
        return False


def preparar_formulario_busqueda(seleccionar_tipo_beneficiario=True):
    """Entra al iframe, pulsa nuevo y aplica filtros requeridos antes de buscar."""
    print("  [*] Preparando formulario...")
    inicio_preparacion = time.perf_counter()

    try:
        imprimir_estado_sesion("Antes de entrar al iframe")
        iframe = wait.until(EC.presence_of_element_located((By.ID, "frameContent")))
        driver.switch_to.frame(iframe)
        imprimir_estado_sesion("Después de entrar al iframe", restaurar_iframe=True)

        boton_nuevo = wait.until(EC.element_to_be_clickable((By.ID, "btnNuevo")))
        inicio_nuevo = time.perf_counter()
        click_elemento_interactivo(boton_nuevo, espera_post=0)
        print("  [+] Botón '+' clickeado")
        wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDireccionesICBF")))
        esperar_postback_finalizado(8)
        imprimir_tiempo_etapa("Abrir '+ nuevo'", inicio_nuevo)
        imprimir_estado_sesion("Después de hacer click en '+'", restaurar_iframe=True)

        if not aplicar_filtros_formulario(seleccionar_tipo_beneficiario=seleccionar_tipo_beneficiario):
            return False

        print("  [+] Formulario listo")
        imprimir_tiempo_etapa("Preparar formulario", inicio_preparacion)
        imprimir_estado_sesion("Después de preparar formulario", restaurar_iframe=True)
        return True
    except Exception as e:
        print(f"  [!] Error preparando formulario: {e}")
        imprimir_estado_sesion("Error preparando formulario")
        return False


def seleccionar_tipo_beneficiario_formulario():
    inicio_tipo = time.perf_counter()
    try:
        driver.execute_script("window.scrollBy(0, 300);")
        ok = seleccionar_select_si_hace_falta(
            "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdTipoBeneficiario",
            "NIÑO O NIÑA ENTRE 6 MESES",
            "Tipo de beneficiario",
        )
        if ok:
            imprimir_tiempo_etapa("Seleccionar tipo de beneficiario", inicio_tipo)
        return ok
    except Exception as e:
        print(f"  [!] No fue posible seleccionar tipo de beneficiario: {e}")
        return False


def aplicar_filtros_formulario(seleccionar_tipo_beneficiario=True):
    """Aplica los filtros de encabezado requeridos y deja listo el bloque de datos básicos."""
    print("  [*] Aplicando datos de contrato y encabezado...")
    inicio_filtros = time.perf_counter()

    try:
        imprimir_estado_sesion("Antes de aplicar filtros", restaurar_iframe=True)

        # Filtros mínimos necesarios para habilitar el formulario.
        try:
            radios = driver.find_elements(By.XPATH, "//input[@type='radio']")
            for radio in radios:
                valor = radio.get_attribute('value') or ''
                if 'Uno a uno' in valor:
                    if not radio.is_selected():
                        inicio_radio = time.perf_counter()
                        radio.click()
                        esperar_postback_finalizado(6)
                        imprimir_tiempo_etapa("Seleccionar modo 'Uno a uno'", inicio_radio)
                    else:
                        print("  [=] Modo 'Uno a uno': ya estaba seleccionado")
                    break
        except Exception:
            pass

        filtros_select = [
            ("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDireccionesICBF", "Dirección de Primera Infancia", "Dirección ICBF"),
            ("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlRegional", "Boyacá", "Regional"),
            ("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdVigencia", "2026", "Vigencia"),
            ("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNumeroContrato", "OD 15 420272 00015 2026", "Contrato"),
            ("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlNombreServicio", "EDUCACIÓN INICIAL EN EL HOGAR - FAMILIAR Y COMUNITARIA - 420272-2026", "Servicio"),
            ("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdTipoDocumento", "REGISTRO CIVIL", "Tipo documento"),
        ]

        for select_id, texto, descripcion in filtros_select:
            try:
                inicio_select = time.perf_counter()
                if seleccionar_select_si_hace_falta(select_id, texto, descripcion):
                    imprimir_tiempo_etapa(f"Filtro {descripcion}", inicio_select)
            except Exception:
                pass

        inicio_uds = time.perf_counter()
        if seleccionar_uds_si_hace_falta():
            imprimir_tiempo_etapa("Filtro UDS", inicio_uds)

        if seleccionar_tipo_beneficiario:
            seleccionar_tipo_beneficiario_formulario()

        esperar_postback_finalizado(6)
        imprimir_tiempo_etapa("Aplicar filtros", inicio_filtros)
        imprimir_estado_sesion("Después de aplicar filtros", restaurar_iframe=True)
        return True
    except Exception as e:
        print(f"  [!] Error aplicando filtros del formulario: {e}")
        imprimir_estado_sesion("Error aplicando filtros", restaurar_iframe=True)
        return False


def buscar_y_extraer_datos(documento):
    """Busca el documento y devuelve los valores cargados en el formulario."""
    print("  [*] Ejecutando búsqueda en formulario...")
    inicio_busqueda = time.perf_counter()
    imprimir_estado_sesion("Antes de ingresar documento", restaurar_iframe=True)
    ultimo_error = None
    for _ in range(3):
        try:
            campo_doc = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdentificacion")))
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_doc)
            campo_doc.click()
            campo_doc.clear()
            campo_doc.send_keys(documento)
            esperar_valor_input("cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtIdentificacion", documento, timeout=3)
            print(f"    [+] Documento ingresado: {documento}")

            lupa = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnBuscar")))
            driver.execute_script("arguments[0].scrollIntoView(true);", lupa)
            click_elemento_interactivo(lupa, espera_post=0)
            print("    [+] Click en lupa realizado")
            esperar_postback_finalizado(max(4, ESPERA_BUSQUEDA + 1))
            wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtPrimerNombre")))
            imprimir_tiempo_etapa(f"Búsqueda documento {documento}", inicio_busqueda)
            imprimir_estado_sesion(f"Después de buscar documento {documento}", restaurar_iframe=True)
            return leer_datos_basicos_formulario()
        except StaleElementReferenceException as e:
            ultimo_error = e
            esperar_postback_finalizado(2)
            continue
        except Exception as e:
            ultimo_error = e
            break

    if ultimo_error:
        raise ultimo_error

    return leer_datos_basicos_formulario()


def leer_datos_basicos_formulario():
    for _ in range(3):
        try:
            datos_formulario = {}
            campos = {
                'primer_nombre': "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtPrimerNombre",
                'segundo_nombre': "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtSegundoNombre",
                'primer_apellido': "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtPrimerApellido",
                'segundo_apellido': "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtSegundoApellido",
            }

            for llave, campo_id in campos.items():
                try:
                    datos_formulario[llave] = driver.find_element(By.ID, campo_id).get_attribute("value") or ""
                except Exception:
                    datos_formulario[llave] = ""

            try:
                select_sexo = Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdSexo"))
                datos_formulario['sexo'] = select_sexo.first_selected_option.text.strip()
            except Exception:
                datos_formulario['sexo'] = ""

            try:
                datos_formulario['fecha_atencion'] = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_cuwFechaAtencion_txtFecha").get_attribute("value") or ""
            except Exception:
                datos_formulario['fecha_atencion'] = ""

            try:
                select_discapacidad = Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlPresentaDiscapacidad"))
                datos_formulario['discapacidad'] = select_discapacidad.first_selected_option.text.strip()
            except Exception:
                datos_formulario['discapacidad'] = ""

            return datos_formulario
        except StaleElementReferenceException:
            esperar_postback_finalizado(2)
            continue

    return {
        'primer_nombre': "",
        'segundo_nombre': "",
        'primer_apellido': "",
        'segundo_apellido': "",
        'sexo': "",
        'fecha_atencion': "",
        'discapacidad': "",
    }


def formulario_basico_sin_informacion(datos_formulario):
    texto_basico = "".join([
        texto_excel(datos_formulario.get('primer_nombre', '')),
        texto_excel(datos_formulario.get('segundo_nombre', '')),
        texto_excel(datos_formulario.get('primer_apellido', '')),
        texto_excel(datos_formulario.get('segundo_apellido', '')),
    ]).strip()
    sexo = normalizar_texto(datos_formulario.get('sexo', ''))
    return (not texto_basico) and sexo in {'', 'SELECCIONE'}


def llenar_datos_basicos_beneficiario_desde_excel(datos_excel):
    print("  [*] No se encontró información tras la lupa; creando beneficiario desde Excel...")

    def asignar_input(ids, etiqueta, valor, descripcion):
        if not valor:
            return False
        for element_id in ids:
            if asignar_valor_input_por_id(element_id, valor, descripcion):
                return True
        return asignar_valor_input_por_etiqueta(etiqueta, valor, descripcion)

    def asignar_select(ids, etiqueta, valor, descripcion, esperar_valor=None):
        if not valor:
            return False
        for element_id in ids:
            if esperar_valor:
                esperar_opcion_en_select(element_id, esperar_valor)
            if asignar_select_por_id(element_id, valor, descripcion):
                return True
        return asignar_select_por_etiqueta(etiqueta, valor, descripcion)

    def asignar_select_cascada(ids, etiqueta, valor, descripcion, esperar_id_siguiente=None, esperar_valor_siguiente=None):
        if not valor:
            return False
        asignado = asignar_select(ids, etiqueta, valor, descripcion)
        if not asignado:
            return False
        if esperar_id_siguiente and esperar_valor_siguiente:
            print(f"  [*] Esperando habilitación de {descripcion.lower()} -> siguiente selector...")
            time.sleep(1)
            esperar_opcion_en_select(esperar_id_siguiente, esperar_valor_siguiente, timeout=12)
        return True

    def asignar_select_cascada_por_id(element_id, etiqueta, valor, descripcion, esperar_id_siguiente=None, esperar_valor_siguiente=None):
        if not valor:
            return False
        for intento in range(5):
            asignado = asignar_select_por_id(element_id, valor, descripcion, reintentos=3)
            if not asignado:
                try:
                    asignado = asignar_select_por_etiqueta(etiqueta, valor, descripcion)
                except Exception:
                    asignado = False
            if asignado:
                if esperar_id_siguiente and esperar_valor_siguiente:
                    print(f"  [*] Esperando habilitación de {descripcion.lower()} -> siguiente selector...")
                    time.sleep(1)
                    esperar_opcion_en_select(esperar_id_siguiente, esperar_valor_siguiente, timeout=12)
                return True
            print(f"  [~] Reintentando {descripcion.lower()} ({intento + 1}/5)...")
            time.sleep(1.5)
            esperar_documento_listo(timeout=5)
        return False

    print(f"  [=] Documento beneficiario: se conserva el valor buscado {datos_excel.get('documento', '')}")
    asignar_input(
        ["cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtPrimerNombre"],
        "Primer Nombre",
        datos_excel.get('primer_nombre', ''),
        "Primer nombre beneficiario",
    )
    asignar_input(
        ["cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtSegundoNombre"],
        "Segundo Nombre",
        datos_excel.get('segundo_nombre', ''),
        "Segundo nombre beneficiario",
    )
    asignar_input(
        ["cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtPrimerApellido"],
        "Primer Apellido",
        datos_excel.get('primer_apellido', ''),
        "Primer apellido beneficiario",
    )
    asignar_input(
        ["cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_txtSegundoApellido"],
        "Segundo Apellido",
        datos_excel.get('segundo_apellido', ''),
        "Segundo apellido beneficiario",
    )
    asignar_input(
        [
            "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_cuwFechaNacimiento_txtFecha",
            "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_cuwFechaNacimiento_txtFecha",
        ],
        "Fecha de Nacimiento",
        datos_excel.get('fecha_nacimiento_beneficiario', ''),
        "Fecha nacimiento beneficiario",
    )
    asignar_select(
        ["cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdSexo"],
        "Sexo",
        'Hombre' if normalizar_sexo(datos_excel.get('sexo', '')) == 'MASCULINO' else 'Mujer',
        "Sexo beneficiario",
    )

    foto_cargada = cargar_foto(datos_excel['documento'])
    print(f"  [*] Foto de alta nueva: {'OK' if foto_cargada else 'NO'}")

    asignar_select_cascada_por_id(
        "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdPaisNacimiento",
        "Pais nacimiento",
        datos_excel.get('pais_nacimiento_beneficiario', '') or 'COLOMBIA',
        "País nacimiento beneficiario",
        esperar_id_siguiente="cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDepartamento",
        esperar_valor_siguiente=datos_excel.get('departamento_nacimiento_beneficiario', ''),
    )
    asignar_select_cascada_por_id(
        "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdDepartamento",
        "Departamento nacimiento",
        datos_excel.get('departamento_nacimiento_beneficiario', ''),
        "Departamento nacimiento beneficiario",
        esperar_id_siguiente="cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_ddlIdMunicipioNacimiento",
        esperar_valor_siguiente=datos_excel.get('municipio_nacimiento_beneficiario', ''),
    )
    asignar_select_cascada_por_id(
        "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlIdMunicipioNacimiento",
        "Municipio nacimiento",
        datos_excel.get('municipio_nacimiento_beneficiario', ''),
        "Municipio nacimiento beneficiario",
    )

    completar_campos_faltantes(datos_excel)

    datos_releidos = leer_datos_basicos_formulario()
    creado_ok = all([
        normalizar_texto(datos_releidos.get('primer_nombre', '')) == normalizar_texto(datos_excel.get('primer_nombre', '')),
        normalizar_texto(datos_releidos.get('primer_apellido', '')) == normalizar_texto(datos_excel.get('primer_apellido', '')),
        normalizar_sexo(datos_releidos.get('sexo', '')) == normalizar_sexo(datos_excel.get('sexo', '')),
    ])
    print(f"  [{'+' if creado_ok else '!'}] Alta básica desde Excel: {'OK' if creado_ok else 'REVISAR'}")
    return {
        'creado_ok': creado_ok,
        'foto_cargada': foto_cargada,
    }


def completar_campos_faltantes(datos_excel):
    """Completa fecha de atencion y discapacidad despues de la consulta."""
    print("  [*] Completando campos faltantes del formulario...")

    fecha_objetivo = ""
    if datos_excel.get('fecha_ingreso'):
        fecha_objetivo = datos_excel['fecha_ingreso'].strftime("%d/%m/%Y")

    try:
        fecha_asignada = False
        for _ in range(5):
            try:
                campo_fecha_atencion = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_cuwFechaAtencion_txtFecha")))
                driver.execute_script("arguments[0].scrollIntoView(true);", campo_fecha_atencion)
                time.sleep(0.3)
                fecha_actual = (campo_fecha_atencion.get_attribute("value") or "").strip()
                if not fecha_actual and fecha_objetivo:
                    driver.execute_script(
                        """
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                        """,
                        campo_fecha_atencion,
                        fecha_objetivo,
                    )
                    time.sleep(1)
                    campo_fecha_atencion = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_cuwFechaAtencion_txtFecha")))
                    fecha_actual = (campo_fecha_atencion.get_attribute("value") or "").strip()
                if fecha_actual:
                    print(f"    [+] Fecha de atencion: {fecha_actual}")
                    fecha_asignada = True
                    break
                time.sleep(1)
            except Exception:
                time.sleep(1)

        if not fecha_asignada:
            print("    [!] No fue posible completar fecha de atencion")
    except Exception as e:
        print(f"    [!] Error completando fecha de atencion: {e}")

    try:
        select_discapacidad = Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlPresentaDiscapacidad"))
        discapacidad_objetivo = 'Sí' if normalizar_respuesta_si_no(datos_excel.get('discapacidad_excel', 'NO')) == 'SI' else 'No'
        if not texto_select_coincide(select_discapacidad.first_selected_option.text.strip(), discapacidad_objetivo):
            select_discapacidad.select_by_visible_text(discapacidad_objetivo)
        print(f"    [+] Discapacidad: {discapacidad_objetivo}")
    except Exception as e:
        print(f"    [!] No fue posible completar discapacidad: {e}")

    time.sleep(1)


def buscar_foto_por_documento(documento):
    for ruta_fotos in rutas_fotos_unidad:
        coincidencias = sorted(ruta_fotos.glob(f"*{documento}*"))
        if coincidencias:
            return str(coincidencias[0])
    return None


def cargar_foto(documento):
    """Carga la foto correspondiente al documento antes de guardar."""
    print("  [*] Cargando foto...")

    ruta_foto = buscar_foto_por_documento(documento)
    if not ruta_foto:
        print(f"  [!] No se encontró foto para el documento {documento}")
        return False

    try:
        input_foto = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_FileUploadControl")))
        driver.execute_script("arguments[0].scrollIntoView(true);", input_foto)
        time.sleep(0.5)
        input_foto.send_keys(ruta_foto)
        print(f"  [+] Foto seleccionada: {Path(ruta_foto).name}")
        time.sleep(1)

        boton_cargar = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_datosBasicosPersona_btnCargar")))
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_cargar)
        time.sleep(0.5)
        boton_cargar.click()
        print("  [+] Botón 'Cargar foto' pulsado")
        print(f"  [*] Esperando {ESPERA_CARGA_FOTO} segundos a que termine la carga de la foto...")
        time.sleep(ESPERA_CARGA_FOTO)
        return True
    except Exception as e:
        print(f"  [!] Error cargando foto: {e}")
        return False


def imprimir_campos_con_valor():
    """Muestra los controles que quedaron con contenido después de la consulta."""
    if not DEBUG_CAMPOS_CON_VALOR:
        return

    print("\n  [DEBUG CAMPOS CON VALOR]:")

    try:
        inputs = driver.find_elements(By.XPATH, "//input")
        for elemento in inputs:
            try:
                valor = (elemento.get_attribute("value") or "").strip()
                elemento_id = elemento.get_attribute("id") or ""
                nombre = elemento.get_attribute("name") or ""
                tipo = elemento.get_attribute("type") or ""
                if valor and tipo not in ["hidden", "password"]:
                    print(f"    INPUT id={elemento_id} name={nombre} type={tipo} value={valor}")
            except StaleElementReferenceException:
                continue
    except Exception as e:
        print(f"    [!] No fue posible listar inputs: {e}")

    try:
        selects = driver.find_elements(By.TAG_NAME, "select")
        for select_elem in selects:
            try:
                select_id = select_elem.get_attribute("id") or ""
                nombre = select_elem.get_attribute("name") or ""
                seleccionado = Select(select_elem).first_selected_option.text.strip()
                if seleccionado:
                    print(f"    SELECT id={select_id} name={nombre} selected={seleccionado}")
            except StaleElementReferenceException:
                continue
    except Exception as e:
        print(f"    [!] No fue posible listar selects: {e}")


def normalizar_sexo(valor):
    texto = (valor or "").strip().upper()
    equivalencias = {
        "HOMBRE": "MASCULINO",
        "MUJER": "FEMENINO",
    }
    return equivalencias.get(texto, texto)


def normalizar_respuesta_si_no(valor, default='NO'):
    texto = normalizar_texto(valor)
    if texto in {'SI', 'S', 'YES', 'TRUE', '1'}:
        return 'SI'
    if texto in {'NO', 'N', 'FALSE', '0'}:
        return 'NO'
    return default


def normalizar_texto(valor):
    texto = " ".join((valor or "").strip().upper().split())
    texto = unicodedata.normalize('NFD', texto)
    return ''.join(ch for ch in texto if unicodedata.category(ch) != 'Mn')


def es_valor_ausente(valor):
    texto = normalizar_texto(valor)
    if texto in {"", "N/A", "NA", "NO", "N0", "NO APLICA", "NO APLICA.", "NINGUNO", "NINGUNA", "NONE"}:
        return True
    # Detecta nombres compuestos de puras celdas vacías p.ej. "N/A N/A N/A N/A"
    tokens_ausentes = {"", "N/A", "NA", "NO", "N0", "NO APLICA", "NINGUNO", "NINGUNA", "NONE"}
    tokens = [t.strip() for t in texto.split()]
    if tokens and all(t in tokens_ausentes for t in tokens):
        return True
    return False


def valor_excel_familia(valor):
    return "" if es_valor_ausente(valor) else (valor or "")


DEPARTAMENTOS_COLOMBIA = {
    'AMAZONAS', 'ANTIOQUIA', 'ARAUCA', 'ATLANTICO', 'BOGOTA', 'BOLIVAR', 'BOYACA',
    'CALDAS', 'CAQUETA', 'CASANARE', 'CAUCA', 'CESAR', 'CHOCO', 'CORDOBA',
    'CUNDINAMARCA', 'GUAINIA', 'GUAVIARE', 'HUILA', 'LA GUAJIRA', 'MAGDALENA',
    'META', 'NARINO', 'NORTE DE SANTANDER', 'PUTUMAYO', 'QUINDIO', 'RISARALDA',
    'SAN ANDRES', 'SANTANDER', 'SUCRE', 'TOLIMA', 'VALLE DEL CAUCA', 'VAUPES',
    'VICHADA', 'ARCHIPIELAGO DE SAN ANDRES', 'BOGOTA D.C', 'DISTRITO CAPITAL',
}

PAISES_REFERENCIA_FAMILIA = {
    'COLOMBIA', 'VENEZUELA', 'ECUADOR', 'PERU', 'BRASIL', 'PANAMA', 'CHILE',
    'ARGENTINA', 'MEXICO', 'BOLIVIA', 'PARAGUAY', 'URUGUAY', 'ESTADOS UNIDOS',
    'ESPANA', 'CUBA', 'REPUBLICA DOMINICANA',
}


def fecha_excel_familia(valor):
    if not valor:
        return ""
    if hasattr(valor, 'strftime'):
        return valor.strftime("%d/%m/%Y")

    texto = texto_excel(valor)
    coincidencia = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4,5})$', texto)
    if not coincidencia:
        return ""

    dia, mes, anio = coincidencia.groups()
    candidatos = []

    def es_fecha_valida(anio_texto):
        try:
            fecha = datetime(int(anio_texto), int(mes), int(dia))
            if 1900 <= fecha.year <= datetime.now().year + 1:
                return fecha.strftime("%d/%m/%Y")
        except Exception:
            return ""
        return ""

    if len(anio) == 4:
        return es_fecha_valida(anio)

    if len(anio) == 5 and anio.count('9') >= 2:
        anio_corregido = anio.replace('9', '', 1)
        fecha_corregida = es_fecha_valida(anio_corregido)
        if fecha_corregida:
            return fecha_corregida

    for indice in range(len(anio)):
        candidato = anio[:indice] + anio[indice + 1:]
        if len(candidato) != 4:
            continue
        fecha_valida = es_fecha_valida(candidato)
        if fecha_valida and fecha_valida not in candidatos:
            candidatos.append(fecha_valida)

    return candidatos[0] if len(candidatos) == 1 else ""


def resolver_nacimiento_familia(departamento, municipio):
    departamento_texto = texto_excel(valor_excel_familia(departamento))
    municipio_texto = texto_excel(valor_excel_familia(municipio))
    departamento_norm = normalizar_texto(departamento_texto)
    municipio_norm = normalizar_texto(municipio_texto)

    if ESTANDARIZAR_CUNDINAMARCA_BOGOTA_DC:
        bogota_alias = {'BOGOTA', 'BOGOTA DC', 'BOGOTA D C', 'DISTRITO CAPITAL'}
        if departamento_norm == 'CUNDINAMARCA' and municipio_norm in bogota_alias:
            departamento_texto = 'BOGOTA D.C.'
            municipio_texto = 'BOGOTA D.C.'
            departamento_norm = normalizar_texto(departamento_texto)
            municipio_norm = normalizar_texto(municipio_texto)

    if not departamento_norm and not municipio_norm:
        return {
            'pais_nacimiento': 'COLOMBIA',
            'departamento_nacimiento': '',
            'municipio_nacimiento': '',
        }

    if departamento_norm in PAISES_REFERENCIA_FAMILIA and departamento_norm != 'COLOMBIA':
        return {
            'pais_nacimiento': departamento_texto,
            'departamento_nacimiento': '',
            'municipio_nacimiento': '',
        }

    if municipio_norm in PAISES_REFERENCIA_FAMILIA and municipio_norm != 'COLOMBIA' and departamento_norm == municipio_norm:
        return {
            'pais_nacimiento': municipio_texto,
            'departamento_nacimiento': '',
            'municipio_nacimiento': '',
        }

    if departamento_norm and departamento_norm == municipio_norm and departamento_norm not in DEPARTAMENTOS_COLOMBIA:
        return {
            'pais_nacimiento': departamento_texto,
            'departamento_nacimiento': '',
            'municipio_nacimiento': '',
        }

    return {
        'pais_nacimiento': 'COLOMBIA',
        'departamento_nacimiento': departamento_texto,
        'municipio_nacimiento': municipio_texto,
    }


def construir_beneficiario_grupo_familiar_desde_excel(datos_excel):
    nacimiento = resolver_nacimiento_familia(
        datos_excel.get('departamento_nacimiento_beneficiario', ''),
        datos_excel.get('municipio_nacimiento_beneficiario', ''),
    )
    pais_beneficiario_excel = texto_excel(datos_excel.get('pais_nacimiento_beneficiario', ''))
    pais_beneficiario = pais_beneficiario_excel or nacimiento['pais_nacimiento'] or 'COLOMBIA'

    return {
        'tipo': 'beneficiario',
        'nombre_completo': nombre_completo(
            datos_excel.get('primer_nombre', ''),
            datos_excel.get('segundo_nombre', ''),
            datos_excel.get('primer_apellido', ''),
            datos_excel.get('segundo_apellido', ''),
        ),
        'doc_tipo': mapear_tipo_documento_grupo_familiar(
            datos_excel.get('tipo_documento_beneficiario', ''),
            datos_excel.get('nacionalidad_beneficiario', ''),
        ),
        'documento': texto_excel(datos_excel.get('documento', '')),
        'primer_nombre': texto_excel(datos_excel.get('primer_nombre', '')),
        'segundo_nombre': texto_excel(datos_excel.get('segundo_nombre', '')),
        'primer_apellido': texto_excel(datos_excel.get('primer_apellido', '')),
        'segundo_apellido': texto_excel(datos_excel.get('segundo_apellido', '')),
        'fecha_nacimiento': texto_excel(datos_excel.get('fecha_nacimiento_beneficiario', '')),
        'pais_nacimiento': pais_beneficiario,
        'departamento_nacimiento': nacimiento['departamento_nacimiento'],
        'municipio_nacimiento': nacimiento['municipio_nacimiento'],
        'parentesco_beneficiario': determinar_parentesco_beneficiario_desde_excel(datos_excel),
        'sexo': texto_excel(datos_excel.get('sexo', '')),
        'es_responsable': False,
    }


def texto_contiene_nombre(texto, nombre):
    texto_normalizado = normalizar_texto(texto)
    tokens = [token for token in normalizar_texto(nombre).split() if token]
    return bool(tokens) and all(token in texto_normalizado for token in tokens)


def obtener_valor_por_xpath(xpath, tipo='input'):
    try:
        elemento = driver.find_element(By.XPATH, xpath)
        if tipo == 'select':
            return Select(elemento).first_selected_option.text.strip()
        if tipo == 'checkbox':
            return 'SI' if elemento.is_selected() else 'NO'
        return (elemento.get_attribute('value') or '').strip()
    except Exception:
        return ""


def textos_equivalentes_select(texto_objetivo):
    objetivo = normalizar_texto(texto_objetivo)
    equivalencias = {
        'MASCULINO': ['MASCULINO', 'HOMBRE'],
        'FEMENINO': ['FEMENINO', 'MUJER'],
        'JEFE DEL GRUPO FAMILIAR': ['JEFE DEL GRUPO FAMILIAR', 'JEFE DE HOGAR', 'JEFE'],
        'CONYUGE/COMPANERO(A)': ['CONYUGE/COMPANERO(A)', 'CONYUGE', 'COMPANERO(A)', 'COMPANERO'],
        'PARIENTE U OTRO': ['PARIENTE U OTRO', 'PARIENTE Y OTRO', 'PARIENTE', 'OTRO'],
        'ABUELO (A)': ['ABUELO (A)', 'ABUELO(A)', 'ABUELO', 'ABUELA'],
        'TIO (A)': ['TIO (A)', 'TIO(A)', 'TIO', 'TIA'],
        'HERMANO (A)': ['HERMANO (A)', 'HERMANO(A)', 'HERMANO', 'HERMANA'],
    }
    opciones = equivalencias.get(objetivo, [texto_objetivo])
    if objetivo in EQUIVALENCIAS_TIPO_VIA:
        opciones = EQUIVALENCIAS_TIPO_VIA[objetivo]
    else:
        for candidatos in EQUIVALENCIAS_TIPO_VIA.values():
            normalizados = [normalizar_texto(candidato) for candidato in candidatos]
            if objetivo in normalizados:
                opciones = candidatos
                break
    return [normalizar_texto(opcion) for opcion in opciones if opcion]


def texto_select_coincide(texto_actual, texto_objetivo):
    actual = normalizar_texto(texto_actual)
    if not actual:
        return False
    for esperado in textos_equivalentes_select(texto_objetivo):
        if actual == esperado or esperado in actual or actual in esperado:
            return True
    return False


def seleccionar_por_texto_normalizado(select_elem, texto_objetivo):
    objetivos = textos_equivalentes_select(texto_objetivo)
    if not objetivos:
        return False

    for option in select_elem.find_elements(By.TAG_NAME, "option"):
        texto_opcion = normalizar_texto(option.text)
        if any(texto_opcion == objetivo for objetivo in objetivos):
            Select(select_elem).select_by_visible_text(option.text)
            return True

    for option in select_elem.find_elements(By.TAG_NAME, "option"):
        texto_opcion = normalizar_texto(option.text)
        if any(objetivo in texto_opcion or texto_opcion in objetivo for objetivo in objetivos):
            Select(select_elem).select_by_visible_text(option.text)
            return True

    return False


def asignar_select_elemento_con_candidatos(select_elem, candidatos, descripcion, solo_si_vacio=False):
    candidatos_limpios = [texto_excel(candidato) for candidato in candidatos if texto_excel(candidato)]
    if not candidatos_limpios or select_elem is None:
        return False

    try:
        driver.execute_script(
            """
            arguments[0].removeAttribute('disabled');
            arguments[0].removeAttribute('readonly');
            arguments[0].style.visibility = 'visible';
            """,
            select_elem,
        )
    except Exception:
        pass

    try:
        actual = Select(select_elem).first_selected_option.text.strip()
        if any(texto_select_coincide(actual, candidato) for candidato in candidatos_limpios):
            print(f"  [=] {descripcion}: ya tiene '{actual}' (sin cambios)")
            return True
        if solo_si_vacio and not texto_select_esta_vacio(actual):
            print(f"  [=] {descripcion}: ya tiene '{actual}' (sin cambios)")
            return True
    except Exception:
        actual = ""

    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_elem)
    except Exception:
        pass

    for candidato in candidatos_limpios:
        try:
            if not seleccionar_por_texto_normalizado(select_elem, candidato):
                continue
            driver.execute_script(
                """
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                """,
                select_elem,
            )
            time.sleep(0.8)
            actual_final = Select(select_elem).first_selected_option.text.strip()
            print(f"  [+] {descripcion}: {actual_final or candidato}")
            return True
        except Exception:
            continue

    return False


def obtener_texto_select_por_id(element_id):
    try:
        elementos = driver.find_elements(By.ID, element_id)
        if not elementos:
            return ""
        return Select(elementos[0]).first_selected_option.text.strip()
    except Exception:
        return ""


def control_esta_habilitado(elemento):
    try:
        if not elemento.is_enabled():
            return False
        atributo_disabled = elemento.get_attribute("disabled")
        clases = (elemento.get_attribute("class") or "").lower()
        aria_disabled = (elemento.get_attribute("aria-disabled") or "").lower()
        if atributo_disabled:
            return False
        if "aspnetdisabled" in clases:
            return False
        if aria_disabled == "true":
            return False
        return True
    except Exception:
        return False


def esperar_control_habilitado(element_id, timeout=4.0):
    fin = time.time() + timeout
    ultimo_elemento = None
    while time.time() < fin:
        try:
            esperar_postback_finalizado(1.0)
            elementos = driver.find_elements(By.ID, element_id)
            if elementos:
                ultimo_elemento = elementos[0]
                if control_esta_habilitado(ultimo_elemento):
                    return ultimo_elemento
        except Exception:
            pass
        time.sleep(0.2)
    return ultimo_elemento


def asignar_valor_input_por_id(element_id, valor, descripcion, timeout=4.0):
    try:
        elemento = esperar_control_habilitado(element_id, timeout=timeout)
        if elemento is None:
            return False
        if not control_esta_habilitado(elemento):
            print(f"  [~] {descripcion}: control bloqueado, se omite")
            return False
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
        valor_actual = (elemento.get_attribute("value") or "").strip()
        if valor_actual == valor:
            print(f"  [=] {descripcion}: ya tiene '{valor}' (sin cambios)")
            return True
        driver.execute_script(
            """
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
            """,
            elemento,
            valor,
        )
        time.sleep(0.5)
        print(f"  [+] {descripcion}: {valor}")
        return True
    except Exception as e:
        print(f"  [!] No fue posible asignar {descripcion}: {e}")
        return False


def asignar_valor_input_verificando_por_id(element_id, valor, descripcion, timeout=4.0, reintentos=3):
    valor_esperado = "" if valor is None else str(valor)
    ultimo_error = None

    for intento in range(reintentos):
        try:
            elemento = esperar_control_habilitado(element_id, timeout=timeout)
            if elemento is None:
                ultimo_error = "No se encontró el control"
                continue
            if not control_esta_habilitado(elemento):
                print(f"  [~] {descripcion}: control bloqueado, se omite")
                return False

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
            valor_actual = (elemento.get_attribute("value") or "").strip()
            if valor_actual == valor_esperado:
                print(f"  [=] {descripcion}: ya tiene '{valor_esperado}' (sin cambios)")
                return True

            driver.execute_script(
                """
                arguments[0].removeAttribute('disabled');
                arguments[0].removeAttribute('readonly');
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                """,
                elemento,
                valor_esperado,
            )
            time.sleep(0.4)
            esperar_postback_finalizado(1.0)
            valor_final = (driver.find_element(By.ID, element_id).get_attribute("value") or "").strip()
            if valor_final == valor_esperado:
                print(f"  [+] {descripcion}: {valor_esperado}")
                return True

            ultimo_error = f"quedó '{valor_final}'"
            if intento < reintentos - 1:
                print(f"  [~] {descripcion}: no persistió ({ultimo_error}), reintentando...")
                time.sleep(0.6)
        except Exception as e:
            ultimo_error = e
            if intento < reintentos - 1:
                print(f"  [~] {descripcion}: reintentando por refresco del DOM...")
                time.sleep(0.6)

    print(f"  [!] No fue posible asignar {descripcion}: {ultimo_error}")
    return False


def completar_campos_georreferencia(latitud, longitud):
    campos = [
        ("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtGradosLatitud", latitud[0], "Grados latitud"),
        ("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtMinutosLatitud", latitud[1], "Minutos latitud"),
        ("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtSegundosLatitud", latitud[2], "Segundos latitud"),
        ("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtGradosLongitud", longitud[0], "Grados longitud"),
        ("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtMinutosLongitud", longitud[1], "Minutos longitud"),
        ("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtSegundosLongitud", longitud[2], "Segundos longitud"),
    ]

    ok = True
    for element_id, valor, descripcion in campos:
        ok = asignar_valor_input_verificando_por_id(element_id, valor, descripcion, timeout=6.0, reintentos=3) and ok

    return ok


def asignar_select_por_id(element_id, valor, descripcion, reintentos=2):
    for intento in range(reintentos + 1):
        try:
            select_elem = esperar_control_habilitado(element_id, timeout=4.0)
            if select_elem is None:
                if intento < reintentos:
                    time.sleep(1.5)
                    continue
                return False
            if not control_esta_habilitado(select_elem):
                print(f"  [~] {descripcion}: control bloqueado, se omite")
                return False
            # Si ya tiene el valor correcto, no tocarlo (evita reset de cascada)
            try:
                actual = Select(select_elem).first_selected_option.text.strip()
                if texto_select_coincide(actual, valor):
                    print(f"  [=] {descripcion}: ya tiene '{valor}' (sin cambios)")
                    return True
            except Exception:
                pass
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_elem)
            if seleccionar_por_texto_normalizado(select_elem, valor):
                driver.execute_script(
                    """
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                    """,
                    select_elem,
                )
                time.sleep(1)
                actual_final = obtener_texto_select_por_id(element_id)
                if texto_select_coincide(actual_final, valor):
                    print(f"  [+] {descripcion}: {actual_final or valor}")
                    return True
                if intento < reintentos:
                    print(f"  [~] {descripcion}: el portal dejó '{actual_final or 'Seleccione'}', reintentando...")
                    time.sleep(1.5)
                    continue
                print(f"  [!] {descripcion}: el portal no conservó el valor '{valor}' (quedó '{actual_final or 'Seleccione'}')")
                return False
            print(f"  [!] No se encontró opción para {descripcion}: {valor}")
            return False
        except Exception as e:
            msg = str(e)
            if 'stale element' in msg.lower() and intento < reintentos:
                print(f"  [~] Elemento obsoleto al asignar {descripcion}, reintentando ({intento+1}/{reintentos})...")
                time.sleep(1.5)
                continue
            print(f"  [!] No fue posible asignar {descripcion}: {msg.splitlines()[0][:120]}")
            return False
    return False


def completar_ubicacion_nacimiento_grupo_familiar(persona, prefijo_log="Grupo familiar"):
    id_pais = "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdPaisNacimiento"
    id_departamento = "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdDepartamento"
    id_municipio = "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdMunicipioNacimiento"

    pais_objetivo = persona.get('pais_nacimiento', '')
    departamento_objetivo = persona.get('departamento_nacimiento', '')
    municipio_objetivo = persona.get('municipio_nacimiento', '')

    if not pais_objetivo and (departamento_objetivo or municipio_objetivo):
        pais_objetivo = 'COLOMBIA'

    if not pais_objetivo:
        return True

    def asignar_select_cascada(element_id, valor, descripcion):
        if not valor:
            return False
        return asignar_select_por_id(element_id, valor, descripcion)

    # CRITICAL: Ensure country is assigned and stays selected
    for intento in range(3):
        exito = asignar_select_cascada(id_pais, pais_objetivo, f"{prefijo_log} pais_nacimiento (intento {intento+1})")
        esperar_postback_finalizado(1.0)
        esperar_detalle_grupo_familiar(1.0)
        
        pais_final = obtener_texto_select_por_id(id_pais)
        if texto_select_coincide(pais_final, pais_objetivo) or texto_select_coincide(pais_final, 'COLOMBIA'):
            print(f"  [+] {prefijo_log} pais_nacimiento: {pais_final} ✓")
            break
        elif intento == 2:
            print(f"  [!] {prefijo_log} pais_nacimiento: no se logró asignar después de {intento+1} intentos (quedó: {pais_final})")
            return False

    pais_es_colombia = texto_select_coincide(pais_final, 'COLOMBIA') or texto_select_coincide(pais_objetivo, 'COLOMBIA')

    if not pais_es_colombia:
        return True

    if not departamento_objetivo:
        departamento_actual = obtener_texto_select_por_id(id_departamento)
        if departamento_actual and not texto_select_coincide(departamento_actual, 'Seleccione'):
            departamento_objetivo = departamento_actual

    if departamento_objetivo:
        asignar_select_cascada(id_departamento, departamento_objetivo, f"{prefijo_log} departamento_nacimiento")
        esperar_postback_finalizado(1.2)
        esperar_detalle_grupo_familiar(1.0)
    else:
        print(f"  [~] {prefijo_log} departamento_nacimiento: vacío (solo país asignado)")

    if not municipio_objetivo:
        municipio_actual = obtener_texto_select_por_id(id_municipio)
        if municipio_actual and not texto_select_coincide(municipio_actual, 'Seleccione'):
            municipio_objetivo = municipio_actual

    if municipio_objetivo:
        esperar_opcion_en_select(id_municipio, municipio_objetivo, timeout=8)
        asignar_select_cascada(id_municipio, municipio_objetivo, f"{prefijo_log} municipio_nacimiento")
        esperar_postback_finalizado(0.8)
    else:
        print(f"  [~] {prefijo_log} municipio_nacimiento: vacío (solo país/depto asignados)")

    # Final validation of country selection (critical for form submission)
    pais_final_validacion = obtener_texto_select_por_id(id_pais)
    if texto_select_coincide(pais_final_validacion, 'Seleccione') or not pais_final_validacion:
        print(f"  [!] {prefijo_log}: país de nacimiento no quedó seleccionado (validación final)")
        return False

    departamento_final = obtener_texto_select_por_id(id_departamento)
    municipio_final = obtener_texto_select_por_id(id_municipio)
    if departamento_objetivo and texto_select_coincide(departamento_final, 'Seleccione'):
        print(f"  [!] {prefijo_log}: departamento de nacimiento no quedó seleccionado")
        return False
    if municipio_objetivo and texto_select_coincide(municipio_final, 'Seleccione'):
        print(f"  [!] {prefijo_log}: municipio de nacimiento no quedó seleccionado")
        return False

    return True


def encontrar_control_por_etiqueta(etiqueta, tag, solo_visibles=True):
    variantes = [etiqueta, etiqueta.replace('*', '').strip()]
    for texto in variantes:
        xpaths = [
            f"//label[contains(normalize-space(.), \"{texto}\")]/following::{tag}[1]",
            f"//*[contains(normalize-space(.), \"{texto}\")]/following::{tag}[1]",
            f"//td[contains(normalize-space(), \"{texto}\")]/following-sibling::td//{tag}[1]",
            f"//span[contains(normalize-space(.), \"{texto}\")]/following::{tag}[1]",
        ]
        for xpath in xpaths:
            try:
                elementos = driver.find_elements(By.XPATH, xpath)
                if solo_visibles:
                    visibles = [el for el in elementos if el.is_displayed()]
                    if visibles:
                        return visibles[0]
                elif elementos:
                    return elementos[0]
            except Exception:
                pass
    return None


def asignar_valor_input_por_etiqueta(etiqueta, valor, descripcion):
    try:
        elemento = encontrar_control_por_etiqueta(etiqueta, 'input')
        if elemento is None:
            return False
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
        driver.execute_script(
            """
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
            """,
            elemento,
            valor,
        )
        time.sleep(0.5)
        print(f"  [+] {descripcion}: {valor}")
        return True
    except Exception as e:
        print(f"  [!] No fue posible asignar {descripcion}: {e}")
        return False


def asignar_select_por_etiqueta(etiqueta, valor, descripcion):
    try:
        select_elem = encontrar_control_por_etiqueta(etiqueta, 'select')
        if select_elem is None:
            return False
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_elem)
        return asignar_select_elemento_con_candidatos(select_elem, [valor], descripcion)
    except Exception as e:
        print(f"  [!] No fue posible asignar {descripcion}: {e}")
        return False


def texto_select_esta_vacio(texto):
    return normalizar_texto(texto) in {'', 'SELECCIONE', '-1'}


def asignar_select_por_etiquetas_con_candidatos(etiquetas, candidatos, descripcion, solo_si_vacio=False, solo_visibles=True):
    candidatos_limpios = [texto_excel(candidato) for candidato in candidatos if texto_excel(candidato)]
    if not candidatos_limpios:
        return False

    for etiqueta in etiquetas:
        try:
            select_elem = encontrar_control_por_etiqueta(etiqueta, 'select', solo_visibles=solo_visibles)
            if select_elem is None:
                continue
            if asignar_select_elemento_con_candidatos(select_elem, candidatos_limpios, descripcion, solo_si_vacio=solo_si_vacio):
                return True
        except Exception:
            continue
    print(f"  [~] No se encontró opción compatible para {descripcion}: {', '.join(candidatos_limpios)}")
    return False


def asignar_select_por_xpath_con_candidatos(xpath, candidatos, descripcion, solo_si_vacio=False, solo_visibles=True):
    candidatos_limpios = [texto_excel(candidato) for candidato in candidatos if texto_excel(candidato)]
    if not candidatos_limpios:
        return False

    try:
        elementos = driver.find_elements(By.XPATH, xpath)
        if solo_visibles:
            elementos = [elemento for elemento in elementos if elemento.is_displayed()]
        for select_elem in elementos:
            if asignar_select_elemento_con_candidatos(select_elem, candidatos_limpios, descripcion, solo_si_vacio=solo_si_vacio):
                return True
    except Exception:
        pass

    return False


def asignar_select_por_opciones_disponibles(candidatos, descripcion, solo_si_vacio=False):
    candidatos_limpios = [texto_excel(candidato) for candidato in candidatos if texto_excel(candidato)]
    if not candidatos_limpios:
        return False

    try:
        selects = driver.find_elements(By.TAG_NAME, 'select')
    except Exception:
        return False

    for select_elem in selects:
        try:
            opciones = [normalizar_texto(opcion.text) for opcion in select_elem.find_elements(By.TAG_NAME, 'option')]
            if not opciones:
                continue
            if not any(
                any(opcion == esperado or esperado in opcion or opcion in esperado for esperado in textos_equivalentes_select(candidato))
                for opcion in opciones for candidato in candidatos_limpios
            ):
                continue
            if asignar_select_elemento_con_candidatos(select_elem, candidatos_limpios, descripcion, solo_si_vacio=solo_si_vacio):
                return True
        except Exception:
            continue

    return False


def limpiar_selects_por_etiquetas(etiquetas, descripcion, candidatos_limpieza=None):
    candidatos_limpieza = candidatos_limpieza or ['Seleccione', 'SELECCIONE', 'NO APLICA', 'NINGUNO', 'NINGUNA']
    for etiqueta in etiquetas:
        try:
            select_elem = encontrar_control_por_etiqueta(etiqueta, 'select')
            if select_elem is None:
                continue
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_elem)
            for candidato in candidatos_limpieza:
                if seleccionar_por_texto_normalizado(select_elem, candidato):
                    driver.execute_script(
                        """
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                        """,
                        select_elem,
                    )
                    time.sleep(0.5)
                    actual_final = Select(select_elem).first_selected_option.text.strip()
                    print(f"  [+] {descripcion}: {actual_final or candidato}")
                    return True
        except Exception:
            continue
    return False


def esperar_opcion_en_select(element_id, texto_objetivo, timeout=8):
    objetivo = normalizar_texto(texto_objetivo)
    if not objetivo:
        return False

    def _opcion_disponible(_driver):
        elementos = _driver.find_elements(By.ID, element_id)
        if not elementos:
            return False
        opciones = elementos[0].find_elements(By.TAG_NAME, "option")
        for opcion in opciones:
            texto_opcion = normalizar_texto(opcion.text)
            if texto_opcion == objetivo or objetivo in texto_opcion or texto_opcion in objetivo:
                return True
        return False

    try:
        WebDriverWait(driver, timeout).until(_opcion_disponible)
        return True
    except Exception:
        return False


def pulsar_boton_azul_grupo_familiar():
    selectores = [
        (By.XPATH, "//input[contains(@id, 'tbngrupofamiliar') and (@type='image' or @type='button') and (contains(@id, 'btnBuscar') or contains(@id, 'btnInfo'))]"),
        (By.XPATH, "//a[contains(@id, 'tbngrupofamiliar')]//img[contains(@src, 'info')]/parent::a"),
        (By.XPATH, "//input[contains(@id, 'tbngrupofamiliar') and @type='image' and contains(@src, 'info')]"),
        (By.XPATH, "//input[contains(@id, 'tbngrupofamiliar') and @type='image' and contains(@src, 'buscar')]"),
    ]

    for selector in selectores:
        try:
            botones = [elemento for elemento in driver.find_elements(selector[0], selector[1]) if elemento.is_displayed()]
            if not botones:
                continue
            boton = botones[0]
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton)
            time.sleep(0.5)
            try:
                boton.click()
            except Exception:
                driver.execute_script("arguments[0].click();", boton)
            esperar_postback_finalizado(1.0)
            esperar_detalle_grupo_familiar(1.0)
            print("  [+] Botón azul de Grupo Familiar pulsado")
            return True
        except Exception:
            pass

    print("  [!] No se encontró botón azul de Grupo Familiar después de seleccionar el ID")
    return False


def pulsar_lupa_buscar_grupo_familiar(contexto="Grupo familiar"):
    selectores = [
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_btnBuscar"),
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_btnBuscar"),
        (By.XPATH, "//input[contains(@id, 'tbngrupofamiliar') and contains(@id, 'btnBuscar') and (@type='image' or @type='button') and not(contains(@id, 'BeneficiarioVincula'))]"),
    ]

    for by, selector in selectores:
        try:
            elementos = driver.find_elements(by, selector)
            visibles = [el for el in elementos if el.is_displayed()]
            if not visibles:
                continue
            lupa = visibles[0]
            if not control_esta_habilitado(lupa):
                continue

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", lupa)
            try:
                lupa.click()
            except Exception:
                driver.execute_script("arguments[0].click();", lupa)
            esperar_postback_finalizado(1.2)
            esperar_detalle_grupo_familiar(1.0)
            print(f"  [+] {contexto}: lupa de búsqueda pulsada")
            return True
        except Exception:
            pass

    print(f"  [~] {contexto}: lupa de búsqueda no disponible o bloqueada")
    return False


def mapear_tipo_documento_grupo_familiar(valor_excel, nacionalidad=''):
    tipo = normalizar_texto(valor_excel)
    nacionalidad_normalizada = normalizar_texto(nacionalidad)

    if 'VENEZ' in nacionalidad_normalizada and tipo in {'CEDULA', 'CC', 'CEDULA DE CIUDADANIA'}:
        return 'PERMISO ESPECIAL DE PERMANENCIA'

    equivalencias = {
        'CEDULA': 'CEDULA DE CIUDADANIA',
        'CC': 'CEDULA DE CIUDADANIA',
        'CEDULA DE CIUDADANIA': 'CEDULA DE CIUDADANIA',
        'CE': 'CEDULA DE EXTRANJERIA',
        'CEDULA DE EXTRANJERIA': 'CEDULA DE EXTRANJERIA',
        'PEP': 'PERMISO ESPECIAL DE PERMANENCIA',
        'PERMISO ESPECIAL DE PERMANENCIA': 'PERMISO ESPECIAL DE PERMANENCIA',
        'PPT': 'PERMISO ESPECIAL DE PERMANENCIA',
        'PERMISO POR PROTECCION TEMPORAL': 'PERMISO ESPECIAL DE PERMANENCIA',
        'RC': 'REGISTRO CIVIL',
        'REGISTRO CIVIL': 'REGISTRO CIVIL',
        'TI': 'TARJETA DE IDENTIDAD',
        'TARJETA DE IDENTIDAD': 'TARJETA DE IDENTIDAD',
        'PASAPORTE': 'PASAPORTE',
    }
    return equivalencias.get(tipo, valor_excel)


def normalizar_zona_excel(valor):
    zona = normalizar_texto(valor)
    equivalencias_zona = {
        'URBANO': 'CABECERA',
        'URBANA': 'CABECERA',
        'RURAL': 'RESTO',
        'RESTO': 'RESTO',
    }
    zona_normalizada = equivalencias_zona.get(zona, valor)
    if not zona_normalizada:
        zona_normalizada = 'RESTO'  # Si no hay zona en Excel, asumir RESTO
    return zona_normalizada


def asegurar_zona_ubicacion_requerida(datos_excel):
    zona_objetivo = normalizar_zona_excel(datos_excel.get('zona_residencia', ''))
    if not zona_objetivo:
        return False

    element_id = "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlZonaUbicacion"
    actual = obtener_texto_select_por_id(element_id)
    if texto_select_coincide(actual, zona_objetivo):
        return True

    print(f"  [~] Reforzando Zona residencia requerida hacia '{zona_objetivo}'...")
    if asignar_select_por_id(element_id, zona_objetivo, "Zona residencia", reintentos=3):
        actual = obtener_texto_select_por_id(element_id)
        if texto_select_coincide(actual, zona_objetivo):
            return True

    try:
        elementos = driver.find_elements(By.ID, element_id)
        if elementos and asignar_select_elemento_con_candidatos(elementos[0], [zona_objetivo], "Zona residencia (forzada)"):
            actual = obtener_texto_select_por_id(element_id)
            if texto_select_coincide(actual, zona_objetivo):
                return True
    except Exception:
        pass

    print(f"  [!] Zona residencia sigue sin quedar en '{zona_objetivo}'")
    return False


def corregir_datos_ubicacion(datos_excel, es_alta_nueva=False):
    print("  [*] Corrigiendo Datos de Ubicación según Excel...")
    tiene_georreferencia = excel_tiene_georreferencia(datos_excel)
    datos_ubicacion_actuales = leer_datos_ubicacion()
    georeferenciado_actual = normalizar_texto(datos_ubicacion_actuales.get('georeferenciado', '')) == 'SI'
    debe_gestionar_georreferencia = tiene_georreferencia or georeferenciado_actual
    modo_georreferencia = 'excel' if tiene_georreferencia else ('existente_sin_excel' if georeferenciado_actual else 'sin_georreferencia')

    # Orden requerido por el formulario de RUB.
    asignar_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdPaisResidencia", datos_excel['pais_residencia'], "País residencia")
    asignar_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdDepartamentoResidencia", datos_excel['departamento_residencia'], "Departamento residencia")
    asignar_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdMunicipioResidencia", datos_excel['municipio_residencia'], "Municipio residencia")
    asignar_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdRancheria", "Seleccione", "Ranchería")
    asignar_select_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlZonaUbicacion", normalizar_zona_excel(datos_excel['zona_residencia']), "Zona residencia")
    asegurar_zona_ubicacion_requerida(datos_excel)
    completar_controles_adicionales_ubicacion(datos_excel)
    completar_direccion_residencia(datos_excel['direccion_residencia'])
    asignar_valor_input_por_id("cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtTelefono", datos_excel['telefono'], "Teléfono")
    asegurar_zona_ubicacion_requerida(datos_excel)

    latitud = parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')
    longitud = parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')
    direccion_latitud = obtener_direccion_georreferencia(datos_excel['latitud_excel'], 'latitud')
    direccion_longitud = obtener_direccion_georreferencia(datos_excel['longitud_excel'], 'longitud')

    if tiene_georreferencia:
        asignar_checkbox_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_chkGeoreferenciado",
            True,
            "Georeferenciado",
            timeout=6.0,
        )
    elif es_alta_nueva and not georeferenciado_actual:
        asignar_checkbox_por_id(
            "cphCont_TabContainer1_tbnDatosGeo_chkGeoreferenciado",
            False,
            "Georeferenciado",
            timeout=6.0,
        )
    elif georeferenciado_actual:
        print("  [=] Georeferenciado: se conserva en 'SI' porque el formulario ya tiene coordenadas")
    else:
        print("  [=] Georeferenciado: permanece en 'NO' al no existir coordenadas en Excel ni en formulario")

    if debe_gestionar_georreferencia:
        completar_fecha_captura_ubicacion(FECHA_CAPTURA_UBICACION)
        completar_hora_captura_ubicacion(HORA_CAPTURA_UBICACION)

    if not tiene_georreferencia:
        if georeferenciado_actual:
            print("  [~] Excel sin coordenadas: se validan fecha y hora, y se conservan las coordenadas existentes del formulario")
        elif es_alta_nueva:
            print("  [~] Alta nueva sin coordenadas en Excel: se deja Georeferenciado = NO; fecha y hora no aplican")
        else:
            print("  [~] Registro sin coordenadas en Excel ni formulario: no hay coordenadas que corregir")
        return {
            'modo_georreferencia': modo_georreferencia,
            'tiene_georreferencia_excel': False,
            'debe_gestionar_georreferencia': debe_gestionar_georreferencia,
        }

    if direccion_latitud == 'N':
        asignar_radio_por_id("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_rblLatitudRancheria_1", "Latitud hemisferio: N")
    elif direccion_latitud == 'S':
        asignar_radio_por_id("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_rblLatitudRancheria_0", "Latitud hemisferio: S")

    if direccion_longitud == 'O':
        asignar_radio_por_id("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_rblLongitudRancheria_1", "Longitud hemisferio: O")
    elif direccion_longitud == 'E':
        asignar_radio_por_id("cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_rblLongitudRancheria_0", "Longitud hemisferio: E")

    completar_campos_georreferencia(latitud, longitud)
    completar_campos_georreferencia(latitud, longitud)
    return {
        'modo_georreferencia': modo_georreferencia,
        'tiene_georreferencia_excel': True,
        'debe_gestionar_georreferencia': True,
    }


def abrir_tab_pertenencia_etnica():
    selectores_tab = [
        (By.ID, "__tab_cphCont_TabContainer1_tbnPertenenciaEtnica"),
        (By.XPATH, "//span[normalize-space()='Pertenencia Étnica']"),
        (By.XPATH, "//a[normalize-space()='Pertenencia Étnica']"),
        (By.XPATH, "//span[contains(normalize-space(.), 'Pertenencia')]"),
        (By.XPATH, "//a[contains(normalize-space(.), 'Pertenencia')]"),
    ]
    return abrir_tab_con_espera(
        "Pertenencia Étnica",
        selectores_tab,
        tab_pertenencia_etnica_lista,
        "pertenencia_etnica",
        timeout_postback=1.8,
        timeout_listo=4.5,
    )


def forzar_grupo_etnico_no_autoreconoce():
    texto_objetivo = 'NO SE AUTORRECONOCE EN NINGUNO DE LOS ANTERIORES'
    selectores = [
        (By.ID, "cphCont_TabContainer1_tbnDatosAut_CtlAutoReconocimiento_ddlGrupoEtnico"),
        (By.NAME, "ctl00$cphCont$TabContainer1$tbnDatosAut$CtlAutoReconocimiento$ddlGrupoEtnico"),
        (By.XPATH, "//select[@id='cphCont_TabContainer1_tbnDatosAut_CtlAutoReconocimiento_ddlGrupoEtnico' or @name='ctl00$cphCont$TabContainer1$tbnDatosAut$CtlAutoReconocimiento$ddlGrupoEtnico']"),
    ]

    for by, selector in selectores:
        try:
            elementos = driver.find_elements(by, selector)
            if not elementos:
                continue

            select_elem = elementos[0]
            valor_actual = Select(select_elem).first_selected_option.text.strip()
            if texto_select_coincide(valor_actual, texto_objetivo):
                print(f"  [=] Grupo étnico: ya tiene '{valor_actual}' (sin cambios)")
                return True

            driver.execute_script(
                """
                arguments[0].removeAttribute('disabled');
                arguments[0].removeAttribute('readonly');
                arguments[0].value = '8';
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                """,
                select_elem,
            )
            time.sleep(0.4)
            esperar_postback_finalizado(2)

            valor_actual = Select(driver.find_element(By.XPATH, "//select[@id='cphCont_TabContainer1_tbnDatosAut_CtlAutoReconocimiento_ddlGrupoEtnico' or @name='ctl00$cphCont$TabContainer1$tbnDatosAut$CtlAutoReconocimiento$ddlGrupoEtnico']")).first_selected_option.text.strip()
            if texto_select_coincide(valor_actual, texto_objetivo):
                print(f"  [+] Grupo étnico: {valor_actual}")
                return True
        except Exception:
            continue

    return False


def completar_pertenencia_etnica(datos_excel):
    print("  [*] Completando Pertenencia Étnica según Excel...")

    tab_abierta = abrir_tab_pertenencia_etnica()
    if not tab_abierta:
        print("  [~] Pertenencia Étnica no abrió; se intentará por DOM directamente")

    territorio_etnico = texto_excel(datos_excel.get('territorio_etnico', ''))
    hay_territorio_etnico = bool(territorio_etnico and not es_valor_ausente(territorio_etnico))
    candidatos_grupo_etnico = (
        ['AFROCOLOMBIANO O COMUNIDAD NEGRA', 'AFROCOLOMBIANO', 'COMUNIDAD NEGRA']
        if hay_territorio_etnico
        else ['NO SE AUTORRECONOCE EN NINGUNO DE LOS ANTERIORES', 'NO SE AUTORECONOCE', 'NO SE AUTORRECONOCE']
    )

    se_actualizo = False
    grupo_etnico_resuelto = False

    if not hay_territorio_etnico and forzar_grupo_etnico_no_autoreconoce():
        return True

    configuracion_selects = [
        {
            'etiquetas': [
                'Grupo étnico al que pertenece el beneficiario',
                'Grupo etnico al que pertenece el beneficiario',
            ],
            'candidatos': candidatos_grupo_etnico,
            'descripcion': 'Grupo étnico',
            'solo_si_vacio': grupo_etnico_resuelto,
        },
        {
            'etiquetas': [
                '¿El beneficiario reside en el lugar de origen de la comunidad?',
                'El beneficiario reside en el lugar de origen de la comunidad?',
            ],
            'candidatos': ['NO'] if hay_territorio_etnico else ['NO APLICA', 'NO', 'Seleccione'],
            'descripcion': 'Reside en lugar de origen',
            'solo_si_vacio': False,
        },
        {
            'etiquetas': [
                'Si el beneficiario es del grupo étnico afrocolombiano o comunidad negra, indique el nombre del consejo comunitario al que pertenece',
                'Si el beneficiario es del grupo étnico afrocolombiano o comunidad negra indique la comunidad a la que pertenece',
                'Si el beneficiario es del grupo étnico indique comunidad a la que pertenece',
            ],
            'candidatos': [territorio_etnico] if hay_territorio_etnico else [],
            'descripcion': 'Territorio/comunidad étnica',
            'solo_si_vacio': False,
        },
    ]

    for config in configuracion_selects:
        if asignar_select_por_etiquetas_con_candidatos(
            config['etiquetas'],
            config['candidatos'],
            config['descripcion'],
            solo_si_vacio=config.get('solo_si_vacio', False),
            solo_visibles=tab_abierta,
        ):
            se_actualizo = True
            esperar_postback_finalizado(2)

    if not se_actualizo and not grupo_etnico_resuelto:
        if asignar_select_por_xpath_con_candidatos(
            "//select[contains(@id, 'ddlIdGrupoEtnico') or contains(@name, 'ddlIdGrupoEtnico') or contains(@id, 'ddlGrupoEtnico') or contains(@name, 'ddlGrupoEtnico') or contains(@id, 'GrupoEtnico') or contains(@name, 'GrupoEtnico')]",
            candidatos_grupo_etnico,
            'Grupo étnico',
            solo_si_vacio=False,
            solo_visibles=False,
        ):
            se_actualizo = True
            esperar_postback_finalizado(2)

    if not se_actualizo and not grupo_etnico_resuelto:
        if asignar_select_por_opciones_disponibles(
            candidatos_grupo_etnico,
            'Grupo étnico',
            solo_si_vacio=False,
        ):
            se_actualizo = True
            esperar_postback_finalizado(2)

    if asignar_select_por_opciones_disponibles(
        ['NO'] if hay_territorio_etnico else ['NO APLICA', 'NO', 'Seleccione'],
        'Reside en lugar de origen',
        solo_si_vacio=False,
    ):
        se_actualizo = True
        esperar_postback_finalizado(1)

    if not hay_territorio_etnico:
        dependientes = [
            [
                'Si el beneficiario es del grupo étnico afrocolombiano o comunidad negra, indique el nombre del consejo comunitario al que pertenece',
                'Si el beneficiario es del grupo étnico afrocolombiano o comunidad negra indique la comunidad a la que pertenece',
                'Si el beneficiario es del grupo étnico afrocolombiano comunidad negra, palenquero o raizal indique la asociación y organización a la que pertenece',
                'Si el beneficiario es del grupo étnico indígena, indique el pueblo al que pertenece',
                'Tipo de resguardo/comunidad',
                'Si el beneficiario es del grupo étnico indigena indique el resguardo al que pertenece',
                'Si el beneficiario es del grupo étnico indique comunidad a la que pertenece',
                'Si el beneficiario es del grupo étnico rrom/gitano, indique la comunidad kumpania u organización a la que pertenece',
                '¿Cual es la lengua propia?',
                '¿En que idioma o lengua se comunica habitualmente?',
            ],
        ]
        for etiquetas_dependientes in dependientes:
            if limpiar_selects_por_etiquetas(etiquetas_dependientes, 'Limpieza pertenencia étnica'):
                se_actualizo = True
                esperar_postback_finalizado(1)

    orden_validado = [
        'Grupo étnico al que pertenece el beneficiario',
        '¿El beneficiario reside en el lugar de origen de la comunidad?',
        'Si el beneficiario es del grupo étnico afrocolombiano o comunidad negra, indique el nombre del consejo comunitario al que pertenece',
        'Si el beneficiario es del grupo étnico afrocolombiano o comunidad negra indique la comunidad a la que pertenece',
        'Si el beneficiario es del grupo étnico afrocolombiano comunidad negra, palenquero o raizal indique la asociación y organización a la que pertenece',
        'Si el beneficiario es del grupo étnico indígena, indique el pueblo al que pertenece',
        'Tipo de resguardo/comunidad',
        'Si el beneficiario es del grupo étnico indigena indique el resguardo al que pertenece',
        'Si el beneficiario es del grupo étnico indique comunidad a la que pertenece',
        'Si el beneficiario es del grupo étnico rrom/gitano, indique la comunidad kumpania u organización a la que pertenece',
        '¿Cual es la lengua propia?',
        '¿En que idioma o lengua se comunica habitualmente?',
    ]
    print("  [=] Orden de revisión Pertenencia Étnica:")
    for etiqueta in orden_validado:
        print(f"      - {etiqueta}")

    return se_actualizo


def abrir_tab_ubicacion():
    selectores_tab = [
        (By.ID, "__tab_cphCont_TabContainer1_tbnDatosU"),
        (By.XPATH, "//span[normalize-space()='Datos de Ubicación']"),
        (By.XPATH, "//a[normalize-space()='Datos de Ubicación']"),
    ]
    return abrir_tab_con_espera(
        "Datos de Ubicación",
        selectores_tab,
        tab_ubicacion_lista,
        "ubicacion",
        timeout_postback=1.2,
        timeout_listo=2.0,
    )


def leer_datos_ubicacion():
    datos = {
        'pais_residencia': '',
        'departamento_residencia': '',
        'municipio_residencia': '',
        'zona_residencia': '',
        'centro_poblado': '',
        'tipo_cabecera': '',
        'comuna': '',
        'barrio': '',
        'observacion_ubicacion': '',
        'tipo_via_direccion': '',
        'complemento_direccion': '',
        'direccion_resumen': '',
        'telefono': '',
        'correo': '',
        'georeferenciado': '',
        'grados_latitud': '',
        'minutos_latitud': '',
        'segundos_latitud': '',
        'grados_longitud': '',
        'minutos_longitud': '',
        'segundos_longitud': '',
        'fecha_captura': '',
        'hora_captura': '',
    }

    select_ids = {
        'pais_residencia': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdPaisResidencia",
        'departamento_residencia': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdDepartamentoResidencia",
        'municipio_residencia': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdMunicipioResidencia",
        'zona_residencia': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlZonaUbicacion",
        'centro_poblado': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdCentroPoblado",
        'tipo_cabecera': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdTipoCabecera",
        'comuna': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdComuna",
        'barrio': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_ddlIdBarrio",
        'tipo_via_direccion': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlVia",
        'complemento_direccion': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtDireccionResidencia_ddlManzana",
    }

    for llave, element_id in select_ids.items():
        try:
            datos[llave] = Select(driver.find_element(By.ID, element_id)).first_selected_option.text.strip()
        except Exception:
            datos[llave] = ""

    input_ids = {
        'observacion_ubicacion': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtObservacionAdicionalUbicacion",
        'telefono': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtTelefono",
        'correo': "cphCont_TabContainer1_tbnDatosGeo_UbicacionPersona_txtCorreoElectronico",
        'grados_latitud': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtGradosLatitud",
        'minutos_latitud': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtMinutosLatitud",
        'segundos_latitud': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtSegundosLatitud",
        'grados_longitud': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtGradosLongitud",
        'minutos_longitud': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtMinutosLongitud",
        'segundos_longitud': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtSegundosLongitud",
        'fecha_captura': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_fecha_txtFecha",
        'hora_captura': "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtHora",
    }

    for llave, element_id in input_ids.items():
        try:
            datos[llave] = (driver.find_element(By.ID, element_id).get_attribute('value') or '').strip()
        except Exception:
            datos[llave] = ""

    try:
        datos['georeferenciado'] = 'SI' if driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosGeo_chkGeoreferenciado").is_selected() else 'NO'
    except Exception:
        datos['georeferenciado'] = ''

    datos['direccion_resumen'] = leer_direccion_residencia_compuesta()

    return datos


def completar_fecha_captura_ubicacion(fecha_objetivo):
    print(f"  [*] Ajustando fecha de captura de ubicación a {fecha_objetivo}...")
    element_id = "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_fecha_txtFecha"
    ultimo_error = None

    for intento in range(3):
        try:
            campo_fecha = esperar_control_habilitado(element_id, timeout=4.0)
            if campo_fecha is None:
                ultimo_error = "No se encontró el campo de fecha"
                continue
            if not control_esta_habilitado(campo_fecha):
                print("  [~] Fecha de captura: control bloqueado, se omite")
                return False

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", campo_fecha)
            time.sleep(0.3)
            driver.execute_script(
                """
                arguments[0].removeAttribute('disabled');
                arguments[0].removeAttribute('readonly');
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                if (typeof __doPostBack === 'function') {
                    __doPostBack(arguments[0].name, '');
                }
                arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                """,
                campo_fecha,
                fecha_objetivo,
            )
            esperar_postback_finalizado(2.0)
            time.sleep(0.5)
            fecha_final = (driver.find_element(By.ID, element_id).get_attribute('value') or '').strip()
            print(f"  [+] Fecha de captura final: {fecha_final}")
            if normalizar_texto(fecha_final) == normalizar_texto(fecha_objetivo):
                return True
            ultimo_error = f"La fecha no persistió (quedó '{fecha_final}')"
        except Exception as e:
            ultimo_error = e
            if intento < 2:
                print(f"  [~] Reintentando ajuste de fecha de captura ({intento + 1}/3)...")
                time.sleep(0.8)

    print(f"  [!] No fue posible ajustar la fecha de captura: {ultimo_error}")
    return False


def completar_hora_captura_ubicacion(hora_objetivo):
    print(f"  [*] Ajustando hora de captura de ubicación a {hora_objetivo}...")

    id_hora_exacto = "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtHora"
    if asignar_valor_input_verificando_por_id(id_hora_exacto, hora_objetivo, "Hora de captura de coordenada", timeout=6.0, reintentos=3):
        return True

    ids_probables = [
        "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_fecha_txtHora",
        "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtHora",
        "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_txtHoraCaptura",
        "cphCont_TabContainer1_tbnDatosGeo_GeoBeneficiario_hora_txtHora",
    ]

    for element_id in ids_probables:
        if asignar_valor_input_verificando_por_id(element_id, hora_objetivo, "Hora de captura de coordenada", timeout=6.0, reintentos=3):
            return True

    etiquetas = [
        'Hora de Captura de la coordenada(Sistema horario 24 horas)',
        'Hora de Captura de la coordenada',
        'Hora de captura de la coordenada(Sistema horario 24 horas)',
        'Hora de captura de la coordenada',
    ]
    for etiqueta in etiquetas:
        if asignar_valor_input_por_etiqueta(etiqueta, hora_objetivo, "Hora de captura de coordenada"):
            return True

    xpaths_probables = [
        "//div[contains(@id, 'GeoBeneficiario')]//input[contains(@id, 'Hora')]",
        "//div[contains(@id, 'GeoBeneficiario')]//input[contains(@name, 'Hora')]",
        "//*[contains(normalize-space(.), 'Hora de Captura de la coordenada')]/following::input[1]",
        "//*[contains(normalize-space(.), 'Sistema horario 24 horas')]/following::input[1]",
    ]
    for xpath in xpaths_probables:
        try:
            for elemento in driver.find_elements(By.XPATH, xpath):
                if not elemento.is_displayed() or not control_esta_habilitado(elemento):
                    continue
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento)
                driver.execute_script(
                    """
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
                    """,
                    elemento,
                    hora_objetivo,
                )
                time.sleep(0.4)
                valor_final = (elemento.get_attribute('value') or '').strip()
                if valor_final == hora_objetivo:
                    print(f"  [+] Hora de captura de coordenada: {valor_final}")
                    return True
        except Exception:
            pass

    print("  [!] No fue posible ajustar la hora de captura de ubicación")
    return False


def imprimir_campos_ubicacion():
    print("\n  [DEBUG CAMPOS UBICACION]:")

    try:
        for elemento in driver.find_elements(By.XPATH, "//input | //select"):
            try:
                elemento_id = elemento.get_attribute("id") or ""
                nombre = elemento.get_attribute("name") or ""
                tag = (elemento.tag_name or "").lower()
                tipo = elemento.get_attribute("type") or ""
                visible = elemento.is_displayed()
                if not visible:
                    continue
                if tag == 'select':
                    try:
                        valor = Select(elemento).first_selected_option.text.strip()
                    except Exception:
                        valor = ""
                elif tipo == 'checkbox' or tipo == 'radio':
                    valor = 'SI' if elemento.is_selected() else 'NO'
                else:
                    valor = (elemento.get_attribute("value") or "").strip()
                print(f"    {tag.upper()} id={elemento_id} name={nombre} type={tipo} visible={visible} value={valor}")
            except Exception:
                pass
    except Exception as e:
        print(f"    [!] No fue posible inspeccionar campos de ubicación: {e}")


def comparar_datos_ubicacion(datos_excel, datos_formulario):
    print("\n  [DATOS DE UBICACIÓN]:")
    print(f"    País residencia: {datos_formulario['pais_residencia']}")
    print(f"    Departamento residencia: {datos_formulario['departamento_residencia']}")
    print(f"    Municipio residencia: {datos_formulario['municipio_residencia']}")
    print(f"    Zona residencia: {datos_formulario['zona_residencia']}")
    print(f"    Centro poblado: {datos_formulario['centro_poblado']}")
    print(f"    Tipo cabecera: {datos_formulario['tipo_cabecera']}")
    print(f"    Barrio: {datos_formulario['barrio']}")
    print(f"    Observación ubicación: {datos_formulario['observacion_ubicacion']}")
    print(f"    Dirección resumen: {datos_formulario['direccion_resumen']}")
    print(f"    Teléfono: {datos_formulario['telefono']}")
    print(f"    Georeferenciado: {datos_formulario['georeferenciado']}")
    print(f"    Fecha captura: {datos_formulario['fecha_captura']}")
    print(f"    Hora captura: {datos_formulario.get('hora_captura', '')}")
    print(f"    Latitud: {datos_formulario['grados_latitud']} {datos_formulario['minutos_latitud']} {datos_formulario['segundos_latitud']}")
    print(f"    Longitud: {datos_formulario['grados_longitud']} {datos_formulario['minutos_longitud']} {datos_formulario['segundos_longitud']}")

    zona_excel = normalizar_texto(datos_excel['zona_residencia'])
    zona_form = normalizar_texto(datos_formulario['zona_residencia'])
    equivalencias_zona = {
        'URBANO': 'CABECERA',
        'URBANA': 'CABECERA',
        'RURAL': 'RESTO',
    }
    if zona_excel in equivalencias_zona:
        zona_excel = equivalencias_zona[zona_excel]

    comparacion = {
        'pais_residencia': normalizar_texto(datos_excel['pais_residencia']) == normalizar_texto(datos_formulario['pais_residencia']),
        'departamento_residencia': normalizar_texto(datos_excel['departamento_residencia']) == normalizar_texto(datos_formulario['departamento_residencia']),
        'municipio_residencia': normalizar_texto(datos_excel['municipio_residencia']) == normalizar_texto(datos_formulario['municipio_residencia']),
        'zona_residencia': zona_excel == zona_form,
        'direccion_resumen': direccion_residencia_coincide(datos_excel['direccion_residencia'], datos_formulario['direccion_resumen']),
        'telefono': normalizar_texto(datos_excel['telefono']) == normalizar_texto(datos_formulario['telefono']),
    }

    valores_excel = {
        'pais_residencia': datos_excel.get('pais_residencia', ''),
        'departamento_residencia': datos_excel.get('departamento_residencia', ''),
        'municipio_residencia': datos_excel.get('municipio_residencia', ''),
        'zona_residencia': datos_excel.get('zona_residencia', ''),
        'direccion_resumen': datos_excel.get('direccion_residencia', ''),
        'telefono': datos_excel.get('telefono', ''),
    }

    print("\n  [COMPARACIÓN UBICACIÓN]:")
    for llave, coincide in comparacion.items():
        estado = '✓' if coincide else '✗'
        print(f"    {estado} {llave}: Excel={valores_excel.get(llave, '')} | Form={datos_formulario.get(llave, '')}")

    return comparacion


def comparar_datos_basicos_excel_formulario(datos_excel, datos_formulario):
    sexo_excel = normalizar_sexo(datos_excel.get('sexo', ''))
    sexo_formulario = normalizar_sexo(datos_formulario.get('sexo', ''))
    return {
        'primer_nombre': datos_excel.get('primer_nombre', '').upper() == datos_formulario.get('primer_nombre', '').upper() if datos_formulario.get('primer_nombre') else False,
        'segundo_nombre': datos_excel.get('segundo_nombre', '').upper() == datos_formulario.get('segundo_nombre', '').upper() if datos_formulario.get('segundo_nombre') else False,
        'primer_apellido': datos_excel.get('primer_apellido', '').upper() == datos_formulario.get('primer_apellido', '').upper() if datos_formulario.get('primer_apellido') else False,
        'segundo_apellido': datos_excel.get('segundo_apellido', '').upper() == datos_formulario.get('segundo_apellido', '').upper() if datos_formulario.get('segundo_apellido') else False,
        'sexo': sexo_excel == sexo_formulario if datos_formulario.get('sexo') else False,
    }


def recolectar_mensajes_validacion_guardado():
    selectores_mensaje = [
        (By.XPATH, "//*[self::span or self::div or self::label or self::li][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'la operación no se completó satisfactoriamente') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
        (By.XPATH, "//*[self::span or self::div or self::label or self::li][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'se debe seleccionar') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
        (By.XPATH, "//*[self::span or self::div or self::label or self::li][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'debe diligenciar') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
        (By.XPATH, "//*[self::span or self::div or self::label or self::li][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'campo obligatorio') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
        (By.XPATH, "//*[self::span or self::div or self::label or self::li][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÑÁÉÍÓÚabcdefghijklmnopqrstuvwxyz', 'abcdefghijklmnopqrstuvwxyznaeiouabcdefghijklmnopqrstuvwxyz'), 'verifique') and (contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÑÁÉÍÓÚabcdefghijklmnopqrstuvwxyz', 'abcdefghijklmnopqrstuvwxyznaeiouabcdefghijklmnopqrstuvwxyz'), 'pestana') or contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÑÁÉÍÓÚabcdefghijklmnopqrstuvwxyz', 'abcdefghijklmnopqrstuvwxyznaeiouabcdefghijklmnopqrstuvwxyz'), 'ventana')) and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
        (By.XPATH, "//*[contains(@class, 'validation') or contains(@class, 'error') or contains(@class, 'danger') or contains(@class, 'alert-danger')][normalize-space()]"),
        (By.XPATH, "//*[@style[contains(., 'red')] or @style[contains(., '#ff')]][normalize-space()]")
    ]

    mensajes = []
    for selector in selectores_mensaje:
        try:
            for elemento in driver.find_elements(selector[0], selector[1]):
                if not elemento.is_displayed():
                    continue
                texto = " ".join((elemento.text or '').split()).strip()
                if not texto:
                    continue
                if len(texto) > 220:
                    texto = texto[:220] + "..."
                if texto not in mensajes:
                    mensajes.append(texto)
        except Exception:
            pass

    selectores_controles_invalidos = [
        (By.XPATH, "//input[@aria-invalid='true' or contains(@class, 'error') or contains(@class, 'invalid') or contains(@class, 'is-invalid') ]"),
        (By.XPATH, "//select[@aria-invalid='true' or contains(@class, 'error') or contains(@class, 'invalid') or contains(@class, 'is-invalid') ]"),
        (By.XPATH, "//textarea[@aria-invalid='true' or contains(@class, 'error') or contains(@class, 'invalid') or contains(@class, 'is-invalid') ]"),
    ]

    controles_invalidos = []
    for selector in selectores_controles_invalidos:
        try:
            for elemento in driver.find_elements(selector[0], selector[1]):
                if not elemento.is_displayed():
                    continue
                descripcion = ''
                element_id = (elemento.get_attribute('id') or '').strip()
                element_name = (elemento.get_attribute('name') or '').strip()
                if element_id:
                    try:
                        etiqueta = driver.find_element(By.XPATH, f"//label[@for='{element_id}']")
                        descripcion = " ".join((etiqueta.text or '').split()).strip()
                    except Exception:
                        pass
                descripcion = descripcion or element_id or element_name
                if descripcion:
                    mensaje_control = f"Campo con validación: {descripcion}"
                    if mensaje_control not in controles_invalidos:
                        controles_invalidos.append(mensaje_control)
        except Exception:
            pass

    if mensajes:
        return mensajes[:5]
    return controles_invalidos[:5]


def leer_estado_guardado():
    try:
        alerta = driver.switch_to.alert
        texto_alerta = (alerta.text or '').strip()
        if texto_alerta:
            try:
                alerta.accept()
            except Exception:
                pass
            return False, texto_alerta, 'error_alerta'
    except Exception:
        pass

    selectores_exito = [
        (By.XPATH, "//*[self::span or self::div or self::label][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'fue vinculado correctamente') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
    ]

    mensajes_exito = []
    for selector in selectores_exito:
        try:
            for elemento in driver.find_elements(selector[0], selector[1]):
                if not elemento.is_displayed():
                    continue
                texto = " ".join((elemento.text or '').split()).strip()
                if texto and texto not in mensajes_exito:
                    mensajes_exito.append(texto)
        except Exception:
            pass

    if mensajes_exito:
        return True, " | ".join(mensajes_exito[:2]), 'exito'

    selectores_ya_atendido = [
        (By.XPATH, "//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'ya esta') or contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'ya está')][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'atendido') and not(ancestor-or-self::*[contains(@style,'display:none')])]"),
    ]
    
    for selector in selectores_ya_atendido:
        try:
            elementos = driver.find_elements(selector[0], selector[1])
            for elemento in elementos:
                if elemento.is_displayed():
                    texto = " ".join((elemento.text or '').split()).strip()
                    if texto:
                        texto_norm = normalizar_texto(texto)
                        if (
                            'otra unidad' in texto_norm
                            or 'otra uds' in texto_norm
                            or 'otra unidad de atencion' in texto_norm
                            or 'otra unidad de atención' in texto_norm
                        ):
                            return False, f"✓ {texto}", 'ya_atendido_otra_unidad'
                        return False, texto, 'ya_atendido_mismo_servicio'
        except Exception:
            pass

    mensajes = recolectar_mensajes_validacion_guardado()

    if mensajes:
        return False, " | ".join(mensajes[:5]), 'error_validacion'

    return False, "No apareció el mensaje verde de confirmación de guardado", 'sin_mensaje_verde'


def confirmar_guardado_sin_mensaje_verde(contexto):
    try:
        esperar_postback_finalizado(2.0)
    except Exception:
        pass

    indicadores = []

    try:
        boton_nuevo = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "btnNuevo")))
        if boton_nuevo.is_displayed() and control_esta_habilitado(boton_nuevo):
            indicadores.append("botón '+' disponible")
    except Exception:
        pass

    try:
        boton_guardar = driver.find_element(By.ID, "btnGuardar")
        if boton_guardar.is_displayed():
            indicadores.append("disquete visible")
    except Exception:
        pass

    if contexto == "final" and indicadores:
        return True, f"Guardado ejecutado sin mensaje verde ni errores visibles; estado estable detectado ({', '.join(indicadores[:2])})"

    return False, "No apareció el mensaje verde de confirmación de guardado"


def puede_guardar_final_segun_excel(coincidencias, ubicacion_ok, grupo_familiar_ok):
    if not all(coincidencias.values()):
        return False, "Se omitió el guardado final porque los datos básicos no coinciden completamente con el Excel"
    if not ubicacion_ok:
        return False, "Se omitió el guardado final porque los datos de ubicación no coinciden completamente con el Excel"
    if not grupo_familiar_ok:
        return False, "Se omitió el guardado final porque el grupo familiar no coincide completamente con el Excel"
    return True, "Listo para guardar"


def autocorregir_pais_nacimiento_grupo_familiar(datos_excel, force_update=False):
    """Abre cada miembro del grupo familiar y completa pais/depto/municipio si faltan.

    Cuando force_update=True, vuelve a grabar los integrantes identificables desde Excel
    aunque los selects ya parezcan diligenciados, para estabilizar postbacks del portal.
    """
    print("  [~] Auto-corrección: verificando pais/depto/municipio en cada integrante del grupo familiar...")
    if not abrir_tab_grupo_familiar():
        return False
    filas = leer_filas_grupo_familiar()
    corregido_alguno = False
    beneficiario_doc = normalizar_texto(datos_excel.get('documento', ''))
    for fila in filas:
        doc = fila.get('documento', '')
        if not doc:
            continue
        if not abrir_detalle_persona_grupo_familiar(documento=doc, requiere_edicion=True):
            print(f"    [!] {doc}: no se pudo abrir detalle para validar pais_nacimiento")
            continue
        try:
            select_pais = driver.find_element(By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdPaisNacimiento")
            pais_val = Select(select_pais).first_selected_option.get_attribute("value")
            pais_texto = Select(select_pais).first_selected_option.text.strip()
        except Exception:
            continue

        departamento_actual = obtener_texto_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdDepartamento"
        )
        municipio_actual = obtener_texto_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdMunicipioNacimiento"
        )
        pais_es_colombia = texto_select_coincide(pais_texto, 'COLOMBIA')
        ubicacion_incompleta = pais_es_colombia and (
            texto_select_coincide(departamento_actual, 'Seleccione')
            or texto_select_coincide(municipio_actual, 'Seleccione')
        )

        padre_doc = normalizar_texto(valor_excel_familia(datos_excel.get('padre_documento', '')))
        madre_doc = normalizar_texto(valor_excel_familia(datos_excel.get('madre_documento', '')))
        responsable_doc = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', '')))
        doc_norm = normalizar_texto(doc)

        # Verificar también el sexo (puede haberse borrado por postback de ASP.NET)
        sexo_esperado = ''
        if doc_norm == beneficiario_doc:
            sexo_esperado = normalizar_sexo(datos_excel.get('sexo', ''))
        elif doc_norm == padre_doc:
            sexo_esperado = 'MASCULINO'
        elif doc_norm == madre_doc:
            sexo_esperado = 'FEMENINO'
        elif doc_norm == responsable_doc:
            sexo_esperado = determinar_sexo_responsable_desde_excel(datos_excel)
        try:
            sel_sexo = driver.find_element(By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_ddlIdSexo")
            sexo_val_actual = Select(sel_sexo).first_selected_option.get_attribute("value")
        except Exception:
            sexo_val_actual = None
        sexo_ok = (not sexo_esperado) or (sexo_val_actual and sexo_val_actual != "-1")

        persona_datos = None
        if doc_norm == beneficiario_doc:
            persona_datos = construir_beneficiario_grupo_familiar_desde_excel(datos_excel)
        elif doc_norm == padre_doc:
            persona_datos = construir_persona_desde_excel(datos_excel, 'padre')
        elif doc_norm == madre_doc:
            persona_datos = construir_persona_desde_excel(datos_excel, 'madre')
        elif doc_norm == responsable_doc:
            persona_datos = construir_persona_desde_excel(datos_excel, 'responsable')

        if pais_val and pais_val != "-1" and sexo_ok and not ubicacion_incompleta and not (force_update and persona_datos):
            print(f"    [=] {doc}: pais '{pais_texto}', ubicación y sexo OK (sin cambios)")
            continue

        # Pais/ubicación/sexo vacíos → completar
        necesita_ubicacion = force_update or not (pais_val and pais_val != "-1") or ubicacion_incompleta
        print(f"    [*] {doc}: completando campos faltantes (ubicacion={necesita_ubicacion}, sexo={not sexo_ok}, forzar={force_update})...")
        if not persona_datos:
            persona_datos = {'pais_nacimiento': 'COLOMBIA', 'departamento_nacimiento': '', 'municipio_nacimiento': '', 'sexo': ''}

        if not persona_datos:
            continue

        if necesita_ubicacion:
            completar_ubicacion_nacimiento_grupo_familiar(
                persona_datos,
                prefijo_log="Grupo familiar (auto-corrección)",
            )

        # Sexo DESPUÉS de pais/depto/municipio (igual que en llenar_formulario_persona_grupo_familiar)
        if not sexo_ok and sexo_esperado:
            asignar_select_por_id(
                "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_ddlIdSexo",
                sexo_esperado,
                "sexo (auto-corrección)",
            )
            esperar_postback_finalizado(0.6)

        # Actualizar persona
        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
            try:
                btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
            print(f"    [+] {doc}: pais completado y persona actualizada")
            esperar_postback_finalizado(max(0.8, ESPERA_ACTUALIZAR_PERSONA))
            esperar_grupo_familiar_listo(1.2)
            corregido_alguno = True
        except Exception as e:
            print(f"    [!] {doc}: no se pudo actualizar persona: {e}")

    return corregido_alguno


def autocorregir_jefe_grupo_familiar(datos_excel, jefe_forzado=""):
    """Corrige cuando no hay jefe o hay más de uno."""
    print("  [~] Auto-corrección: revisando jefe del grupo familiar...")
    if not abrir_tab_grupo_familiar():
        return False
    filas = leer_filas_grupo_familiar()
    jefes = [f for f in filas if normalizar_texto(f.get('parentesco_jefe', '')) == normalizar_texto('JEFE DEL GRUPO FAMILIAR')]

    jefe_objetivo = normalizar_texto(jefe_forzado)
    if jefe_objetivo not in {'padre', 'madre'}:
        jefe_objetivo = determinar_jefe_objetivo(datos_excel)
    persona_jefe = construir_persona_desde_excel(datos_excel, jefe_objetivo) if jefe_objetivo in ('padre', 'madre', 'responsable') else None

    doc_beneficiario = normalizar_texto(datos_excel.get('documento', ''))
    doc_padre = normalizar_texto(valor_excel_familia(datos_excel.get('padre_documento', '')))
    doc_madre = normalizar_texto(valor_excel_familia(datos_excel.get('madre_documento', '')))
    doc_responsable = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', ''))) if responsable_adicional_es_valido(datos_excel) else ''
    parentesco_padre_objetivo, parentesco_madre_objetivo, _ = calcular_parentescos_padres_objetivo(datos_excel)

    def parentesco_correcto_para_documento(documento_normalizado):
        if documento_normalizado == doc_beneficiario:
            return determinar_parentesco_beneficiario_desde_excel(datos_excel)
        if documento_normalizado == doc_padre:
            return parentesco_padre_objetivo
        if documento_normalizado == doc_madre:
            return parentesco_madre_objetivo
        if documento_normalizado == doc_responsable:
            return 'JEFE DEL GRUPO FAMILIAR' if jefe_objetivo == 'responsable' else 'PARIENTE U OTRO'
        return 'CONYUGE/COMPAÑERO(A)'

    def cambiar_parentesco(documento_raw, parentesco_nuevo):
        if not documento_raw or not parentesco_nuevo:
            return False
        print(f"    [*] {documento_raw}: ajustando parentesco a {parentesco_nuevo}...")
        if not abrir_detalle_persona_grupo_familiar(documento=documento_raw, requiere_edicion=True):
            return False
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
            parentesco_nuevo,
            "parentesco_jefe (auto-corrección)",
        )
        esperar_postback_finalizado(1.0)
        esperar_detalle_grupo_familiar(1.0)
        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
            driver.execute_script("arguments[0].click();", btn)
            esperar_postback_finalizado(max(0.8, ESPERA_ACTUALIZAR_PERSONA))
            esperar_grupo_familiar_listo(1.2)
            print(f"    [+] {documento_raw}: parentesco corregido a {parentesco_nuevo}")
            return True
        except Exception as e:
            print(f"    [!] {documento_raw}: no se pudo actualizar: {e}")
            return False

    if persona_jefe or len(jefes) > 1:
        doc_jefe_correcto = normalizar_texto(persona_jefe['documento']) if persona_jefe else ''
        if len(jefes) > 1 and not doc_jefe_correcto:
            # Fallback defensivo: si no se pudo inferir jefe objetivo desde Excel,
            # conservar temporalmente el primer jefe actual y degradar los demás.
            doc_jefe_correcto = normalizar_texto(jefes[0].get('documento', ''))
            print(f"    [~] No se pudo inferir jefe objetivo por Excel; se conservará provisionalmente como jefe el documento {doc_jefe_correcto}")
        jefe_actual_correcto = any(normalizar_texto(f.get('documento', '')) == doc_jefe_correcto for f in jefes)

        if len(jefes) > 1 or (len(jefes) == 1 and not jefe_actual_correcto):
            corregido = False
            for fila_extra in jefes:
                doc_extra = normalizar_texto(fila_extra.get('documento', ''))
                if doc_extra == doc_jefe_correcto:
                    continue
                corregido = cambiar_parentesco(fila_extra.get('documento', ''), parentesco_correcto_para_documento(doc_extra)) or corregido

            fila_jefe_correcto, _ = buscar_fila_grupo_familiar_por_documento(doc_jefe_correcto)
            parentesco_actual_jefe_correcto = normalizar_texto(fila_jefe_correcto.get('parentesco_jefe', '')) if fila_jefe_correcto else ''
            if parentesco_actual_jefe_correcto != normalizar_texto('JEFE DEL GRUPO FAMILIAR'):
                corregido = cambiar_parentesco(doc_jefe_correcto, 'JEFE DEL GRUPO FAMILIAR') or corregido

            return corregido

    if len(jefes) == 0 and persona_jefe:
        # Sin jefe: buscar la persona correcta en la tabla y cambiar su parentesco
        doc_jefe = normalizar_texto(persona_jefe['documento'])
        for fila in filas:
            if normalizar_texto(fila.get('documento', '')) == doc_jefe:
                print(f"    [*] {doc_jefe}: 0 jefes en grilla → asignando JEFE DEL GRUPO FAMILIAR...")
                if cambiar_parentesco(persona_jefe['documento'], 'JEFE DEL GRUPO FAMILIAR'):
                    return True
                break
    return False


def autocorregir_parentesco_beneficiario(datos_excel):
    """Corrige parentesco del beneficiario en grupo familiar si está vacío."""
    print("  [~] Auto-corrección: corrigiendo parentesco del beneficiario en grupo familiar...")
    if not abrir_tab_grupo_familiar():
        return False
    doc_ben = datos_excel.get('documento', '')
    parentesco_esperado = determinar_parentesco_beneficiario_desde_excel(datos_excel)
    if not doc_ben or not parentesco_esperado:
        return False
    fila, _ = buscar_fila_grupo_familiar_por_documento(doc_ben)
    if not fila:
        return False
    parentesco_actual = normalizar_texto(fila.get('parentesco_jefe', ''))
    if parentesco_actual == normalizar_texto(parentesco_esperado):
        return False
    print(f"    [*] Beneficiario {doc_ben}: parentesco actual='{fila.get('parentesco_jefe','')}' esperado='{parentesco_esperado}'")
    if abrir_detalle_persona_grupo_familiar(documento=doc_ben, requiere_edicion=True):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
            parentesco_esperado,
            "parentesco beneficiario (auto-corrección)",
        )
        esperar_postback_finalizado(1.0)
        esperar_detalle_grupo_familiar(1.0)
        try:
            btn = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
            driver.execute_script("arguments[0].click();", btn)
            esperar_postback_finalizado(max(0.8, ESPERA_ACTUALIZAR_PERSONA))
            esperar_grupo_familiar_listo(1.2)
            print(f"    [+] Beneficiario {doc_ben}: parentesco corregido a {parentesco_esperado}")
            return True
        except Exception as e:
            print(f"    [!] No se pudo actualizar parentesco beneficiario: {e}")
    return False


def autocorregir_parentesco_filas_vacias(datos_excel):
    """Corrige cualquier fila del grupo familiar cuyo parentesco_jefe esté vacío.
    Se usa cuando la validación dice 'Parentesco con respecto al Jefe del Grupo es obligatorio'
    para personas distintas del propio beneficiario."""
    print("  [~] Auto-corrección: buscando filas con parentesco vacío en grupo familiar...")
    if not abrir_tab_grupo_familiar():
        return False
    filas = leer_filas_grupo_familiar()
    corregido = False
    doc_beneficiario = normalizar_texto(datos_excel.get('documento', ''))
    doc_padre = normalizar_texto(valor_excel_familia(datos_excel.get('padre_documento', '')))
    doc_madre = normalizar_texto(valor_excel_familia(datos_excel.get('madre_documento', '')))
    doc_responsable = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', '')))
    responsable_obj = determinar_responsable_objetivo(datos_excel)
    parentesco_padre_objetivo, parentesco_madre_objetivo, jefe_objetivo = calcular_parentescos_padres_objetivo(datos_excel)

    for fila in filas:
        doc_fila_raw = fila.get('documento', '') or ''
        doc_fila = normalizar_texto(doc_fila_raw)
        parentesco_actual = fila.get('parentesco_jefe', '').strip()
        if parentesco_actual:
            continue  # Ya tiene parentesco, no tocar

        # Determinar parentesco correcto según quién es
        if doc_fila == doc_beneficiario:
            parentesco_nuevo = determinar_parentesco_beneficiario_desde_excel(datos_excel)
        elif doc_padre and doc_fila == doc_padre:
            parentesco_nuevo = parentesco_padre_objetivo
        elif doc_madre and doc_fila == doc_madre:
            parentesco_nuevo = parentesco_madre_objetivo
        elif responsable_obj == 'responsable' and doc_responsable and doc_fila == doc_responsable:
            parentesco_nuevo = 'JEFE DEL GRUPO FAMILIAR' if jefe_objetivo == 'responsable' else 'PARIENTE U OTRO'
        else:
            print(f"  [!] Fila {doc_fila_raw}: parentesco vacío pero no identificable en Excel, se omite")
            continue

        print(f"  [*] Fila {doc_fila_raw}: parentesco vacío → asignando '{parentesco_nuevo}'")
        if abrir_detalle_persona_grupo_familiar(documento=doc_fila_raw, requiere_edicion=True):
            asignar_select_por_id(
                "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
                parentesco_nuevo,
                "parentesco fila vacía (auto-corrección)",
            )
            esperar_postback_finalizado(1.0)
            esperar_detalle_grupo_familiar(1.0)
            try:
                btn = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
                driver.execute_script("arguments[0].click();", btn)
                esperar_postback_finalizado(max(0.8, ESPERA_ACTUALIZAR_PERSONA))
                esperar_grupo_familiar_listo(1.2)
                print(f"  [+] Fila {doc_fila_raw}: parentesco corregido a {parentesco_nuevo}")
                corregido = True
            except Exception as e:
                print(f"  [!] No se pudo actualizar fila {doc_fila_raw}: {e}")
    return corregido


def autocorregir_campos_otras_pestanas(datos_excel):
    """Corrector para 'verifique las demás pestañas o ventanas'.
    Cierra cualquier detalle abierto en grupo familiar (que puede tener campos
    requeridos vacíos) y re-verifica pais/nacimiento de cada integrante."""
    print("  [~] Auto-corrección: cerrando formulario detalle y revisando campos en todas las pestañas...")
    corregido = False

    corregido = completar_pertenencia_etnica(datos_excel) or corregido

    # 1) Ir a grupo familiar y limpiar el subformulario (puede tener campos en rojo)
    if abrir_tab_grupo_familiar():
        if limpiar_formulario_grupo_familiar():
            print("  [+] Formulario detalle de grupo familiar limpiado (campos en rojo cerrados)")
            corregido = True
        time.sleep(1)

    # 2) Re-verificar pais/depto/municipio en cada integrante (causa frecuente)
    corregido = autocorregir_pais_nacimiento_grupo_familiar(datos_excel) or corregido

    return corregido


def autocorregir_desde_error(mensaje_error, datos_excel):
    """Dispatcher: detecta el tipo de error y llama al corrector apropiado.
    Retorna True si se corrigió algo (vale la pena reintentar guardar)."""
    msg = normalizar_texto(mensaje_error)
    msg_raw = (mensaje_error or "").lower()
    corregido = False

    # Manejo de "verifique las demás pestañas o ventanas"
    if ('verifique' in msg) and ('pestana' in msg or 'ventana' in msg):
        corregido = autocorregir_campos_otras_pestanas(datos_excel) or corregido

    if (
        ('pais' in msg and 'nacimiento' in msg)
        or ('país' in msg_raw and 'nacimiento' in msg_raw)
        or ('pais' in msg_raw and 'nacimiento' in msg_raw)
    ):
        corregido = autocorregir_pais_nacimiento_grupo_familiar(datos_excel, force_update=True) or corregido

    if 'mas de un jefe' in msg or 'existe mas de un jefe' in msg:
        corregido = autocorregir_jefe_grupo_familiar(datos_excel) or corregido

    if 'no existe un jefe' in msg or 'ingrese los datos del jefe' in msg:
        # Intentar asignar jefe según Excel
        corregido = autocorregir_jefe_grupo_familiar(datos_excel) or corregido

    if 'parentesco con respecto al jefe' in msg and 'obligatorio' in msg:
        corregido = autocorregir_parentesco_beneficiario(datos_excel) or corregido
        corregido = autocorregir_parentesco_filas_vacias(datos_excel) or corregido

    if ('debe seleccionar la lupa' in msg) or ('seleccionar la lupa' in msg):
        corregido = autocorregir_pais_nacimiento_grupo_familiar(datos_excel, force_update=True) or corregido
        corregido = autocorregir_parentesco_filas_vacias(datos_excel) or corregido

    if 'zona ubicacion' in msg or 'zona ubicación' in msg or 'nombrezonaresto' in msg.lower():
        if abrir_tab_ubicacion():
            corregido = corregir_datos_ubicacion(datos_excel, es_alta_nueva=True) is not None or corregido
            corregido = asegurar_zona_ubicacion_requerida(datos_excel) or corregido

    if (
        'reside en el lugar' in msg
        or 'consejo comunitario' in msg
        or 'lengua propia' in msg
        or 'idioma o lengua' in msg
    ):
        if es_valor_ausente(datos_excel.get('territorio_etnico', '')):
            if abrir_tab_pertenencia_etnica() or True:
                corregido = forzar_grupo_etnico_no_autoreconoce() or corregido

    if 'grupo etnico' in msg or 'grupo étnico' in (mensaje_error or '').lower() or 'pertenencia etnica' in msg or 'pertenencia étnica' in (mensaje_error or '').lower():
        if es_valor_ausente(datos_excel.get('territorio_etnico', '')):
            corregido = forzar_grupo_etnico_no_autoreconoce() or corregido
        else:
            corregido = completar_pertenencia_etnica(datos_excel) or corregido

    if 'municipio' in msg and 'nacimiento' in msg:
        # El municipio falta; el corrector de pais también lo rellena
        corregido = autocorregir_pais_nacimiento_grupo_familiar(datos_excel, force_update=True) or corregido

    if not corregido:
        print(f"  [~] Error no reconocido para auto-corrección: {mensaje_error[:200]}")

    return corregido


def guardar_formulario(exigir_mensaje_verde=False, contexto="inicial"):
    """Guarda el formulario usando el disquete superior.

    En RUB el mensaje verde suele aparecer solo en el guardado final,
    así que el guardado inicial acepta la ausencia de ese mensaje siempre
    que no haya alertas o mensajes de validación visibles.
    """
    print(f"  [*] Guardando formulario ({contexto})...")

    selectores_guardar = [
        (By.ID, "btnGuardar"),
        (By.XPATH, "//a[@id='btnGuardar']"),
        (By.XPATH, "//img[@title='Guardar']/parent::a"),
        (By.XPATH, "//img[@alt='Guardar']/parent::a"),
    ]

    for selector in selectores_guardar:
        try:
            boton_guardar = WebDriverWait(driver, 60).until(EC.presence_of_element_located((selector[0], selector[1])))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_guardar)
            print("  [*] Esperando que el disquete quede habilitado...")
            time.sleep(ESPERA_HABILITAR_GUARDAR)
            try:
                boton_guardar.click()
            except Exception:
                try:
                    ActionChains(driver).move_to_element(boton_guardar).click().perform()
                except Exception:
                    driver.execute_script("arguments[0].click();", boton_guardar)
            print(f"  [*] Esperando {ESPERA_POST_GUARDADO} segundos para verificar el guardado...")
            time.sleep(ESPERA_POST_GUARDADO)
            guardado_ok, mensaje_guardado, motivo = leer_estado_guardado()
            
            # Casos especiales permitidos: ya está siendo atendido
            # (en otra unidad o en el mismo servicio)
            if motivo in ('ya_atendido_otra_unidad', 'ya_atendido_mismo_servicio'):
                print(f"  [~] Guardado parcial ({contexto}): {mensaje_guardado}")
                print(f"      Los cambios verdes se confirmaron; beneficiario ya en el sistema.")
                return True, mensaje_guardado
            
            if not guardado_ok:
                if (not exigir_mensaje_verde) and motivo == 'sin_mensaje_verde' and contexto == 'inicial':
                    mensaje_guardado = "Guardado inicial sin mensaje verde; se continúa solo para poder validar pestañas, pero no cuenta como confirmación final"
                    print(f"  [~] Guardado provisional ({contexto}): {mensaje_guardado}")
                    return True, mensaje_guardado
                print(f"  [!] El formulario no confirmó guardado exitoso: {mensaje_guardado}")
                return False, mensaje_guardado
            print(f"  [+] Guardado realizado: {mensaje_guardado}")
            return True, mensaje_guardado
        except Exception:
            pass

    print("  [!] No se encontró el botón guardar (disquete)")
    return False, "No se encontró el botón guardar (disquete)"


def iniciar_nuevo_registro():
    print("  [*] Pulsando '+' para iniciar el siguiente niño...")

    try:
        boton_nuevo = wait.until(EC.element_to_be_clickable((By.ID, "btnNuevo")))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_nuevo)
        time.sleep(0.5)
        try:
            boton_nuevo.click()
        except Exception:
            try:
                ActionChains(driver).move_to_element(boton_nuevo).click().perform()
            except Exception:
                driver.execute_script("arguments[0].click();", boton_nuevo)
        time.sleep(ESPERA_NUEVO_REGISTRO)
        if not aplicar_filtros_formulario():
            return False
        print("  [+] Formulario listo para el siguiente niño")
        return True
    except Exception as e:
        print(f"  [!] No fue posible pulsar el '+': {e}")
        return False


def campos_fallidos(diccionario):
    return [llave for llave, valor in diccionario.items() if not valor]


def resumir_error(error):
    texto = (error or "").strip()
    if not texto:
        return "Sin detalle adicional"
    primera_linea = texto.splitlines()[0].strip()
    return primera_linea[:300]


def valor_reportable(valor):
    return "" if es_valor_ausente(valor) else (valor or "")


def construir_observaciones_resultado(resultado):
    observaciones = []

    observaciones.extend(detalle_basico_para_reporte(resultado))
    observaciones.extend(detalle_ubicacion_para_reporte(resultado))
    observaciones.extend(detalle_grupo_familiar_para_reporte(resultado))
    observaciones.extend(detalle_grupo_familiar_informativo_para_reporte(resultado))

    documento = resultado.get('documento', '')
    if not resultado.get('foto_cargada', True):
        observaciones.append(f"- Foto: no se encontró o no se pudo cargar para el documento {documento}")
    if not resultado.get('siguiente_nino_listo', True):
        observaciones.append("- Siguiente niño: el formulario no quedó listo tras pulsar '+'")
    telefono_original = texto_excel(resultado.get('excel_ubicacion', {}).get('telefono_original_excel', ''))
    telefono_usado = texto_excel(resultado.get('excel_ubicacion', {}).get('telefono', ''))
    if telefono_original and telefono_usado and telefono_original != telefono_usado:
        observaciones.append(f"- Ubicación telefono_normalizado: Excel original='{telefono_original}' | Teléfono usado='{telefono_usado}'")

    return observaciones


def construir_errores_resultado(resultado):
    errores = []

    if not resultado.get('guardado', True):
        errores.append(f"- Guardado inicial: no se completó ({resultado.get('guardado_mensaje', 'sin detalle')})")
    if not resultado.get('guardado_final', True):
        errores.append(f"- Guardado final: no se completó ({resultado.get('guardado_final_mensaje', 'sin detalle')})")
    if resultado.get('error'):
        errores.append(f"- Error técnico: {resumir_error(resultado['error'])}")

    return errores


def codigos_observacion_resultado(resultado):
    codigos = []
    detalles_grupo = resultado.get('detalles_grupo_familiar') or {}

    if not resultado.get('coincide', False):
        codigos.extend([f"datos_basicos:{campo}" for campo in campos_fallidos(resultado.get('detalles', {}))])
    if not resultado.get('ubicacion_ok', False):
        codigos.extend([f"ubicacion:{campo}" for campo in campos_fallidos(resultado.get('detalles_ubicacion', {}))])
    if detalles_grupo and not resultado.get('grupo_familiar_ok', False):
        codigos.extend([f"grupo_familiar:{campo}" for campo in campos_fallidos({
            'padre_encontrado': detalles_grupo.get('padre_encontrado', True),
            'madre_encontrada': detalles_grupo.get('madre_encontrada', True),
            'responsable_extra_encontrado': detalles_grupo.get('responsable_extra_encontrado', True),
            'jefe_ok': detalles_grupo.get('jefe_ok', False),
            'padre_parentesco_ok': detalles_grupo.get('padre_parentesco_ok', True),
            'madre_parentesco_ok': detalles_grupo.get('madre_parentesco_ok', True),
            'beneficiario_parentesco_jefe_ok': detalles_grupo.get('beneficiario_parentesco_jefe_ok', False),
            'responsable_ok': detalles_grupo.get('responsable_ok', False),
        })])
    if 'foto_cargada' in resultado and not resultado.get('foto_cargada', False):
        codigos.append('foto:no_cargada')
    if 'siguiente_nino_listo' in resultado and not resultado.get('siguiente_nino_listo', True):
        codigos.append('siguiente_nino:no_listo')
    telefono_original = texto_excel(resultado.get('excel_ubicacion', {}).get('telefono_original_excel', ''))
    telefono_usado = texto_excel(resultado.get('excel_ubicacion', {}).get('telefono', ''))
    if telefono_original and telefono_usado and telefono_original != telefono_usado:
        codigos.append('ubicacion:telefono_normalizado')

    incidencias_grupo = detalles_grupo.get('incidencias', [])
    for incidencia in incidencias_grupo:
        if incidencia_grupo_familiar_es_bloqueante(incidencia):
            codigos.append('grupo_familiar:incidencias')
            break

    return codigos


def codigos_error_resultado(resultado):
    codigos = []
    if not resultado.get('guardado', True):
        codigos.append('guardado_inicial')
    if not resultado.get('guardado_final', True):
        codigos.append('guardado_final')
    if resultado.get('error'):
        codigos.append(f"error:{resumir_error(resultado['error'])}")
    return codigos


def estado_general_resultado(resultado):
    if codigos_error_resultado(resultado):
        return 'ERROR'
    if construir_observaciones_resultado(resultado):
        return 'CON OBSERVACIONES'
    return 'OK'


def detalle_basico_para_reporte(resultado):
    lineas = []
    detalles = resultado.get('detalles', {})
    esperado = resultado.get('excel_basico', {})
    encontrado = resultado.get('formulario_basico', {})

    etiquetas = {
        'primer_nombre': 'Primer nombre',
        'segundo_nombre': 'Segundo nombre',
        'primer_apellido': 'Primer apellido',
        'segundo_apellido': 'Segundo apellido',
        'sexo': 'Sexo',
    }

    for campo, ok in detalles.items():
        valor_esperado = valor_reportable(esperado.get(campo, ''))
        if not ok and valor_esperado:
            lineas.append(
                f"- {etiquetas.get(campo, campo)}: Excel='{valor_esperado}' | Formulario='{encontrado.get(campo, '')}'"
            )

    return lineas


def detalle_ubicacion_para_reporte(resultado):
    lineas = []
    detalles = resultado.get('detalles_ubicacion', {})
    esperado = resultado.get('excel_ubicacion', {})
    encontrado = resultado.get('formulario_ubicacion', {})

    for campo, ok in detalles.items():
        if not ok:
            esperado_valor = FECHA_CAPTURA_UBICACION if campo == 'fecha_captura' else HORA_CAPTURA_UBICACION if campo == 'hora_captura' else valor_reportable(esperado.get(campo, ''))
            encontrado_valor = encontrado.get(campo, '')
            if esperado_valor:
                lineas.append(f"- Ubicación {campo}: Excel='{esperado_valor}' | Formulario='{encontrado_valor}'")

    return lineas


def detalle_grupo_familiar_para_reporte(resultado):
    lineas = []
    detalles = resultado.get('detalles_grupo_familiar', {})
    padre_requerido = detalles.get('padre_requerido', bool(detalles.get('padre_esperado', '')))
    madre_requerida = detalles.get('madre_requerida', bool(detalles.get('madre_esperada', '')))

    if not detalles.get('jefe_ok', True):
        lineas.append("- Grupo familiar: no quedó una persona marcada como JEFE DEL GRUPO FAMILIAR")
    if padre_requerido and not detalles.get('padre_encontrado', True) and not detalles.get('padre_omitido_duplicidad', False):
        lineas.append(f"- Padre no encontrado: '{detalles.get('padre_esperado', '')}'")
    if madre_requerida and not detalles.get('madre_encontrada', True) and not detalles.get('madre_omitida_duplicidad', False):
        lineas.append(f"- Madre no encontrada: '{detalles.get('madre_esperada', '')}'")
    if not detalles.get('responsable_extra_encontrado', True) and detalles.get('responsable_extra_esperado', ''):
        lineas.append(f"- Responsable adicional no encontrado: '{detalles.get('responsable_extra_esperado', '')}'")
    if padre_requerido and not detalles.get('padre_parentesco_ok', True) and not detalles.get('padre_omitido_duplicidad', False):
        lineas.append(
            f"- Parentesco padre: esperado='{detalles.get('padre_parentesco_esperado', '')}' | final='{detalles.get('padre_parentesco_final', '')}'"
        )
    if madre_requerida and not detalles.get('madre_parentesco_ok', True) and not detalles.get('madre_omitida_duplicidad', False):
        lineas.append(
            f"- Parentesco madre: esperado='{detalles.get('madre_parentesco_esperado', '')}' | final='{detalles.get('madre_parentesco_final', '')}'"
        )
    if not detalles.get('beneficiario_parentesco_jefe_ok', True):
        lineas.append(
            f"- Beneficiario: parentesco con respecto al jefe esperado='{detalles.get('beneficiario_parentesco_jefe_esperado', '')}' | final='{detalles.get('beneficiario_parentesco_jefe_final', '')}'"
        )
    if not detalles.get('responsable_ok', True):
        lineas.append("- Grupo familiar: no quedó una persona marcada como responsable")
    for fila in detalles.get('filas_incompletas', []):
        lineas.append(
            f"- Fila incompleta grupo familiar: documento='{fila.get('documento', '')}' nombre='{fila.get('nombre', '')}' problemas={','.join(fila.get('problemas', []))}"
        )
    for incidencia in detalles.get('incidencias', []):
        if incidencia_grupo_familiar_es_bloqueante(incidencia):
            lineas.append(f"- {incidencia}")

    return lineas


def detalle_grupo_familiar_informativo_para_reporte(resultado):
    lineas = []
    detalles = resultado.get('detalles_grupo_familiar', {})

    if detalles.get('padre_omitido_duplicidad', False):
        lineas.append(f"- Grupo familiar: padre omitido por duplicidad de tipo de documento y se conservó el resto de la configuración ({detalles.get('padre_duplicidad_mensaje', '')})")
    if detalles.get('madre_omitida_duplicidad', False):
        lineas.append(f"- Grupo familiar: madre omitida por duplicidad de tipo de documento y se conservó el resto de la configuración ({detalles.get('madre_duplicidad_mensaje', '')})")

    return lineas


def incidencia_grupo_familiar_es_bloqueante(incidencia):
    texto = normalizar_texto(incidencia)
    if texto.startswith(normalizar_texto("Padre omitido por duplicidad de tipo de documento:")):
        return False
    if texto.startswith(normalizar_texto("Madre omitida por duplicidad de tipo de documento:")):
        return False
    return True


def grupo_familiar_esta_ok(detalles):
    incidencias_bloqueantes = [
        incidencia for incidencia in detalles.get('incidencias', [])
        if incidencia_grupo_familiar_es_bloqueante(incidencia)
    ]

    padre_requerido = detalles.get('padre_requerido', bool(detalles.get('padre_esperado', '')))
    madre_requerida = detalles.get('madre_requerida', bool(detalles.get('madre_esperada', '')))

    padre_resuelto = (not padre_requerido) or detalles.get('padre_encontrado') or detalles.get('padre_omitido_duplicidad')
    madre_resuelta = (not madre_requerida) or detalles.get('madre_encontrada') or detalles.get('madre_omitida_duplicidad')
    padre_parentesco_resuelto = (not padre_requerido) or detalles.get('padre_parentesco_ok') or detalles.get('padre_omitido_duplicidad')
    madre_parentesco_resuelto = (not madre_requerida) or detalles.get('madre_parentesco_ok') or detalles.get('madre_omitida_duplicidad')

    return all([
        detalles.get('jefe_ok'),
        padre_resuelto,
        madre_resuelta,
        padre_parentesco_resuelto,
        madre_parentesco_resuelto,
        detalles.get('beneficiario_parentesco_jefe_ok'),
        detalles.get('responsable_ok'),
        not incidencias_bloqueantes,
    ])


def error_es_reintentable_registro(error):
    texto = normalizar_texto(str(error))
    return any(fragmento in texto for fragmento in [
        'STALE ELEMENT',
        'STALEELEMENTREFERENCE',
        'ELEMENT CLICK INTERCEPTED',
    ])


def detalle_logros_para_reporte(resultado):
    lineas = []

    if resultado.get('coincide', False):
        lineas.append("- Datos básicos: verificados/corregidos")
    if resultado.get('beneficiario_creado_desde_excel', False):
        lineas.append("- Datos básicos: beneficiario creado desde Excel tras búsqueda vacía")
    if resultado.get('foto_cargada', False):
        lineas.append("- Foto: cargada correctamente")
    if resultado.get('guardado', False):
        lineas.append(f"- Guardado inicial: completado ({resultado.get('guardado_mensaje', 'sin detalle')})")
    if resultado.get('ubicacion_ok', False):
        lineas.append("- Datos de ubicación: verificados/corregidos")
    if resultado.get('grupo_familiar_ok', False):
        lineas.append("- Grupo familiar: verificado/corregido")
    for linea in detalle_grupo_familiar_informativo_para_reporte(resultado):
        lineas.append(linea)
    if resultado.get('guardado_final', False):
        lineas.append(f"- Guardado final: completado ({resultado.get('guardado_final_mensaje', 'sin detalle')})")
    if resultado.get('siguiente_nino_listo', False):
        lineas.append("- Siguiente niño: formulario listo")

    return lineas


def detalle_pendientes_para_reporte(resultado):
    return construir_errores_resultado(resultado) + construir_observaciones_resultado(resultado)


def generar_reporte_excel(resultados, ruta_reporte_txt):
    ruta_reporte_xlsx = ruta_reporte_txt.with_suffix('.xlsx')
    libro = openpyxl.Workbook()
    hoja_resumen = libro.active
    hoja_resumen.title = 'Resumen'
    hoja_resumen.append(['Documento', 'Estado general', 'Errores', 'Observaciones', 'Guardado final', 'Mensaje guardado final'])

    hoja_detalle = libro.create_sheet('Detalle')
    hoja_detalle.append(['Documento', 'Estado general', 'Tipo', 'Codigo', 'Explicacion'])

    for resultado in resultados:
        documento = resultado.get('documento', '')
        estado = estado_general_resultado(resultado)
        errores = construir_errores_resultado(resultado)
        observaciones = construir_observaciones_resultado(resultado)
        codigos_error = codigos_error_resultado(resultado)
        codigos_observacion = codigos_observacion_resultado(resultado)

        hoja_resumen.append([
            documento,
            estado,
            ' | '.join(codigos_error),
            ' | '.join(codigos_observacion),
            'SI' if resultado.get('guardado_final', False) else 'NO',
            resultado.get('guardado_final_mensaje', ''),
        ])

        if not errores and not observaciones:
            hoja_detalle.append([documento, estado, 'OK', '', 'Sin novedades'])
            continue

        for codigo, explicacion in zip(codigos_error or [''], errores):
            hoja_detalle.append([documento, estado, 'ERROR', codigo, explicacion.lstrip('- ').strip()])
        for codigo, explicacion in zip(codigos_observacion or [''], observaciones):
            hoja_detalle.append([documento, estado, 'OBSERVACION', codigo, explicacion.lstrip('- ').strip()])

    for hoja in [hoja_resumen, hoja_detalle]:
        for columna in hoja.columns:
            ancho = max(len(str(celda.value or '')) for celda in columna)
            hoja.column_dimensions[columna[0].column_letter].width = min(max(ancho + 2, 14), 80)

    libro.save(ruta_reporte_xlsx)
    print(f"[+] Reporte Excel generado en: {ruta_reporte_xlsx}")
    return ruta_reporte_xlsx


def generar_reporte_inconsistencias(resultados):
    ruta_reporte = obtener_ruta_reporte_txt()

    lineas = []
    lineas.append(f"REPORTE DE INCONSISTENCIAS {CONFIG_EJECUCION['descripcion']}")
    lineas.append("=" * 80)
    lineas.append(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    lineas.append(f"Total registros procesados: {len(resultados)}")

    con_errores = []
    con_observaciones = []
    for resultado in resultados:
        if codigos_error_resultado(resultado):
            con_errores.append(resultado)
        elif construir_observaciones_resultado(resultado):
            con_observaciones.append(resultado)

    lineas.append(f"Registros con errores: {len(con_errores)}")
    lineas.append(f"Registros con observaciones: {len(con_observaciones)}")
    lineas.append(f"Registros sin novedades: {len(resultados) - len(con_errores) - len(con_observaciones)}")
    lineas.append("")

    lineas.append("DETALLE FINAL POR REGISTRO")
    lineas.append("-" * 80)
    for resultado in resultados:
        documento = resultado.get('documento', '')
        errores = construir_errores_resultado(resultado)
        observaciones = construir_observaciones_resultado(resultado)
        codigos_error = codigos_error_resultado(resultado)
        codigos_observacion = codigos_observacion_resultado(resultado)

        lineas.append(f"Documento: {documento}")
        lineas.append(f"Estado general: {estado_general_resultado(resultado)}")

        logros = detalle_logros_para_reporte(resultado)
        lineas.append("Se pudo:")
        if logros:
            lineas.extend(logros)
        else:
            lineas.append("- Sin acciones exitosas registradas")

        lineas.append("Errores:")
        if errores:
            lineas.extend(errores)
        else:
            lineas.append("- Sin errores")

        lineas.append("Observaciones:")
        if observaciones:
            lineas.extend(observaciones)
        else:
            lineas.append("- Sin observaciones")

        if codigos_error:
            lineas.append(f"Resumen de errores: {', '.join(codigos_error)}")
        if codigos_observacion:
            lineas.append(f"Resumen de observaciones: {', '.join(codigos_observacion)}")
        lineas.append("")

    ruta_reporte.write_text("\n".join(lineas), encoding="utf-8")
    print(f"\n[+] Reporte de inconsistencias generado en: {ruta_reporte}")
    generar_reporte_excel(resultados, ruta_reporte)
    return ruta_reporte


def obtener_documentos_con_novedades_desde_ultimo_reporte():
    reportes = sorted(BASE_DIR.glob(f"reporte_inconsistencias_{CONFIG_EJECUCION['prefijo_reporte']}_*.txt"), key=lambda ruta: ruta.stat().st_mtime, reverse=True)
    if not reportes:
        print("[*] No se encontró reporte previo; se procesarán todos los registros.")
        return None

    estados_recientes_por_documento = {}

    for reporte in reportes:
        try:
            contenido = reporte.read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            print(f"[!] No fue posible leer el reporte {reporte.name}: {e}")
            continue

        print(f"[*] Reporte evaluado para reproceso: {reporte.name}")
        bloques = re.split(r"\n(?=Documento:\s*)", contenido)
        for bloque in bloques:
            coincidencia_documento = re.search(r"Documento:\s*(\d+)", bloque)
            if not coincidencia_documento:
                continue
            documento = coincidencia_documento.group(1)
            if documento in estados_recientes_por_documento:
                continue

            if "Estado general: ERROR" in bloque or "Estado general: CON NOVEDADES" in bloque:
                if bloque_reporte_es_reproceso_ignorable(bloque):
                    estados_recientes_por_documento[documento] = 'ignorable'
                else:
                    estados_recientes_por_documento[documento] = 'error'
            elif "Estado general: CON OBSERVACIONES" in bloque:
                estados_recientes_por_documento[documento] = 'observacion'
            elif "Estado general: OK" in bloque:
                estados_recientes_por_documento[documento] = 'ok'

    documentos_con_novedades = sorted([
        documento for documento, estado in estados_recientes_por_documento.items()
        if estado == 'error'
    ])
    if documentos_con_novedades:
        print(f"[*] Documentos pendientes según estado más reciente: {', '.join(documentos_con_novedades)}")
        return documentos_con_novedades

    print("[*] No se encontraron pendientes reales en reportes previos; no se reprocesará ningún registro.")
    return []


def obtener_documentos_ya_procesados_desde_reportes():
    reportes = sorted(BASE_DIR.glob(f"reporte_inconsistencias_{CONFIG_EJECUCION['prefijo_reporte']}_*.txt"), key=lambda ruta: ruta.stat().st_mtime)
    if not reportes:
        return set()

    documentos = set()
    for reporte in reportes:
        try:
            contenido = reporte.read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            print(f"[!] No fue posible leer el reporte {reporte.name} para procesados: {e}")
            continue

        for coincidencia in re.findall(r"^Documento:\s*(\d+)", contenido, flags=re.MULTILINE):
            if coincidencia:
                documentos.add(coincidencia)

    return documentos


def bloque_reporte_es_reproceso_ignorable(bloque):
    texto = normalizar_texto(bloque)

    if normalizar_texto("No se completó por interrupción manual") in texto:
        return True
    if normalizar_texto("Ejecución interrumpida manualmente por el usuario") in texto:
        return True
    if "Estado general: CON OBSERVACIONES" in bloque:
        return True

    lineas_pendientes = []
    en_observaciones = False
    for linea in bloque.splitlines():
        linea_limpia = linea.strip()
        if linea_limpia == "Observaciones:":
            en_observaciones = True
            continue
        if en_observaciones and (linea_limpia.startswith("Resumen de observaciones:") or linea_limpia.startswith("Resumen de errores:")):
            break
        if en_observaciones and linea_limpia.startswith("-"):
            lineas_pendientes.append(normalizar_texto(linea_limpia))

    if not lineas_pendientes:
        en_no_se_pudo = False
        for linea in bloque.splitlines():
            linea_limpia = linea.strip()
            if linea_limpia == "No se pudo:":
                en_no_se_pudo = True
                continue
            if en_no_se_pudo and linea_limpia.startswith("Resumen de novedades:"):
                break
            if en_no_se_pudo and linea_limpia.startswith("-"):
                lineas_pendientes.append(normalizar_texto(linea_limpia))

    if not lineas_pendientes:
        return False

    prefijos_ignorables = [
        normalizar_texto("- Madre no creada por duplicidad de tipo de documento:"),
        normalizar_texto("- Padre no creado por duplicidad de tipo de documento:"),
        normalizar_texto("- Madre omitida por duplicidad de tipo de documento:"),
        normalizar_texto("- Padre omitida por duplicidad de tipo de documento:"),
        normalizar_texto("- Padre omitido por duplicidad de tipo de documento:"),
        normalizar_texto("- Parentesco madre: esperado='CONYUGE/COMPAÑERO(A)' | final=''"),
        normalizar_texto("- Parentesco padre: esperado='CONYUGE/COMPAÑERO(A)' | final=''"),
    ]

    return all(any(linea.startswith(prefijo) for prefijo in prefijos_ignorables) for linea in lineas_pendientes)


def obtener_documentos_forzados_desde_entorno():
    valor = (os.environ.get("DOCUMENTOS_REPROCESO") or "").strip()
    if not valor:
        return []
    documentos = [texto_excel(parte) for parte in re.split(r"[,\s]+", valor) if texto_excel(parte)]
    if documentos:
        print(f"[*] Reproceso forzado por entorno: {', '.join(documentos)}")
    return documentos


def abrir_tab_grupo_familiar():
    selectores_tab = [
        (By.ID, "__tab_cphCont_TabContainer1_tbngrupofamiliar"),
        (By.XPATH, "//span[normalize-space()='Grupo Familiar']"),
        (By.XPATH, "//a[normalize-space()='Grupo Familiar']"),
    ]
    return abrir_tab_con_espera(
        "Grupo Familiar",
        selectores_tab,
        esperar_grupo_familiar_listo,
        "grupo_familiar",
        timeout_postback=1.5,
        timeout_listo=2.0,
    )


def esperar_grupo_familiar_listo(timeout=3.0):
    selectores = [
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_GwvGrupoFamiliar"),
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona"),
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe"),
    ]
    fin = time.time() + timeout
    while time.time() < fin:
        esperar_postback_finalizado(0.8)
        for by, selector in selectores:
            try:
                elementos = driver.find_elements(by, selector)
                if any(elemento.is_displayed() for elemento in elementos):
                    return True
            except Exception:
                pass
        time.sleep(0.15)
    return False


def esperar_detalle_grupo_familiar(timeout=3.0):
    selectores = [
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe"),
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona"),
    ]
    fin = time.time() + timeout
    while time.time() < fin:
        esperar_postback_finalizado(0.8)
        for by, selector in selectores:
            try:
                elementos = driver.find_elements(by, selector)
                if any(elemento.is_displayed() for elemento in elementos):
                    return True
            except Exception:
                pass
        time.sleep(0.15)
    return False


def leer_filas_grupo_familiar():
    filas = []
    tabla = None
    selectores_tabla = [
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_GwvGrupoFamiliar"),
        (By.XPATH, "//table[contains(@id, 'GwvGrupoFamiliar')]"),
    ]

    for selector in selectores_tabla:
        try:
            tabla = wait.until(EC.presence_of_element_located((selector[0], selector[1])))
            break
        except Exception:
            pass

    if tabla is None:
        return filas

    for fila in tabla.find_elements(By.XPATH, ".//tr[td]"):
        celdas = [celda.text.strip() for celda in fila.find_elements(By.TAG_NAME, "td")]
        texto_fila = " ".join(celdas).strip()
        if not texto_fila:
            continue

        boton_detalle_id = ""
        boton_eliminar_id = ""
        try:
            boton = fila.find_element(By.XPATH, ".//input[contains(@id, 'btnInfo') and @type='image']")
            boton_detalle_id = boton.get_attribute("id") or ""
        except Exception:
            pass
        try:
            boton_el = fila.find_element(By.XPATH, ".//input[@type='image' and (contains(@id, 'btnEliminar') or contains(@id, 'BtnEliminar'))]")
            boton_eliminar_id = boton_el.get_attribute("id") or ""
        except Exception:
            pass

        documento = ""
        nombre = ""
        parentesco_jefe = ""
        parentesco_beneficiario = ""
        responsable = ""
        no_hace_parte = ""
        estado_familiar = ""
        indice_documento = -1

        for idx, celda in enumerate(celdas):
            if re.fullmatch(r"\d{6,}", celda or ""):
                indice_documento = idx
                documento = celda.strip()
                break

        if indice_documento >= 1:
            if indice_documento + 1 < len(celdas):
                nombre = celdas[indice_documento + 1].strip()
            if indice_documento + 2 < len(celdas):
                parentesco_jefe = celdas[indice_documento + 2].strip()
            if indice_documento + 3 < len(celdas):
                parentesco_beneficiario = celdas[indice_documento + 3].strip()
            if indice_documento + 4 < len(celdas):
                responsable = celdas[indice_documento + 4].strip()
            if indice_documento + 5 < len(celdas):
                no_hace_parte = celdas[indice_documento + 5].strip()
            if indice_documento + 6 < len(celdas):
                estado_familiar = celdas[indice_documento + 6].strip()

        filas.append({
            'texto': texto_fila,
            'celdas': celdas,
            'boton_detalle_id': boton_detalle_id,
            'boton_eliminar_id': boton_eliminar_id,
            'documento': documento,
            'nombre': nombre,
            'parentesco_jefe': parentesco_jefe,
            'parentesco_beneficiario': parentesco_beneficiario,
            'responsable': responsable,
            'no_hace_parte': no_hace_parte,
            'estado_familiar': estado_familiar,
        })

    return filas


def evaluar_completitud_grupo_familiar(filas, datos_excel):
    documento_beneficiario = normalizar_texto(datos_excel.get('documento', ''))
    documento_responsable = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', ''))) if responsable_adicional_es_valido(datos_excel) else ''
    parentesco_beneficiario_esperado = normalizar_texto(determinar_parentesco_beneficiario_desde_excel(datos_excel))
    beneficiario = None
    jefes_documento = []
    responsable_ok = False
    responsable_esperado_ok = not responsable_adicional_es_valido(datos_excel)
    filas_incompletas = []

    for fila in filas:
        documento_fila = normalizar_texto(fila.get('documento', ''))
        if normalizar_texto(fila.get('parentesco_jefe', '')) == normalizar_texto('JEFE DEL GRUPO FAMILIAR'):
            doc_jefe = fila.get('documento', '')
            if doc_jefe:
                jefes_documento.append(doc_jefe)
        if normalizar_texto(fila.get('responsable', '')) == 'S':
            responsable_ok = True
            if documento_responsable and documento_fila == documento_responsable:
                responsable_esperado_ok = True

        if documento_fila == documento_beneficiario:
            beneficiario = fila

        problemas = []
        if documento_fila == documento_beneficiario:
            parentesco_jefe_actual = normalizar_texto(fila.get('parentesco_jefe', ''))
            if not parentesco_jefe_actual:
                problemas.append('parentesco_jefe_vacio')
            elif parentesco_jefe_actual != parentesco_beneficiario_esperado:
                problemas.append(f"parentesco_jefe_esperado_{parentesco_beneficiario_esperado}")
        else:
            if fila.get('documento') and not normalizar_texto(fila.get('parentesco_jefe', '')):
                problemas.append('parentesco_jefe_vacio')
            if fila.get('documento') and not normalizar_texto(fila.get('parentesco_beneficiario', '')):
                problemas.append('parentesco_beneficiario_vacio')

        if problemas:
            filas_incompletas.append({
                'documento': fila.get('documento', ''),
                'nombre': fila.get('nombre', ''),
                'problemas': problemas,
            })

    return {
        'beneficiario_encontrado': beneficiario is not None,
        'jefe_ok': len(jefes_documento) == 1,
        'jefe_documento': jefes_documento[0] if jefes_documento else '',
        'jefe_cantidad': len(jefes_documento),
        'beneficiario_parentesco_jefe_esperado': determinar_parentesco_beneficiario_desde_excel(datos_excel),
        'beneficiario_parentesco_jefe_ok': bool(beneficiario and normalizar_texto(beneficiario.get('parentesco_jefe', '')) == parentesco_beneficiario_esperado),
        'responsable_ok': responsable_ok and responsable_esperado_ok,
        'filas_incompletas': filas_incompletas,
    }


def imprimir_filas_grupo_familiar(filas):
    print("\n  [GRUPO FAMILIAR]:")
    for idx, fila in enumerate(filas, 1):
        print(f"    {idx}. {fila['texto']}")


def buscar_fila_grupo_familiar_por_nombre(nombre_persona):
    filas = leer_filas_grupo_familiar()
    for fila in filas:
        if texto_contiene_nombre(fila['texto'], nombre_persona):
            return fila, filas
    return None, filas


def buscar_fila_grupo_familiar_por_documento(documento):
    filas = leer_filas_grupo_familiar()
    documento_normalizado = normalizar_texto(documento)
    for fila in filas:
        if documento_normalizado and documento_normalizado in normalizar_texto(fila['texto']):
            return fila, filas
    return None, filas


def buscar_filas_grupo_familiar_coincidentes(documento="", nombre_persona=""):
    filas = leer_filas_grupo_familiar()
    coincidencias = []
    documento_normalizado = normalizar_texto(documento)
    nombre_normalizado = normalizar_texto(nombre_persona)

    for fila in filas:
        texto_fila = normalizar_texto(fila.get('texto', ''))
        documento_fila = normalizar_texto(fila.get('documento', ''))
        nombre_fila = normalizar_texto(fila.get('nombre', ''))

        coincide_documento = bool(documento_normalizado and (documento_fila == documento_normalizado or documento_normalizado in texto_fila))
        coincide_nombre = bool(nombre_normalizado and (nombre_normalizado in texto_fila or nombre_normalizado == nombre_fila or nombre_fila in nombre_normalizado or nombre_normalizado in nombre_fila))

        if coincide_documento or coincide_nombre:
            coincidencias.append(fila)

    return coincidencias, filas


def buscar_fila_grupo_familiar_existente(persona):
    coincidencias, filas = buscar_filas_grupo_familiar_coincidentes(persona.get('documento', ''), persona.get('nombre_completo', ''))
    if coincidencias:
        return coincidencias[0], filas, coincidencias
    return None, filas, coincidencias


def buscar_fila_grupo_familiar_estable(documento="", nombre_persona="", reintentos=2, espera=1.5):
    """Relee la grilla varias veces porque RUB a veces tarda en reflejar altas/actualizaciones."""
    for intento in range(reintentos + 1):
        abrir_tab_grupo_familiar()
        esperar_grupo_familiar_listo(max(0.8, min(espera, 1.2)))

        fila = None
        filas = []
        if documento:
            fila, filas = buscar_fila_grupo_familiar_por_documento(documento)
        if not fila and nombre_persona:
            fila, filas = buscar_fila_grupo_familiar_por_nombre(nombre_persona)

        if fila:
            return fila, filas

        if intento < reintentos:
            print(f"  [~] Aún no aparece en la grilla ({documento or nombre_persona}); releyendo ({intento + 1}/{reintentos})...")
            esperar_postback_finalizado(0.8)
    return None, filas


def determinar_responsable_objetivo(datos_excel):
    # El responsable debe ser el que venga en Excel. Solo usamos respaldo cuando Excel no trae un responsable identificable.
    responsable_documento = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', '')))
    padre_documento = normalizar_texto(valor_excel_familia(datos_excel.get('padre_documento', '')))
    madre_documento = normalizar_texto(valor_excel_familia(datos_excel.get('madre_documento', '')))

    if responsable_documento:
        if responsable_documento == padre_documento and padre_documento:
            return 'padre'
        if responsable_documento == madre_documento and madre_documento:
            return 'madre'
        return 'responsable'

    responsable_parentesco = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_parentesco', '')))
    if responsable_parentesco == 'PADRE':
        return 'padre'
    if responsable_parentesco == 'MADRE':
        return 'madre'
    if responsable_parentesco:
        return 'responsable'

    responsable_nombre = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_nombre', '')))
    padre_nombre = normalizar_texto(valor_excel_familia(datos_excel.get('padre_nombre', '')))
    madre_nombre = normalizar_texto(valor_excel_familia(datos_excel.get('madre_nombre', '')))

    if responsable_nombre:
        if responsable_nombre and padre_nombre and (responsable_nombre in padre_nombre or padre_nombre in responsable_nombre):
            return 'padre'
        if responsable_nombre and madre_nombre and (responsable_nombre in madre_nombre or madre_nombre in responsable_nombre):
            return 'madre'
        return 'responsable'

    jefe_hogar = normalizar_texto(valor_excel_familia(datos_excel.get('jefe_hogar', '')))
    if jefe_hogar == 'PADRE':
        return 'padre'
    if jefe_hogar == 'MADRE':
        return 'madre'

    if madre_documento or madre_nombre:
        return 'madre'
    if padre_documento or padre_nombre:
        return 'padre'
    return ''


def determinar_jefe_objetivo(datos_excel):
    jefe_hogar = normalizar_texto(valor_excel_familia(datos_excel.get('jefe_hogar', '')))
    if jefe_hogar == 'PADRE':
        return 'padre'
    if jefe_hogar == 'MADRE':
        return 'madre'
    return determinar_responsable_objetivo(datos_excel)


def responsable_adicional_es_valido(datos_excel):
    if determinar_responsable_objetivo(datos_excel) != 'responsable':
        return False

    responsable_documento = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', '')))
    responsable_nombre = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_nombre', '')))
    return bool(responsable_documento or responsable_nombre)


def determinar_sexo_responsable_desde_excel(datos_excel):
    responsable_parentesco = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_parentesco', '')))
    masculinos = {'ABUELO', 'TIO', 'HERMANO', 'ESPOSO', 'PADRASTRO'}
    femeninos = {'ABUELA', 'TIA', 'HERMANA', 'ESPOSA', 'MADRASTRA'}

    if responsable_parentesco in masculinos:
        return 'MASCULINO'
    if responsable_parentesco in femeninos:
        return 'FEMENINO'
    return ''


def determinar_parentesco_beneficiario_responsable(datos_excel):
    responsable_parentesco = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_parentesco', '')))
    equivalencias = {
        'PADRE': 'PADRE',
        'MADRE': 'MADRE',
        'ABUELO': 'ABUELO (A)',
        'ABUELA': 'ABUELO (A)',
        'ABUELOS': 'ABUELO (A)',
        'TIO': 'TIO (A)',
        'TIA': 'TIO (A)',
        'TIO/A': 'TIO (A)',
        'HERMANO': 'HERMANO (A)',
        'HERMANA': 'HERMANO (A)',
        'ESPOSO': 'PADRASTRO',
        'ESPOSA': 'MADRASTRA',
        'PADRASTRO': 'PADRASTRO',
        'MADRASTRA': 'MADRASTRA',
    }
    parentesco = equivalencias.get(responsable_parentesco, valor_excel_familia(datos_excel.get('responsable_parentesco', '')))
    return normalizar_parentesco_beneficiario_para_select(parentesco)


def normalizar_parentesco_beneficiario_para_select(parentesco):
    """Normaliza parentesco para ddlIdParentescoBen usando solo opciones válidas del portal.
    Si el parentesco no existe en el combo, se convierte a OTRO para evitar 'Seleccione'."""
    parentesco_texto = texto_excel(parentesco)
    parentesco_norm = normalizar_texto(parentesco_texto)
    if not parentesco_norm:
        return ''

    equivalencias = {
        'PADRE': 'PADRE',
        'MADRE': 'MADRE',
        'HIJO (A)': 'HIJO (A)',
        'HIJASTRO(A)': 'HIJASTRO(A)',
        'HERMANO (A)': 'HERMANO (A)',
        'MADRASTRA': 'MADRASTRA',
        'PADRASTRO': 'PADRASTRO',
        'YERNO / NUERA': 'YERNO / NUERA',
        'SUEGRO (A)': 'SUEGRO (A)',
        'NIETO (A)': 'NIETO (A)',
        'CONYUGE/COMPANERO(A)': 'CONYUGE/COMPAÑERO(A)',
        # Parentescos frecuentes en Excel que no existen en este combo
        'ABUELO (A)': 'OTRO',
        'TIO (A)': 'OTRO',
        'SOBRINO (A)': 'OTRO',
        'PARIENTE U OTRO': 'OTRO',
        'PARIENTE': 'OTRO',
        'OTRO': 'OTRO',
    }

    return equivalencias.get(parentesco_norm, 'OTRO')


def responsable_es_pareja_de_madre(datos_excel):
    responsable_parentesco = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_parentesco', '')))
    return responsable_adicional_es_valido(datos_excel) and responsable_parentesco in {'ESPOSO', 'COMPANERO', 'COMPANERO(A)'}


def responsable_es_pareja_de_padre(datos_excel):
    responsable_parentesco = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_parentesco', '')))
    return responsable_adicional_es_valido(datos_excel) and responsable_parentesco in {'ESPOSA', 'COMPANERA', 'COMPANERA(A)'}


def calcular_parentescos_padres_objetivo(datos_excel):
    jefe_objetivo = determinar_jefe_objetivo(datos_excel)
    parentesco_padre_objetivo = 'PADRE'
    parentesco_madre_objetivo = 'MADRE'

    if jefe_objetivo == 'padre':
        parentesco_padre_objetivo = 'JEFE DEL GRUPO FAMILIAR'
        parentesco_madre_objetivo = 'CONYUGE/COMPAÑERO(A)'
    elif jefe_objetivo == 'madre':
        parentesco_padre_objetivo = 'CONYUGE/COMPAÑERO(A)'
        parentesco_madre_objetivo = 'JEFE DEL GRUPO FAMILIAR'

    if responsable_es_pareja_de_madre(datos_excel) and parentesco_madre_objetivo == 'CONYUGE/COMPAÑERO(A)':
        parentesco_madre_objetivo = 'PARIENTE U OTRO'
    if responsable_es_pareja_de_padre(datos_excel) and parentesco_padre_objetivo == 'CONYUGE/COMPAÑERO(A)':
        parentesco_padre_objetivo = 'PARIENTE U OTRO'

    return parentesco_padre_objetivo, parentesco_madre_objetivo, jefe_objetivo


def determinar_parentesco_beneficiario_desde_excel(datos_excel):
    jefe_hogar = normalizar_texto(valor_excel_familia(datos_excel.get('jefe_hogar', '')))
    responsable_parentesco = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_parentesco', '')))

    # Cuando el jefe o acudiente es padre/madre, el beneficiario debe quedar como HIJO(A).
    if jefe_hogar in {'PADRE', 'MADRE'} or responsable_parentesco in {'PADRE', 'MADRE'}:
        return 'HIJO (A)'

    equivalencias = {
        'ABUELO': 'NIETO (A)',
        'ABUELA': 'NIETO (A)',
        'ABUELOS': 'NIETO (A)',
        'TIO': 'SOBRINO (A)',
        'TIA': 'SOBRINO (A)',
        'TIO/A': 'SOBRINO (A)',
        'HERMANO': 'HERMANO (A)',
        'HERMANA': 'HERMANO (A)',
        'PADRASTRO': 'HIJASTRO(A)',
        'MADRASTRA': 'HIJASTRO(A)',
    }

    for fuente in [jefe_hogar, responsable_parentesco]:
        for clave, parentesco in equivalencias.items():
            if clave in fuente:
                return parentesco

    return 'HIJO (A)'


def construir_persona_desde_excel(datos_excel, tipo_persona):
    if tipo_persona == 'responsable':
        responsable_objetivo = determinar_responsable_objetivo(datos_excel)
        nacimiento = resolver_nacimiento_familia(
            datos_excel.get('responsable_departamento_nacimiento', ''),
            datos_excel.get('responsable_municipio_nacimiento', ''),
        )
        return {
            'tipo': 'responsable',
            'nombre_completo': valor_excel_familia(datos_excel.get('responsable_nombre', '')),
            'doc_tipo': mapear_tipo_documento_grupo_familiar(datos_excel.get('responsable_doc_tipo', ''), datos_excel.get('nacionalidad_beneficiario', '')),
            'documento': valor_excel_familia(datos_excel.get('responsable_documento', '')),
            'primer_nombre': valor_excel_familia(datos_excel.get('responsable_primer_nombre', '')),
            'segundo_nombre': valor_excel_familia(datos_excel.get('responsable_segundo_nombre', '')),
            'primer_apellido': valor_excel_familia(datos_excel.get('responsable_primer_apellido', '')),
            'segundo_apellido': valor_excel_familia(datos_excel.get('responsable_segundo_apellido', '')),
            'fecha_nacimiento': fecha_excel_familia(datos_excel.get('responsable_fecha_nacimiento', '')),
            'pais_nacimiento': nacimiento['pais_nacimiento'],
            'departamento_nacimiento': nacimiento['departamento_nacimiento'],
            'municipio_nacimiento': nacimiento['municipio_nacimiento'],
            'parentesco_beneficiario': determinar_parentesco_beneficiario_responsable(datos_excel),
            'sexo': determinar_sexo_responsable_desde_excel(datos_excel),
            'es_responsable': responsable_objetivo == 'responsable',
        }

    prefijo = 'padre' if tipo_persona == 'padre' else 'madre'
    sexo = 'MASCULINO' if prefijo == 'padre' else 'FEMENINO'
    responsable_objetivo = determinar_responsable_objetivo(datos_excel)
    es_responsable = responsable_objetivo == prefijo
    nacimiento = resolver_nacimiento_familia(
        datos_excel.get(f'{prefijo}_departamento_nacimiento', ''),
        datos_excel.get(f'{prefijo}_municipio_nacimiento', ''),
    )

    return {
        'tipo': prefijo,
        'nombre_completo': valor_excel_familia(datos_excel.get(f'{prefijo}_nombre', '')),
        'doc_tipo': mapear_tipo_documento_grupo_familiar(datos_excel.get(f'{prefijo}_doc_tipo', ''), datos_excel.get('nacionalidad_beneficiario', '')),
        'documento': valor_excel_familia(datos_excel.get(f'{prefijo}_documento', '')),
        'primer_nombre': valor_excel_familia(datos_excel.get(f'{prefijo}_primer_nombre', '')),
        'segundo_nombre': valor_excel_familia(datos_excel.get(f'{prefijo}_segundo_nombre', '')),
        'primer_apellido': valor_excel_familia(datos_excel.get(f'{prefijo}_primer_apellido', '')),
        'segundo_apellido': valor_excel_familia(datos_excel.get(f'{prefijo}_segundo_apellido', '')),
        'fecha_nacimiento': fecha_excel_familia(datos_excel.get(f'{prefijo}_fecha_nacimiento', '')),
        'pais_nacimiento': nacimiento['pais_nacimiento'],
        'departamento_nacimiento': nacimiento['departamento_nacimiento'],
        'municipio_nacimiento': nacimiento['municipio_nacimiento'],
        'parentesco_beneficiario': 'PADRE' if prefijo == 'padre' else 'MADRE',
        'sexo': sexo,
        'es_responsable': es_responsable,
    }


def completar_campos_basicos_persona_grupo_familiar(persona, consultar_por_id=False):
    doc_tipo = persona.get('doc_tipo', '')
    if doc_tipo:
        posibles_selects = [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_ddlIdTipoDocumento",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdTipoDocumento",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlTipoDocumento",
        ]
        for select_id in posibles_selects:
            try:
                if asignar_select_por_id(select_id, doc_tipo, "Tipo documento grupo familiar"):
                    break
            except Exception:
                pass
        else:
            asignar_select_por_etiqueta("Tipo de Documento de Identidad", doc_tipo, "Tipo documento grupo familiar")

    documento_ids = [
        "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_txtNumeroDocumento",
        "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_txtNumeroDocumento",
        "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_txtIdentificacion",
    ]

    documento_asignado = False
    for element_id in documento_ids:
        if asignar_valor_input_por_id(element_id, persona.get('documento', ''), "Grupo familiar documento"):
            documento_asignado = True
            break
    if not documento_asignado and persona.get('documento', ''):
        documento_asignado = asignar_valor_input_por_etiqueta("Número de Documento de Identidad", persona['documento'], "Grupo familiar documento")

    # Si el documento quedó escrito y la lupa se habilita, pulsarla para que el portal
    # cargue/valide datos del integrante antes de completar el resto del detalle.
    if documento_asignado and persona.get('documento', ''):
        pulsar_lupa_buscar_grupo_familiar("Grupo familiar documento")

    posibles_inputs = {
        'primer_nombre': [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_txtPrimerNombre",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_txtPrimerNombre",
        ],
        'segundo_nombre': [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_txtSegundoNombre",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_txtSegundoNombre",
        ],
        'primer_apellido': [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_txtPrimerApellido",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_txtPrimerApellido",
        ],
        'segundo_apellido': [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_txtSegundoApellido",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_txtSegundoApellido",
        ],
        'fecha_nacimiento': [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_cuwFechaNacimiento_txtFecha",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_cuwFechaNacimiento_txtFecha",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_fecha_txtFecha",
        ],
    }

    for llave, ids in posibles_inputs.items():
        valor = persona.get(llave, '')
        if not valor:
            continue
        asignado = False
        for element_id in ids:
            if asignar_valor_input_por_id(element_id, valor, f"Grupo familiar {llave}"):
                asignado = True
                break
        if not asignado:
            etiquetas = {
                'primer_nombre': 'Primer Nombre',
                'segundo_nombre': 'Segundo Nombre',
                'primer_apellido': 'Primer Apellido',
                'segundo_apellido': 'Segundo Apellido',
                'fecha_nacimiento': 'Fecha de Nacimiento',
            }
            asignar_valor_input_por_etiqueta(etiquetas.get(llave, llave), valor, f"Grupo familiar {llave}")


def completar_detalle_existente_grupo_familiar(persona, parentesco_objetivo):
    parentesco_beneficiario = normalizar_parentesco_beneficiario_para_select(persona.get('parentesco_beneficiario', ''))

    if parentesco_objetivo:
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
            parentesco_objetivo,
            "Parentesco grupo familiar",
        )
        esperar_postback_finalizado(1.2)
        esperar_detalle_grupo_familiar(1.5)

    if parentesco_beneficiario:
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen",
            parentesco_beneficiario,
            "Parentesco con el beneficiario",
        )
        esperar_postback_finalizado(1.0)

    completar_campos_basicos_persona_grupo_familiar(persona, consultar_por_id=False)

    if persona.get('sexo', ''):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_ddlIdSexo",
            persona['sexo'],
            "Grupo familiar sexo",
        )
        esperar_postback_finalizado(0.8)

    completar_ubicacion_nacimiento_grupo_familiar(
        persona,
        prefijo_log="Grupo familiar",
    )

    parentesco_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe")
    if parentesco_objetivo and not texto_select_coincide(parentesco_final, parentesco_objetivo):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
            parentesco_objetivo,
            "Parentesco grupo familiar (ajuste final)",
        )
        esperar_postback_finalizado(0.8)

    parentesco_ben_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen")
    if parentesco_beneficiario and not texto_select_coincide(parentesco_ben_final, parentesco_beneficiario):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen",
            parentesco_beneficiario,
            "Parentesco con el beneficiario (ajuste final)",
        )
        esperar_postback_finalizado(0.8)

    parentesco_ben_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen")
    if parentesco_beneficiario and texto_select_coincide(parentesco_ben_final, 'Seleccione'):
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen",
            'OTRO',
            "Parentesco con el beneficiario (fallback OTRO)",
        )
        esperar_postback_finalizado(0.8)

    try:
        chk_responsable = driver.find_element(By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_chk_Responsable")
        if chk_responsable.is_selected() != persona.get('es_responsable', False):
            driver.execute_script("arguments[0].click();", chk_responsable)
            esperar_postback_finalizado(0.8)
        print(f"  [+] Responsable: {'SI' if persona.get('es_responsable', False) else 'NO'}")
    except Exception as e:
        print(f"  [!] No fue posible ajustar responsable: {e}")

    verificar_y_corregir_persona_grupo_familiar(persona, parentesco_objetivo)


def abrir_formulario_grupo_familiar_existente(persona):
    fila, filas, coincidencias = buscar_fila_grupo_familiar_existente(persona)
    imprimir_filas_grupo_familiar(filas)

    if len(coincidencias) > 1:
        print(f"  [~] Se encontraron {len(coincidencias)} coincidencias para {persona['nombre_completo']}; se editará la primera para evitar nuevas duplicaciones")

    if not fila:
        return False

    return abrir_detalle_persona_grupo_familiar(
        nombre_persona=persona.get('nombre_completo', ''),
        documento=persona.get('documento', ''),
        requiere_edicion=True,
    )


def limpiar_formulario_grupo_familiar():
    selectores = [
        (By.XPATH, "//a[contains(normalize-space(), 'Limpiar Pantalla')]"),
        (By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblLimpiarPantalla"),
    ]
    for selector in selectores:
        try:
            boton = driver.find_element(selector[0], selector[1])
            driver.execute_script("arguments[0].click();", boton)
            esperar_postback_finalizado(1.5)
            esperar_detalle_grupo_familiar(1.5)
            return True
        except Exception:
            pass
    return False


ultimo_estado_guardado_grupo_familiar = {'codigo': '', 'mensaje': ''}


def leer_mensaje_duplicidad_tipo_documento_grupo_familiar():
    selectores = [
        (By.XPATH, "//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'diferentes tipos de documento') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
        (By.XPATH, "//*[contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚ', 'abcdefghijklmnopqrstuvwxyzáéíóú'), 'la persona se encuentra registrada') and not(ancestor-or-self::*[contains(@style,'display:none')]) ]"),
    ]
    for by, selector in selectores:
        try:
            for elemento in driver.find_elements(by, selector):
                if not elemento.is_displayed():
                    continue
                texto = " ".join((elemento.text or '').split()).strip()
                if texto and 'diferentes tipos de documento' in normalizar_texto(texto).lower():
                    return texto
        except Exception:
            pass
    return ""


def abrir_modo_agregar_persona_grupo_familiar():
    try:
        boton = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton)
        try:
            boton.click()
        except Exception:
            driver.execute_script("arguments[0].click();", boton)
        esperar_postback_finalizado(max(0.8, ESPERA_MODO_AGREGAR_PERSONA))
        esperar_detalle_grupo_familiar(1.5)
        print("  [+] Modo agregar persona abierto en Grupo Familiar")
        return True
    except Exception as e:
        print(f"  [!] No fue posible abrir el modo agregar persona: {e}")
        return False


def guardar_persona_grupo_familiar():
    global ultimo_estado_guardado_grupo_familiar
    ultimo_estado_guardado_grupo_familiar = {'codigo': '', 'mensaje': ''}
    try:
        boton = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton)
        try:
            boton.click()
        except Exception:
            driver.execute_script("arguments[0].click();", boton)
        esperar_postback_finalizado(max(1.0, ESPERA_GUARDAR_PERSONA))
        esperar_grupo_familiar_listo(1.8)
        mensaje_duplicidad = leer_mensaje_duplicidad_tipo_documento_grupo_familiar()
        if mensaje_duplicidad:
            ultimo_estado_guardado_grupo_familiar = {
                'codigo': 'duplicidad_tipo_documento',
                'mensaje': mensaje_duplicidad,
            }
            print(f"  [!] El integrante no se pudo crear por duplicidad de tipo de documento: {mensaje_duplicidad}")
            return False
        print("  [+] Botón 'Agregar Persona' pulsado para guardar el integrante")
        ultimo_estado_guardado_grupo_familiar = {'codigo': 'ok', 'mensaje': 'guardado'}
        return True
    except Exception as e:
        ultimo_estado_guardado_grupo_familiar = {'codigo': 'error', 'mensaje': str(e)}
        print(f"  [!] No fue posible guardar el integrante con 'Agregar Persona': {e}")
        return False


def verificar_y_corregir_persona_grupo_familiar(persona, parentesco_objetivo):
    """Verifica que todos los campos obligatorios de la persona en grupo familiar estén completos.
    Si falta algo (municipio, parentesco jefe), lo corrige automáticamente."""
    
    campos_a_verificar = []
    
    # Verificar municipio nacimiento
    if persona.get('municipio_nacimiento', ''):
        municipio_actual = obtener_texto_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdMunicipioNacimiento"
        )
        if not municipio_actual or texto_select_coincide(municipio_actual, 'Seleccione'):
            campos_a_verificar.append(('municipio_nacimiento', persona['municipio_nacimiento']))
    
    # Verificar parentesco con jefe
    if parentesco_objetivo:
        parentesco_actual = obtener_texto_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe"
        )
        if not parentesco_actual or texto_select_coincide(parentesco_actual, 'Seleccione'):
            campos_a_verificar.append(('parentesco_jefe', parentesco_objetivo))
    
    # Si hay campos vacíos, corregirlos
    if campos_a_verificar:
        print(f"  [*] Identificados {len(campos_a_verificar)} campo(s) incompleto(s) en grupo familiar; corrigiendo...")
        
        for campo, valor in campos_a_verificar:
            if campo == 'municipio_nacimiento':
                print(f"    [+] Completando municipio nacimiento: {valor}")
                if persona.get('departamento_nacimiento', ''):
                    # Asegurar que el departamento esté correcto
                    asignar_select_por_id(
                        "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdDepartamento",
                        persona['departamento_nacimiento'],
                        "departamento_nacimiento (corrección)",
                    )
                    esperar_postback_finalizado(1.2)
                    esperar_detalle_grupo_familiar(1.0)
                
                # Esperar a que la opción esté disponible
                esperar_opcion_en_select(
                    "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdMunicipioNacimiento",
                    valor,
                    timeout=8,
                )
                # Asignar municipio
                asignar_select_por_id(
                    "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdMunicipioNacimiento",
                    valor,
                    "municipio_nacimiento (corrección)",
                )
                esperar_postback_finalizado(0.8)
            
            elif campo == 'parentesco_jefe':
                print(f"    [+] Completando parentesco con jefe: {valor}")
                asignar_select_por_id(
                    "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
                    valor,
                    "parentesco_jefe (corrección)",
                )
                esperar_postback_finalizado(1.0)
        
        # Reabicar el detalle después de correcciones
        esperar_detalle_grupo_familiar(1.2)
        return True
    
    return False


def llenar_formulario_persona_grupo_familiar(persona, parentesco_objetivo, consultar_por_id=False):
    limpiar_formulario_grupo_familiar()
    parentesco_beneficiario = normalizar_parentesco_beneficiario_para_select(persona.get('parentesco_beneficiario', ''))

    asignar_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe", parentesco_objetivo, "Parentesco grupo familiar") or asignar_select_por_etiqueta("Parentesco con respecto al Jefe del Grupo Familiar", parentesco_objetivo, "Parentesco grupo familiar")
    esperar_postback_finalizado(1.2)
    esperar_detalle_grupo_familiar(1.5)

    if parentesco_beneficiario:
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen",
            parentesco_beneficiario,
            "Parentesco con el beneficiario",
        )
        esperar_postback_finalizado(1.0)

    completar_campos_basicos_persona_grupo_familiar(persona, consultar_por_id=consultar_por_id)

    posibles_selects_detalle = {
        'sexo': ["cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_ddlIdSexo"],
        'pais_nacimiento': ["cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdPaisNacimiento"],
        'departamento_nacimiento': [
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdDepartamento",
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdDepartamentoNacimiento",
        ],
        'municipio_nacimiento': ["cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdMunicipioNacimiento"],
    }

    def asignar_select_detalle(llave, valor):
        ids = posibles_selects_detalle[llave]
        asignado = False
        for element_id in ids:
            if asignar_select_por_id(element_id, valor, f"Grupo familiar {llave}"):
                asignado = True
                break
        if not asignado:
            etiquetas = {
                'sexo': 'Sexo',
                'pais_nacimiento': 'Pais nacimiento',
                'departamento_nacimiento': 'Departamento nacimiento',
                'municipio_nacimiento': 'Municipio nacimiento',
            }
            asignado = asignar_select_por_etiqueta(etiquetas.get(llave, llave), valor, f"Grupo familiar {llave}")
        return asignado

    if persona.get('sexo', ''):
        asignar_select_detalle('sexo', persona['sexo'])
        esperar_postback_finalizado(0.6)

    completar_ubicacion_nacimiento_grupo_familiar(
        persona,
        prefijo_log="Grupo familiar",
    )

    # Reaplicar parentesco al final: algunos postbacks del portal lo devuelven a "Seleccione".
    asignar_select_por_id(
        "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
        parentesco_objetivo,
        "Parentesco grupo familiar (final)",
    ) or asignar_select_por_etiqueta(
        "Parentesco con respecto al Jefe del Grupo Familiar",
        parentesco_objetivo,
        "Parentesco grupo familiar (final)",
    )
    esperar_postback_finalizado(0.8)

    parentesco_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe")
    if parentesco_objetivo and not texto_select_coincide(parentesco_final, parentesco_objetivo):
        print(f"  [!] Parentesco final quedó en '{parentesco_final or 'Seleccione'}'; reintentando una vez más...")
        asignar_select_por_id(
            "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
            parentesco_objetivo,
            "Parentesco grupo familiar (reintento final)",
        )

    if parentesco_beneficiario:
        parentesco_ben_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen")
        if not texto_select_coincide(parentesco_ben_final, parentesco_beneficiario):
            print(f"  [!] Parentesco con el beneficiario final quedó en '{parentesco_ben_final or 'Seleccione'}'; reintentando una vez más...")
            asignar_select_por_id(
                "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen",
                parentesco_beneficiario,
                "Parentesco con el beneficiario (reintento final)",
            )
            esperar_postback_finalizado(0.8)

    if parentesco_beneficiario:
        parentesco_ben_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen")
        if texto_select_coincide(parentesco_ben_final, 'Seleccione'):
            print("  [!] Parentesco con el beneficiario sigue en 'Seleccione'; aplicando fallback a 'OTRO'...")
            asignar_select_por_id(
                "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoBen",
                'OTRO',
                "Parentesco con el beneficiario (fallback OTRO)",
            )
            esperar_postback_finalizado(0.8)
    
    # Verificación final en la misma apertura: corregir obligatorios si quedaron vacíos.
    verificar_y_corregir_persona_grupo_familiar(persona, parentesco_objetivo)

    if persona.get('sexo', ''):
        sexo_final = obtener_texto_select_por_id("cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_DatosBasicos1_ddlIdSexo")
        if not texto_select_coincide(sexo_final, persona['sexo']):
            print(f"  [!] Sexo final quedó en '{sexo_final or 'Seleccione'}'; reintentando una vez más...")
            asignar_select_detalle('sexo', persona['sexo'])
            esperar_postback_finalizado(0.6)

    try:
        chk_responsable = driver.find_element(By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_chk_Responsable")
        if chk_responsable.is_selected() != persona.get('es_responsable', False):
            driver.execute_script("arguments[0].click();", chk_responsable)
            esperar_postback_finalizado(0.8)
        print(f"  [+] Responsable: {'SI' if persona.get('es_responsable', False) else 'NO'}")
    except Exception as e:
        print(f"  [!] No fue posible ajustar responsable: {e}")


def agregar_o_actualizar_persona_grupo_familiar(persona, parentesco_objetivo):
    for intento in range(2):
        fila_existente, _, coincidencias = buscar_fila_grupo_familiar_existente(persona)
        existe = abrir_formulario_grupo_familiar_existente(persona) if fila_existente else False
        if existe:
            print(f"  [*] Editando persona existente en grupo familiar: {persona['nombre_completo']}")
            actualizado = actualizar_parentesco_grupo_familiar(
                persona.get('nombre_completo', ''),
                parentesco_objetivo,
                documento=persona.get('documento', ''),
                persona=persona,
            )
            return {
                'ok': bool(actualizado),
                'omitido_duplicidad': False,
                'mensaje': '' if actualizado else 'no_se_pudo_actualizar_existente',
            }
        else:
            if coincidencias:
                print(f"  [~] Se detectaron coincidencias previas para {persona['nombre_completo']}; no se creará una nueva fila")
                return {'ok': True, 'omitido_duplicidad': False, 'mensaje': 'ya_existia_en_grilla'}
            print(f"  [*] Creando persona faltante en grupo familiar: {persona['nombre_completo']}")
            if not abrir_modo_agregar_persona_grupo_familiar():
                return False

        llenar_formulario_persona_grupo_familiar(persona, parentesco_objetivo, consultar_por_id=not existe)

        try:
            if not guardar_persona_grupo_familiar():
                if ultimo_estado_guardado_grupo_familiar.get('codigo') == 'duplicidad_tipo_documento':
                    return {
                        'ok': False,
                        'omitido_duplicidad': True,
                        'mensaje': ultimo_estado_guardado_grupo_familiar.get('mensaje', ''),
                    }
                return {'ok': False, 'omitido_duplicidad': False, 'mensaje': ultimo_estado_guardado_grupo_familiar.get('mensaje', '')}
            print(f"  [+] Persona guardada en grupo familiar: {persona['nombre_completo']}")
            fila, _ = buscar_fila_grupo_familiar_estable(persona.get('documento', ''), persona.get('nombre_completo', ''))
            if fila:
                return {'ok': True, 'omitido_duplicidad': False, 'mensaje': ''}

            if intento == 0:
                print(f"  [~] La persona aún no aparece en la grilla; reintentando guardado de {persona['nombre_completo']}...")
                try:
                    abrir_tab_grupo_familiar()
                    limpiar_formulario_grupo_familiar()
                except Exception:
                    pass
                continue

            print(f"  [!] La persona no apareció en la grilla tras guardar: {persona['nombre_completo']}")
            return {'ok': False, 'omitido_duplicidad': False, 'mensaje': 'no_aparece_en_grilla'}
        except Exception as e:
            print(f"  [!] No fue posible guardar persona de grupo familiar: {e}")
            return {'ok': False, 'omitido_duplicidad': False, 'mensaje': str(e)}

    return {'ok': False, 'omitido_duplicidad': False, 'mensaje': 'reintentos_agotados'}


def abrir_detalle_persona_grupo_familiar(nombre_persona="", documento="", requiere_edicion=False, reintentos=2):
    referencia = documento or nombre_persona

    for intento in range(reintentos + 1):
        fila = None
        filas = []

        if documento:
            fila, filas = buscar_fila_grupo_familiar_por_documento(documento)
        if not fila and nombre_persona:
            fila, filas = buscar_fila_grupo_familiar_por_nombre(nombre_persona)
        imprimir_filas_grupo_familiar(filas)

        if not fila:
            print(f"  [!] No se encontró a {referencia} en Grupo Familiar")
            return False

        if not fila['boton_detalle_id']:
            print(f"  [!] La fila de {referencia} no tiene botón detalle")
            return False

        try:
            boton = wait.until(EC.element_to_be_clickable((By.ID, fila['boton_detalle_id'])))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton)
            try:
                boton.click()
            except Exception:
                driver.execute_script("arguments[0].click();", boton)
            esperar_postback_finalizado(1.0)
            esperar_detalle_grupo_familiar(2.0)
            print(f"  [+] Detalle abierto para {referencia}")

            if not requiere_edicion:
                return True

            ids_validacion = [
                "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdPaisNacimiento",
                "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe",
            ]
            edicion_ok = False
            for element_id in ids_validacion:
                elementos = driver.find_elements(By.ID, element_id)
                if elementos and control_esta_habilitado(elementos[0]):
                    print(f"  [+] {referencia}: detalle en modo edición confirmado por botón azul")
                    edicion_ok = True
                    return True

            if not edicion_ok:
                if intento < reintentos:
                    print(f"  [~] {referencia}: detalle aún en solo lectura tras botón azul; reintentando apertura ({intento + 1}/{reintentos})...")
                    abrir_tab_grupo_familiar()
                    esperar_grupo_familiar_listo(1.2)
                    continue

            print(f"  [!] {referencia}: no quedó en modo edición después del botón azul")
            return False
        except Exception as e:
            if intento < reintentos:
                print(f"  [~] No fue posible abrir detalle de {referencia} (reintento {intento + 1}/{reintentos}): {e}")
                time.sleep(1.0)
                continue
            print(f"  [!] No fue posible abrir el detalle de {referencia}: {e}")
            return False

    return False


def leer_parentesco_actual_grupo_familiar(nombre_persona="", documento=""):
    if not abrir_detalle_persona_grupo_familiar(nombre_persona, documento):
        return ""

    try:
        select_par = Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe"))
        return select_par.first_selected_option.text.strip()
    except Exception:
        return ""


def actualizar_parentesco_grupo_familiar(nombre_persona, parentesco_objetivo, documento="", persona=None):
    # Primero leer la tabla sin abrir ningún formulario
    fila, filas = buscar_fila_grupo_familiar_por_documento(documento) if documento else (None, [])
    if not fila and nombre_persona:
        fila, filas = buscar_fila_grupo_familiar_por_nombre(nombre_persona)
    imprimir_filas_grupo_familiar(filas)

    if not fila:
        referencia = documento or nombre_persona
        print(f"  [!] No se encontró a {referencia} en Grupo Familiar")
        return False

    # Comparar desde la tabla (sin abrir el detalle)
    referencia = documento or nombre_persona
    parentesco_jefe_tabla = fila.get('parentesco_jefe', '').strip()
    parentesco_ben_tabla = fila.get('parentesco_beneficiario', '').strip()
    responsable_tabla = fila.get('responsable', '').strip() == 'S'

    objetivo_parentesco_jefe = normalizar_texto(parentesco_objetivo) if parentesco_objetivo else ''
    objetivo_parentesco_ben = normalizar_texto(persona.get('parentesco_beneficiario', '')) if persona else ''
    objetivo_responsable = persona.get('es_responsable', False) if persona else False

    jefe_ok = (not objetivo_parentesco_jefe) or (normalizar_texto(parentesco_jefe_tabla) == objetivo_parentesco_jefe)
    ben_ok = (not objetivo_parentesco_ben) or (normalizar_texto(parentesco_ben_tabla) == objetivo_parentesco_ben)
    responsable_ok = responsable_tabla == objetivo_responsable

    # Detectar conflicto de múltiple JEFE antes de abrir detalle
    if not jefe_ok and normalizar_texto(parentesco_objetivo) == normalizar_texto('JEFE DEL GRUPO FAMILIAR'):
        for fila_check in leer_filas_grupo_familiar():
            doc_check = normalizar_texto(fila_check.get('documento', ''))
            if doc_check and doc_check != normalizar_texto(documento) and normalizar_texto(fila_check.get('parentesco_jefe', '')) == normalizar_texto('JEFE DEL GRUPO FAMILIAR'):
                print(f"  [~] {referencia}: conflicto — {fila_check.get('documento')} ya es JEFE DEL GRUPO FAMILIAR. Se intentará corregir el jefe antes de continuar.")
                return False

    if jefe_ok and ben_ok and responsable_ok:
        # Parentesco correcto en tabla, pero verificar pais_nacimiento dentro del detalle
        if persona and persona.get('pais_nacimiento', ''):
            if abrir_detalle_persona_grupo_familiar(nombre_persona, documento, requiere_edicion=True):
                try:
                    select_pais = driver.find_element(By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdPaisNacimiento")
                    pais_valor = Select(select_pais).first_selected_option.get_attribute("value")
                    if pais_valor and pais_valor != "-1":
                        print(f"  [=] {referencia}: todo correcto incluyendo pais_nacimiento (sin modificar)")
                        return True
                    # pais_nacimiento vacío → completar sin tocar parentesco
                    print(f"  [*] {referencia}: parentesco OK en tabla pero pais_nacimiento vacío, completando...")
                    completar_detalle_existente_grupo_familiar(persona, None)
                    actualizar_btn = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", actualizar_btn)
                    try:
                        actualizar_btn.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", actualizar_btn)
                    print(f"  [+] pais_nacimiento completado para: {referencia}")
                    esperar_postback_finalizado(max(0.8, ESPERA_ACTUALIZAR_PERSONA))
                    esperar_grupo_familiar_listo(1.2)
                    return True
                except Exception as e:
                    print(f"  [!] Error verificando pais_nacimiento de {referencia}: {e}")
                    return True
        print(f"  [=] {referencia}: todo correcto en tabla (sin modificar)")
        return True

    # Algo difiere → abrir detalle y corregir solo lo necesario
    print(f"  [*] {referencia}: requiere cambios (jefe_ok={jefe_ok}, ben_ok={ben_ok}, resp_ok={responsable_ok})")
    if not abrir_detalle_persona_grupo_familiar(nombre_persona, documento, requiere_edicion=True):
        return False

    try:
        select_elem = wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe")))
        Select(select_elem)  # solo confirma que cargó

        completar_detalle_existente_grupo_familiar(persona or {}, parentesco_objetivo)
        actualizar = wait.until(EC.element_to_be_clickable((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_LblAgregarPersona")))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", actualizar)
        try:
            actualizar.click()
        except Exception:
            driver.execute_script("arguments[0].click();", actualizar)
        print(f"  [+] Persona actualizada: {referencia} -> {parentesco_objetivo}")
        esperar_postback_finalizado(max(0.8, ESPERA_ACTUALIZAR_PERSONA))
        esperar_grupo_familiar_listo(1.2)
        return True
    except Exception as e:
        print(f"  [!] Error actualizando parentesco de {referencia}: {e}")
        return False


def eliminar_persona_grupo_familiar(boton_eliminar_id):
    """Pulsa el botón eliminar de una fila de la grilla y confirma el alert si aparece."""
    try:
        btn = wait.until(EC.element_to_be_clickable((By.ID, boton_eliminar_id)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
        esperar_postback_finalizado(0.8)
        # Confirmar cualquier diálogo de confirmación
        try:
            alerta = driver.switch_to.alert
            alerta.accept()
            esperar_postback_finalizado(0.8)
        except Exception:
            pass
        esperar_grupo_familiar_listo(1.5)
        return True
    except Exception as e:
        print(f"  [!] No se pudo eliminar (id={boton_eliminar_id}): {e}")
        return False


def limpiar_duplicados_integrantes_validos(datos_excel):
    """Elimina filas repetidas de beneficiario/padre/madre dentro de la misma grilla.
    Conserva una sola fila por persona y deja la corrección fina de parentescos para pasos posteriores."""
    print("  [*] Revisando duplicados de beneficiario/padre/madre en Grupo Familiar...")

    personas_objetivo = [
        {
            'descripcion': 'beneficiario',
            'documento': datos_excel.get('documento', ''),
            'nombre': nombre_completo(
                datos_excel.get('primer_nombre', ''),
                datos_excel.get('segundo_nombre', ''),
                datos_excel.get('primer_apellido', ''),
                datos_excel.get('segundo_apellido', ''),
            ),
        },
        {
            'descripcion': 'padre',
            'documento': valor_excel_familia(datos_excel.get('padre_documento', '')),
            'nombre': valor_excel_familia(datos_excel.get('padre_nombre', '')),
        },
        {
            'descripcion': 'madre',
            'documento': valor_excel_familia(datos_excel.get('madre_documento', '')),
            'nombre': valor_excel_familia(datos_excel.get('madre_nombre', '')),
        },
    ]

    if responsable_adicional_es_valido(datos_excel):
        personas_objetivo.append({
            'descripcion': 'responsable',
            'documento': valor_excel_familia(datos_excel.get('responsable_documento', '')),
            'nombre': valor_excel_familia(datos_excel.get('responsable_nombre', '')),
        })

    def _coincide_persona(fila, persona):
        documento_persona = normalizar_texto(persona.get('documento', ''))
        nombre_persona = normalizar_texto(persona.get('nombre', ''))
        documento_fila = normalizar_texto(fila.get('documento', ''))
        nombre_fila = normalizar_texto(fila.get('nombre', ''))
        texto_fila = normalizar_texto(fila.get('texto', ''))

        if documento_persona and documento_fila == documento_persona:
            return True
        if nombre_persona and (nombre_persona == nombre_fila or nombre_persona in texto_fila or nombre_fila in nombre_persona):
            return True
        return False

    def _puntaje_conservacion(fila, persona):
        puntaje = 0
        documento_persona = normalizar_texto(persona.get('documento', ''))
        nombre_persona = normalizar_texto(persona.get('nombre', ''))
        if documento_persona and normalizar_texto(fila.get('documento', '')) == documento_persona:
            puntaje += 100
        if nombre_persona and normalizar_texto(fila.get('nombre', '')) == nombre_persona:
            puntaje += 20
        if fila.get('boton_eliminar_id'):
            puntaje += 5
        return puntaje

    eliminado_alguno = False
    for persona in personas_objetivo:
        if not persona.get('documento') and not persona.get('nombre'):
            continue

        filas = leer_filas_grupo_familiar()
        coincidencias = [fila for fila in filas if _coincide_persona(fila, persona)]
        if len(coincidencias) <= 1:
            continue

        coincidencias_ordenadas = sorted(coincidencias, key=lambda fila: _puntaje_conservacion(fila, persona), reverse=True)
        fila_a_conservar = coincidencias_ordenadas[0]
        print(f"  [~] Duplicados detectados para {persona['descripcion']} '{persona.get('nombre', '')}': {len(coincidencias)} filas. Se conservará documento={fila_a_conservar.get('documento', '')}")

        for fila in coincidencias_ordenadas[1:]:
            boton_eliminar_id = fila.get('boton_eliminar_id', '')
            if not boton_eliminar_id:
                print(f"  [!] Duplicado sin botón eliminar para {persona['descripcion']}: documento={fila.get('documento', '')} nombre={fila.get('nombre', '')}")
                continue
            if eliminar_persona_grupo_familiar(boton_eliminar_id):
                print(f"  [+] Duplicado eliminado para {persona['descripcion']}: documento={fila.get('documento', '')} nombre={fila.get('nombre', '')}")
                eliminado_alguno = True

    return eliminado_alguno


def limpiar_integrantes_ajenos_al_excel(datos_excel):
    """Revisa cada fila de la grilla del Grupo Familiar:
    - Si el documento (o nombre) coincide con el beneficiario/padre/madre del Excel → se deja;
      el parentesco se verifica y corrige más adelante en verificar_y_ajustar_grupo_familiar.
    - Si NO existe en el Excel → se elimina.
      Si no hay botón eliminar → se sobreescribe con la persona del Excel que falte en la grilla.
    """
    print("  [*] Revisando si hay integrantes ajenos al Excel en Grupo Familiar...")
    doc_ben   = normalizar_texto(datos_excel.get('documento', ''))
    doc_padre = normalizar_texto(valor_excel_familia(datos_excel.get('padre_documento', '')))
    doc_madre = normalizar_texto(valor_excel_familia(datos_excel.get('madre_documento', '')))
    doc_responsable = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_documento', ''))) if responsable_adicional_es_valido(datos_excel) else ''
    docs_validos = {d for d in [doc_ben, doc_padre, doc_madre, doc_responsable] if d}

    # También comparar por nombre (fallback por si el doc en grilla difiere del Excel)
    nombre_padre = normalizar_texto(valor_excel_familia(datos_excel.get('padre_nombre', '')))
    nombre_madre = normalizar_texto(valor_excel_familia(datos_excel.get('madre_nombre', '')))
    nombre_responsable = normalizar_texto(valor_excel_familia(datos_excel.get('responsable_nombre', ''))) if responsable_adicional_es_valido(datos_excel) else ''
    nombre_ben   = normalizar_texto(nombre_completo(
        datos_excel.get('primer_nombre', ''), datos_excel.get('segundo_nombre', ''),
        datos_excel.get('primer_apellido', ''), datos_excel.get('segundo_apellido', ''),
    ))

    def es_valido_por_nombre(nombre_fila):
        n = normalizar_texto(nombre_fila)
        if not n:
            return False
        for nombre_ref in [nombre_ben, nombre_padre, nombre_madre, nombre_responsable]:
            if nombre_ref and (nombre_ref in n or n in nombre_ref):
                return True
        return False

    parentesco_padre_obj, parentesco_madre_obj, _ = calcular_parentescos_padres_objetivo(datos_excel)

    eliminado_alguno = False
    for _ in range(10):  # máximo 10 iteraciones por seguridad
        filas = leer_filas_grupo_familiar()
        # Determinar qué personas del Excel ya están en la grilla
        docs_en_grilla = {normalizar_texto(f.get('documento', '')) for f in filas if f.get('documento')}
        padre_falta = bool(doc_padre) and doc_padre not in docs_en_grilla
        madre_falta = bool(doc_madre) and doc_madre not in docs_en_grilla

        # Buscar el primer ajeno (no existe por documento NI por nombre en el Excel)
        ajeno = None
        for fila in filas:
            doc = normalizar_texto(fila.get('documento', ''))
            if not doc:
                continue
            if doc in docs_validos:
                continue  # Está en Excel por documento → OK
            if es_valido_por_nombre(fila.get('nombre', '')):
                continue  # Está en Excel por nombre → OK (doc levemente distinto)
            # No está ni por doc ni por nombre → es ajeno
            ajeno = fila
            break
        if not ajeno:
            break

        doc_raw = ajeno.get('documento', '')
        nombre_raw = ajeno.get('nombre', '')
        eliminar_id = ajeno.get('boton_eliminar_id', '')
        detalle_id = ajeno.get('boton_detalle_id', '')

        # Determinar con qué persona del Excel reemplazar (si no se puede eliminar)
        persona_reemplazo = None
        parentesco_reemplazo = ''
        if padre_falta:
            persona_reemplazo = construir_persona_desde_excel(datos_excel, 'padre')
            parentesco_reemplazo = parentesco_padre_obj
        elif madre_falta:
            persona_reemplazo = construir_persona_desde_excel(datos_excel, 'madre')
            parentesco_reemplazo = parentesco_madre_obj
        elif doc_responsable and doc_responsable not in docs_en_grilla:
            persona_reemplazo = construir_persona_desde_excel(datos_excel, 'responsable')
            parentesco_reemplazo = 'JEFE DEL GRUPO FAMILIAR' if determinar_jefe_objetivo(datos_excel) == 'responsable' else 'PARIENTE U OTRO'

        # Intento 1: eliminar
        eliminado = False
        if eliminar_id:
            print(f"  [*] Intentando eliminar integrante ajeno: {nombre_raw} ({doc_raw})")
            eliminado = eliminar_persona_grupo_familiar(eliminar_id)
            if eliminado:
                print(f"  [+] Eliminado: {nombre_raw} ({doc_raw})")
                eliminado_alguno = True
                continue

        # Intento 2: si no se pudo eliminar y hay reemplazo, abrir detalle y sobreescribir
        if not eliminado and persona_reemplazo and detalle_id:
            print(f"  [*] No se pudo eliminar '{nombre_raw}' ({doc_raw}), reemplazando con {persona_reemplazo['nombre_completo']} ({persona_reemplazo['documento']})...")
            try:
                boton = wait.until(EC.element_to_be_clickable((By.ID, detalle_id)))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton)
                time.sleep(0.5)
                try:
                    boton.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", boton)
                wait.until(EC.presence_of_element_located((By.ID, "cphCont_TabContainer1_tbngrupofamiliar_InformacionGrupoFamiliar_ddlIdParentescoJefe")))
                esperar_postback_finalizado(1.0)
                esperar_detalle_grupo_familiar(1.0)
                # Sobreescribir con llenar_formulario_persona_grupo_familiar (que llena todos los campos)
                llenar_formulario_persona_grupo_familiar(persona_reemplazo, parentesco_reemplazo, consultar_por_id=False)
                if guardar_persona_grupo_familiar():
                    print(f"  [+] Reemplazado: '{nombre_raw}' → {persona_reemplazo['nombre_completo']}")
                    eliminado_alguno = True
                    # Marcar el reemplazo como válido para no intentar de nuevo
                    docs_validos.add(normalizar_texto(persona_reemplazo['documento']))
                    esperar_grupo_familiar_listo(1.2)
                    continue
            except Exception as e:
                print(f"  [!] No se pudo reemplazar '{nombre_raw}': {e}")

        # Si ni eliminar ni reemplazar funcionaron, evitar loop infinito
        print(f"  [!] No se pudo limpiar o reemplazar '{nombre_raw}' ({doc_raw}), continuando...")
        break
    return eliminado_alguno


def verificar_y_ajustar_grupo_familiar(datos_excel):
    resultado = {
        'padre_encontrado': False,
        'madre_encontrada': False,
        'responsable_extra_encontrado': not responsable_adicional_es_valido(datos_excel),
        'jefe_ok': False,
        'jefe_documento': '',
        'padre_omitido_duplicidad': False,
        'madre_omitida_duplicidad': False,
        'padre_duplicidad_mensaje': '',
        'madre_duplicidad_mensaje': '',
        'padre_parentesco_ok': False,
        'madre_parentesco_ok': False,
        'beneficiario_parentesco_jefe_ok': False,
        'beneficiario_parentesco_jefe_esperado': determinar_parentesco_beneficiario_desde_excel(datos_excel),
        'beneficiario_parentesco_jefe_final': '',
        'responsable_ok': False,
        'filas_incompletas': [],
        'incidencias': [],
        'jefe_excel': datos_excel.get('jefe_hogar', ''),
        'padre_esperado': '',
        'madre_esperada': '',
        'responsable_extra_esperado': valor_excel_familia(datos_excel.get('responsable_nombre', '')) if responsable_adicional_es_valido(datos_excel) else '',
        'padre_parentesco_esperado': '',
        'madre_parentesco_esperado': '',
        'padre_parentesco_final': '',
        'madre_parentesco_final': '',
    }

    if not abrir_tab_grupo_familiar():
        return resultado

    # Primero eliminar cualquier integrante ajeno al Excel
    limpiar_integrantes_ajenos_al_excel(datos_excel)
    limpiar_duplicados_integrantes_validos(datos_excel)

    nombre_padre = valor_excel_familia(datos_excel.get('padre_nombre', ''))
    nombre_madre = valor_excel_familia(datos_excel.get('madre_nombre', ''))
    documento_padre_excel = valor_excel_familia(datos_excel.get('padre_documento', ''))
    documento_madre_excel = valor_excel_familia(datos_excel.get('madre_documento', ''))
    nombre_beneficiario = nombre_completo(
        datos_excel.get('primer_nombre', ''),
        datos_excel.get('segundo_nombre', ''),
        datos_excel.get('primer_apellido', ''),
        datos_excel.get('segundo_apellido', ''),
    )
    jefe_excel = normalizar_texto(valor_excel_familia(datos_excel.get('jefe_hogar', '')))

    persona_padre = construir_persona_desde_excel(datos_excel, 'padre')
    persona_madre = construir_persona_desde_excel(datos_excel, 'madre')
    persona_responsable_extra = construir_persona_desde_excel(datos_excel, 'responsable') if responsable_adicional_es_valido(datos_excel) else None

    padre_requerido = bool(nombre_padre or documento_padre_excel)
    madre_requerida = bool(nombre_madre or documento_madre_excel)
    resultado['padre_requerido'] = padre_requerido
    resultado['madre_requerida'] = madre_requerida
    if not padre_requerido:
        resultado['padre_encontrado'] = True
        resultado['padre_parentesco_ok'] = True
    if not madre_requerida:
        resultado['madre_encontrada'] = True
        resultado['madre_parentesco_ok'] = True

    fila_padre = None
    filas = []
    if padre_requerido and persona_padre:
        fila_padre, filas, _ = buscar_fila_grupo_familiar_existente(persona_padre)
        resultado['padre_encontrado'] = fila_padre is not None

    fila_madre = None
    if madre_requerida and persona_madre:
        fila_madre, _, _ = buscar_fila_grupo_familiar_existente(persona_madre)
        resultado['madre_encontrada'] = fila_madre is not None

    fila_responsable_extra = None
    if persona_responsable_extra:
        fila_responsable_extra, _, _ = buscar_fila_grupo_familiar_existente(persona_responsable_extra)
        resultado['responsable_extra_encontrado'] = fila_responsable_extra is not None

    parentesco_padre_objetivo, parentesco_madre_objetivo, jefe_objetivo_actual = calcular_parentescos_padres_objetivo(datos_excel)

    print("\n  [GRUPO FAMILIAR - EXPECTATIVAS]:")
    print(f"    Padre Excel: {nombre_padre}")
    print(f"    Madre Excel: {nombre_madre}")
    print(f"    Jefe de hogar Excel: {datos_excel.get('jefe_hogar', '')}")
    print(f"    Padre debe quedar como: {parentesco_padre_objetivo}")
    print(f"    Madre debe quedar como: {parentesco_madre_objetivo}")
    if persona_responsable_extra:
        print(f"    Responsable adicional Excel: {persona_responsable_extra['nombre_completo']}")

    resultado['padre_esperado'] = nombre_padre
    resultado['madre_esperada'] = nombre_madre
    resultado['padre_parentesco_esperado'] = parentesco_padre_objetivo
    resultado['madre_parentesco_esperado'] = parentesco_madre_objetivo

    def aplicar_adulto_disponible_como_jefe_y_responsable():
        persona_disponible = None
        if resultado['padre_encontrado'] and not resultado['madre_encontrada']:
            persona_disponible = persona_padre
        elif resultado['madre_encontrada'] and not resultado['padre_encontrado']:
            persona_disponible = persona_madre

        if not persona_disponible:
            return False

        persona_disponible['es_responsable'] = True
        parentesco_disponible = 'JEFE DEL GRUPO FAMILIAR'
        print(f"  [*] Forzando adulto disponible como jefe y responsable: {persona_disponible['nombre_completo']}")
        return actualizar_parentesco_grupo_familiar(
            persona_disponible['nombre_completo'],
            parentesco_disponible,
            persona_disponible['documento'],
            persona_disponible,
        )

    def forzar_jefe_y_responsable_con_cualquier_adulto():
        candidatos = []

        responsable_objetivo_local = determinar_responsable_objetivo(datos_excel)
        if responsable_objetivo_local == 'padre' and resultado['padre_encontrado']:
            candidatos.append((persona_padre, 'adulto responsable esperado'))
        elif responsable_objetivo_local == 'madre' and resultado['madre_encontrada']:
            candidatos.append((persona_madre, 'adulto responsable esperado'))
        elif responsable_objetivo_local == 'responsable' and resultado['responsable_extra_encontrado'] and persona_responsable_extra:
            candidatos.append((persona_responsable_extra, 'responsable adicional esperado'))

        jefe_objetivo_local = determinar_jefe_objetivo(datos_excel)
        if jefe_objetivo_local == 'padre' and resultado['padre_encontrado']:
            candidatos.append((persona_padre, 'jefe esperado'))
        elif jefe_objetivo_local == 'madre' and resultado['madre_encontrada']:
            candidatos.append((persona_madre, 'jefe esperado'))
        elif jefe_objetivo_local == 'responsable' and resultado['responsable_extra_encontrado'] and persona_responsable_extra:
            candidatos.append((persona_responsable_extra, 'jefe esperado'))

        if resultado['padre_encontrado']:
            candidatos.append((persona_padre, 'padre disponible'))
        if resultado['madre_encontrada']:
            candidatos.append((persona_madre, 'madre disponible'))
        if resultado['responsable_extra_encontrado'] and persona_responsable_extra:
            candidatos.append((persona_responsable_extra, 'responsable adicional disponible'))

        vistos = set()
        for persona_candidata, razon in candidatos:
            if not persona_candidata:
                continue
            clave = normalizar_texto(persona_candidata.get('documento', '')) or normalizar_texto(persona_candidata.get('nombre_completo', ''))
            if not clave or clave in vistos:
                continue
            vistos.add(clave)
            persona_candidata['es_responsable'] = True
            print(f"  [*] Forzando jefe y responsable con {persona_candidata['nombre_completo']} ({razon})...")
            if actualizar_parentesco_grupo_familiar(
                persona_candidata['nombre_completo'],
                'JEFE DEL GRUPO FAMILIAR',
                persona_candidata['documento'],
                persona_candidata,
            ):
                return True, persona_candidata

        return False, None

    if padre_requerido and persona_padre:
        alta_padre = agregar_o_actualizar_persona_grupo_familiar(persona_padre, parentesco_padre_objetivo)
        if alta_padre.get('omitido_duplicidad'):
            resultado['padre_omitido_duplicidad'] = True
            resultado['padre_duplicidad_mensaje'] = alta_padre.get('mensaje', '')
            resultado['incidencias'].append(f"Padre omitido por duplicidad de tipo de documento: {alta_padre.get('mensaje', '')}")
        elif not alta_padre.get('ok', False):
            resultado['incidencias'].append(f"Padre no encontrado ni creado: {nombre_padre} (motivo: {alta_padre.get('mensaje', 'sin detalle')})")

        fila_p, _ = buscar_fila_grupo_familiar_estable(persona_padre['documento'], persona_padre['nombre_completo'])
        resultado['padre_encontrado'] = fila_p is not None
        if fila_p:
            resultado['padre_parentesco_final'] = fila_p.get('parentesco_jefe', '')
            resultado['padre_parentesco_ok'] = normalizar_texto(resultado['padre_parentesco_final']) == normalizar_texto(parentesco_padre_objetivo)
        else:
            print(f"  [!] Padre no encontrado ni creado en Grupo Familiar: {nombre_padre}")

    if madre_requerida and persona_madre:
        alta_madre = agregar_o_actualizar_persona_grupo_familiar(persona_madre, parentesco_madre_objetivo)
        if alta_madre.get('omitido_duplicidad'):
            resultado['madre_omitida_duplicidad'] = True
            resultado['madre_duplicidad_mensaje'] = alta_madre.get('mensaje', '')
            resultado['incidencias'].append(f"Madre omitida por duplicidad de tipo de documento: {alta_madre.get('mensaje', '')}")
        elif not alta_madre.get('ok', False):
            resultado['incidencias'].append(f"Madre no encontrada ni creada: {nombre_madre} (motivo: {alta_madre.get('mensaje', 'sin detalle')})")

        fila_m, _ = buscar_fila_grupo_familiar_estable(persona_madre['documento'], persona_madre['nombre_completo'])
        resultado['madre_encontrada'] = fila_m is not None
        if fila_m:
            resultado['madre_parentesco_final'] = fila_m.get('parentesco_jefe', '')
            resultado['madre_parentesco_ok'] = normalizar_texto(resultado['madre_parentesco_final']) == normalizar_texto(parentesco_madre_objetivo)
        else:
            print(f"  [!] Madre no encontrada ni creada en Grupo Familiar: {nombre_madre}")

    parentesco_responsable_extra = 'JEFE DEL GRUPO FAMILIAR' if jefe_objetivo_actual == 'responsable' else 'PARIENTE U OTRO'
    if persona_responsable_extra:
        alta_responsable = agregar_o_actualizar_persona_grupo_familiar(persona_responsable_extra, parentesco_responsable_extra)
        if not alta_responsable.get('ok', False) and not alta_responsable.get('omitido_duplicidad', False):
            resultado['incidencias'].append(f"Responsable adicional no encontrado ni creado: {persona_responsable_extra['nombre_completo']} (motivo: {alta_responsable.get('mensaje', 'sin detalle')})")
        fila_responsable_extra, _ = buscar_fila_grupo_familiar_estable(persona_responsable_extra['documento'], persona_responsable_extra['nombre_completo'])
        resultado['responsable_extra_encontrado'] = fila_responsable_extra is not None
        if not resultado['responsable_extra_encontrado']:
            print(f"  [!] Responsable adicional no encontrado ni creado en Grupo Familiar: {persona_responsable_extra['nombre_completo']}")

    # Ajustar el beneficiario en la misma pasada principal para evitar una edición separada al final.
    parentesco_beneficiario_objetivo = determinar_parentesco_beneficiario_desde_excel(datos_excel)
    persona_beneficiario = {
        'documento': datos_excel.get('documento', ''),
        'nombre_completo': nombre_beneficiario,
        'es_responsable': False,
    }
    if persona_beneficiario.get('documento') and parentesco_beneficiario_objetivo:
        print(f"  [*] Ajustando beneficiario en la misma pasada: {nombre_beneficiario}")
        agregar_o_actualizar_persona_grupo_familiar(persona_beneficiario, parentesco_beneficiario_objetivo)

    if resultado['padre_omitido_duplicidad'] or resultado['madre_omitida_duplicidad']:
        if not aplicar_adulto_disponible_como_jefe_y_responsable():
            resultado['incidencias'].append("No fue posible forzar jefe y responsable con el adulto disponible tras una omisión por duplicidad")
        if resultado['padre_omitido_duplicidad'] and resultado['madre_encontrada']:
            parentesco_madre_objetivo = 'JEFE DEL GRUPO FAMILIAR'
            resultado['madre_parentesco_esperado'] = parentesco_madre_objetivo
            jefe_objetivo_actual = 'madre'
            fila_m, _ = buscar_fila_grupo_familiar_estable(persona_madre['documento'], persona_madre['nombre_completo'])
            resultado['madre_parentesco_final'] = fila_m.get('parentesco_jefe', '') if fila_m else ''
            resultado['madre_parentesco_ok'] = normalizar_texto(resultado['madre_parentesco_final']) == normalizar_texto(parentesco_madre_objetivo)
        elif resultado['madre_omitida_duplicidad'] and resultado['padre_encontrado']:
            parentesco_padre_objetivo = 'JEFE DEL GRUPO FAMILIAR'
            resultado['padre_parentesco_esperado'] = parentesco_padre_objetivo
            jefe_objetivo_actual = 'padre'
            fila_p, _ = buscar_fila_grupo_familiar_estable(persona_padre['documento'], persona_padre['nombre_completo'])
            resultado['padre_parentesco_final'] = fila_p.get('parentesco_jefe', '') if fila_p else ''
            resultado['padre_parentesco_ok'] = normalizar_texto(resultado['padre_parentesco_final']) == normalizar_texto(parentesco_padre_objetivo)

    # Si el padre o la madre deberían ser JEFE y no quedó así, forzar una pasada de corrección.
    if (
        (persona_padre.get('documento') and parentesco_padre_objetivo == 'JEFE DEL GRUPO FAMILIAR' and not resultado['padre_parentesco_ok'])
        or (persona_madre.get('documento') and parentesco_madre_objetivo == 'JEFE DEL GRUPO FAMILIAR' and not resultado['madre_parentesco_ok'])
    ):
        print("  [*] Reintentando corrección de jefe del grupo familiar...")
        autocorregir_jefe_grupo_familiar(datos_excel, jefe_forzado=jefe_objetivo_actual)
        fila_p, _ = buscar_fila_grupo_familiar_estable(persona_padre.get('documento', ''), persona_padre.get('nombre_completo', ''))
        fila_m, _ = buscar_fila_grupo_familiar_estable(persona_madre.get('documento', ''), persona_madre.get('nombre_completo', ''))
        if fila_p:
            resultado['padre_parentesco_final'] = fila_p.get('parentesco_jefe', '')
            resultado['padre_parentesco_ok'] = normalizar_texto(resultado['padre_parentesco_final']) == normalizar_texto(parentesco_padre_objetivo)
        if fila_m:
            resultado['madre_parentesco_final'] = fila_m.get('parentesco_jefe', '')
            resultado['madre_parentesco_ok'] = normalizar_texto(resultado['madre_parentesco_final']) == normalizar_texto(parentesco_madre_objetivo)

    filas_finales = leer_filas_grupo_familiar()
    evaluacion_grilla = evaluar_completitud_grupo_familiar(filas_finales, datos_excel)
    if not evaluacion_grilla['jefe_ok'] or not evaluacion_grilla['responsable_ok']:
        forzado_ok, persona_forzada = forzar_jefe_y_responsable_con_cualquier_adulto()
        if not forzado_ok:
            resultado['incidencias'].append("No se pudo dejar jefe y responsable en Grupo Familiar con ningún adulto disponible")
        else:
            print(f"  [+] Jefe y responsable garantizados con: {persona_forzada['nombre_completo']}")
        filas_finales = leer_filas_grupo_familiar()
        evaluacion_grilla = evaluar_completitud_grupo_familiar(filas_finales, datos_excel)
    resultado['beneficiario_parentesco_jefe_final'] = ''
    for fila in filas_finales:
        if normalizar_texto(fila.get('documento', '')) == normalizar_texto(datos_excel.get('documento', '')):
            resultado['beneficiario_parentesco_jefe_final'] = fila.get('parentesco_jefe', '')
            break

    if not evaluacion_grilla['beneficiario_parentesco_jefe_ok']:
        print(f"  [*] Ajustando parentesco del beneficiario en Grupo Familiar: {nombre_beneficiario}")
        actualizar_parentesco_grupo_familiar(
            nombre_beneficiario,
            evaluacion_grilla['beneficiario_parentesco_jefe_esperado'],
            datos_excel.get('documento', ''),
            {
                'es_responsable': False,
            },
        )
        filas_finales = leer_filas_grupo_familiar()
        evaluacion_grilla = evaluar_completitud_grupo_familiar(filas_finales, datos_excel)
        resultado['beneficiario_parentesco_jefe_final'] = ''
        for fila in filas_finales:
            if normalizar_texto(fila.get('documento', '')) == normalizar_texto(datos_excel.get('documento', '')):
                resultado['beneficiario_parentesco_jefe_final'] = fila.get('parentesco_jefe', '')
                break

    responsable_objetivo = determinar_responsable_objetivo(datos_excel)
    if resultado['padre_omitido_duplicidad'] and resultado['madre_encontrada']:
        responsable_objetivo = 'madre'
    elif resultado['madre_omitida_duplicidad'] and resultado['padre_encontrado']:
        responsable_objetivo = 'padre'
    persona_responsable = persona_padre if responsable_objetivo == 'padre' else persona_madre if responsable_objetivo == 'madre' else persona_responsable_extra if responsable_objetivo == 'responsable' else None
    parentesco_responsable = parentesco_padre_objetivo if responsable_objetivo == 'padre' else parentesco_madre_objetivo if responsable_objetivo == 'madre' else parentesco_responsable_extra if responsable_objetivo == 'responsable' else ""
    if (resultado['padre_omitido_duplicidad'] and responsable_objetivo == 'madre') or (resultado['madre_omitida_duplicidad'] and responsable_objetivo == 'padre'):
        parentesco_responsable = 'JEFE DEL GRUPO FAMILIAR'
        if persona_responsable:
            persona_responsable['es_responsable'] = True
    if persona_responsable and not evaluacion_grilla['responsable_ok']:
        fila_resp, _ = buscar_fila_grupo_familiar_estable(persona_responsable.get('documento', ''), persona_responsable.get('nombre_completo', ''))
        if not fila_resp:
            fila_resp_directa, _, coincidencias_resp = buscar_fila_grupo_familiar_existente(persona_responsable)
            if fila_resp_directa or coincidencias_resp:
                print(f"  [~] Responsable ya existe en la grilla; se evitará una nueva creación para {persona_responsable['nombre_completo']}")
            else:
                print(f"  [*] Responsable no visible en grilla, reintentando alta/actualización de {persona_responsable['nombre_completo']}...")
                agregar_o_actualizar_persona_grupo_familiar(persona_responsable, parentesco_responsable)
        print(f"  [*] Ajustando responsable en Grupo Familiar para: {persona_responsable['nombre_completo']}")
        actualizar_parentesco_grupo_familiar(
            persona_responsable['nombre_completo'],
            parentesco_responsable,
            persona_responsable['documento'],
            persona_responsable,
        )
        filas_finales = leer_filas_grupo_familiar()
        evaluacion_grilla = evaluar_completitud_grupo_familiar(filas_finales, datos_excel)

    resultado['beneficiario_parentesco_jefe_ok'] = evaluacion_grilla['beneficiario_parentesco_jefe_ok']
    resultado['beneficiario_parentesco_jefe_esperado'] = evaluacion_grilla['beneficiario_parentesco_jefe_esperado']
    resultado['jefe_ok'] = evaluacion_grilla.get('jefe_ok', False)
    resultado['jefe_documento'] = evaluacion_grilla.get('jefe_documento', '')
    resultado['responsable_ok'] = evaluacion_grilla['responsable_ok']
    resultado['filas_incompletas'] = evaluacion_grilla['filas_incompletas']

    # Cerrar el subformulario interno de Grupo Familiar para que no queden campos
    # obligatorios en blanco al intentar el guardado final del beneficiario.
    try:
        abrir_tab_grupo_familiar()
        if limpiar_formulario_grupo_familiar():
            print("  [+] Subformulario de Grupo Familiar limpiado antes del guardado final")
        time.sleep(1)
    except Exception as e:
        print(f"  [!] No fue posible limpiar el subformulario de Grupo Familiar al final: {e}")

    return resultado

# ========== PASO 2: VERIFICAR CADA REGISTRO ==========
print("\n[*] Iniciando verificación de registros...\n")

resultados = []
reintentos_registro = {}
formulario_inicializado = False
indice_inicio = CONFIG_EJECUCION['indice_inicio']
ejecucion_interrumpida = False
documento_interrumpido = ""

registros_prueba = registros_unidad[indice_inicio:]
documentos_forzados = set(obtener_documentos_forzados_desde_entorno())
if documentos_forzados:
    registros_prueba = [registro for registro in registros_prueba if texto_excel(registro.get('documento')) in documentos_forzados]
    print(f"[*] Filtro activo: reproceso forzado de {len(registros_prueba)} registro(s)")
else:
    documentos_pendientes_reporte = obtener_documentos_con_novedades_desde_ultimo_reporte()
    documentos_pendientes = set(documentos_pendientes_reporte or [])
    if documentos_pendientes_reporte is None:
        pass
    elif documentos_pendientes:
        registros_prueba = [registro for registro in registros_prueba if texto_excel(registro.get('documento')) in documentos_pendientes]
        print(f"[*] Filtro activo: solo se reprocesarán {len(registros_prueba)} registro(s) con novedades del último reporte")
    else:
        documentos_ya_procesados = obtener_documentos_ya_procesados_desde_reportes()
        registros_prueba = [registro for registro in registros_prueba if texto_excel(registro.get('documento')) not in documentos_ya_procesados]
        if registros_prueba:
            print(f"[*] Filtro activo: no hay errores pendientes; se continuará con {len(registros_prueba)} registro(s) aún no procesados")
        else:
            print("[*] Filtro activo: no hay documentos con errores pendientes ni registros nuevos por procesar")

# Validar fechas: omitir registros donde fecha_ingreso < fecha_nacimiento
registros_con_fecha_valida = []
registros_omitidos_fecha = []

for registro in registros_prueba:
    fecha_ingreso = registro.get('fecha_ingreso')
    fecha_nacimiento = registro.get('fecha_nacimiento_beneficiario')
    
    # Convertir a datetime si es necesario
    try:
        from datetime import datetime as dt
        
        # Convertir fecha_ingreso a datetime
        if fecha_ingreso:
            if isinstance(fecha_ingreso, str):
                fecha_ingreso_obj = dt.strptime(fecha_ingreso, "%d/%m/%Y")
            else:
                fecha_ingreso_obj = fecha_ingreso
        else:
            fecha_ingreso_obj = None
        
        # Convertir fecha_nacimiento a datetime
        if fecha_nacimiento:
            if isinstance(fecha_nacimiento, str):
                fecha_nacimiento_obj = dt.strptime(fecha_nacimiento, "%d/%m/%Y")
            else:
                fecha_nacimiento_obj = fecha_nacimiento
        else:
            fecha_nacimiento_obj = None
        
        # Comparar fechas
        if fecha_ingreso_obj and fecha_nacimiento_obj and fecha_ingreso_obj < fecha_nacimiento_obj:
            registros_omitidos_fecha.append(registro)
        else:
            registros_con_fecha_valida.append(registro)
    except Exception:
        # Si hay error en la conversión, incluir el registro
        registros_con_fecha_valida.append(registro)

if registros_omitidos_fecha:
    print(f"\n[!] REGISTROS OMITIDOS: Se omitieron {len(registros_omitidos_fecha)} registro(s) por fecha de ingreso menor a fecha de nacimiento:")
    for reg in registros_omitidos_fecha:
        fecha_ing = reg.get('fecha_ingreso', 'N/A')
        fecha_nac = reg.get('fecha_nacimiento_beneficiario', 'N/A')
        print(f"    - {reg['documento']}: Ingreso {fecha_ing} < Nacimiento {fecha_nac}")

registros_prueba = registros_con_fecha_valida

if LIMITE_REGISTROS_DIAGNOSTICO:
    registros_prueba = registros_prueba[:LIMITE_REGISTROS_DIAGNOSTICO]
print(f"[*] Procesando {len(registros_prueba)} registro(s) de {CONFIG_EJECUCION['descripcion']} desde la posición {indice_inicio + 1}")

try:
    for idx, datos_excel in enumerate(registros_prueba, 1):
        documento_interrumpido = datos_excel['documento']
        print(f"\n{'='*80}")
        print(f"REGISTRO {idx}: {datos_excel['documento']} - {datos_excel['primer_nombre']} {datos_excel['primer_apellido']}")
        print(f"{'='*80}")
        
        print(f"  [EXCEL DATA]:")
        print(f"    Documento: {datos_excel['documento']}")
        print(f"    Nombre: {datos_excel['primer_nombre']} {datos_excel['segundo_nombre']}")
        print(f"    Apellido: {datos_excel['primer_apellido']} {datos_excel['segundo_apellido']}")
        print(f"    Sexo: {datos_excel['sexo']}")
        print(f"    Padre Excel: {datos_excel['padre_nombre']}")
        print(f"    Madre Excel: {datos_excel['madre_nombre']}")
        print(f"    Jefe Hogar Excel: {datos_excel['jefe_hogar']}")
        
        # Buscar el documento en el formulario
        print(f"\n  [BUSCANDO EN FORMULARIO]...")
        
        try:
            if not formulario_inicializado:
                driver.switch_to.default_content()

                if not navegar_a_formulario():
                    raise RuntimeError("No se pudo navegar al formulario")

                if not preparar_formulario_busqueda(seleccionar_tipo_beneficiario=True):
                    raise RuntimeError("No se pudo preparar el formulario")

                formulario_inicializado = True
            else:
                print("  [+] Continuando desde el formulario ya abierto")

            datos_formulario = buscar_y_extraer_datos(datos_excel['documento'])
            beneficiario_creado_desde_excel = False
            foto_cargada_en_alta_nueva = False
            if formulario_basico_sin_informacion(datos_formulario):
                print("  [*] Primera búsqueda sin datos; completando alta nueva en este mismo formulario...")
                resultado_alta_nueva = llenar_datos_basicos_beneficiario_desde_excel(datos_excel)
                beneficiario_creado_desde_excel = resultado_alta_nueva.get('creado_ok', False)
                foto_cargada_en_alta_nueva = resultado_alta_nueva.get('foto_cargada', False)
                datos_formulario = leer_datos_basicos_formulario()
            imprimir_campos_con_valor()

            # Extraer datos que aparecen en el formulario
            print(f"\n  [FORMULARIO DATA]:")
            
            try:
                print(f"    Primer Nombre Formulario: {datos_formulario['primer_nombre']}")
                print(f"    Segundo Nombre Formulario: {datos_formulario['segundo_nombre']}")
                print(f"    Primer Apellido Formulario: {datos_formulario['primer_apellido']}")
                print(f"    Segundo Apellido Formulario: {datos_formulario['segundo_apellido']}")
                print(f"    Sexo Formulario: {datos_formulario['sexo']}")
                print(f"    Fecha Atencion Formulario: {datos_formulario['fecha_atencion']}")
                print(f"    Discapacidad Formulario: {datos_formulario['discapacidad']}")
                
                # Comparar datos
                print(f"\n  [COMPARACIÓN]:")
                
                sexo_excel = normalizar_sexo(datos_excel['sexo'])
                sexo_formulario = normalizar_sexo(datos_formulario['sexo'])
                coincidencias = comparar_datos_basicos_excel_formulario(datos_excel, datos_formulario)
                
                if coincidencias['primer_nombre']:
                    print(f"    ✓ Primer nombre coincide")
                else:
                    print(f"    ✗ Primer nombre NO coincide (Excel: {datos_excel['primer_nombre']}, Form: {datos_formulario['primer_nombre']})")
                
                if coincidencias['segundo_nombre']:
                    print(f"    ✓ Segundo nombre coincide")
                else:
                    print(f"    ✗ Segundo nombre NO coincide (Excel: {datos_excel['segundo_nombre']}, Form: {datos_formulario['segundo_nombre']})")
                
                if coincidencias['primer_apellido']:
                    print(f"    ✓ Primer apellido coincide")
                else:
                    print(f"    ✗ Primer apellido NO coincide (Excel: {datos_excel['primer_apellido']}, Form: {datos_formulario['primer_apellido']})")

                if coincidencias['segundo_apellido']:
                    print(f"    ✓ Segundo apellido coincide")
                else:
                    print(f"    ✗ Segundo apellido NO coincide (Excel: {datos_excel['segundo_apellido']}, Form: {datos_formulario['segundo_apellido']})")
                
                if coincidencias['sexo']:
                    print(f"    ✓ Sexo coincide")
                else:
                    print(f"    ✗ Sexo NO coincide (Excel: {sexo_excel}, Form: {sexo_formulario})")

                if beneficiario_creado_desde_excel:
                    print("    [+] Beneficiario creado/cargado desde Excel tras búsqueda vacía")

                if not beneficiario_creado_desde_excel:
                    completar_campos_faltantes(datos_excel)
                seleccionar_tipo_beneficiario_formulario()
                datos_formulario_actualizado = leer_datos_basicos_formulario()
                try:
                    datos_formulario_actualizado['fecha_atencion'] = driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_cuwFechaAtencion_txtFecha").get_attribute("value") or ""
                except Exception:
                    pass
                try:
                    select_discapacidad = Select(driver.find_element(By.ID, "cphCont_TabContainer1_tbnDatosB_BeneficiarioVincula_ddlPresentaDiscapacidad"))
                    datos_formulario_actualizado['discapacidad'] = select_discapacidad.first_selected_option.text.strip()
                except Exception:
                    pass
                coincidencias = comparar_datos_basicos_excel_formulario(datos_excel, datos_formulario_actualizado)
                print(f"    Fecha Atencion Final: {datos_formulario_actualizado['fecha_atencion']}")
                print(f"    Discapacidad Final: {datos_formulario_actualizado['discapacidad']}")

                foto_cargada = foto_cargada_en_alta_nueva if beneficiario_creado_desde_excel else cargar_foto(datos_excel['documento'])
                print(f"    Foto cargada: {'OK' if foto_cargada else 'NO'}")

                guardado, guardado_mensaje = guardar_formulario(contexto="inicial")
                print(f"    Guardado: {'OK' if guardado else 'NO'}")

                ubicacion_ok = False
                comparacion_ubicacion = {}
                datos_ubicacion = {}
                if guardado and abrir_tab_ubicacion():
                    # Verificar primero si la ubicación ya está correcta
                    datos_ubicacion = leer_datos_ubicacion()
                    comparacion_ubicacion = comparar_datos_ubicacion(datos_excel, datos_ubicacion)
                    modo_georreferencia = 'excel' if excel_tiene_georreferencia(datos_excel) else ('existente_sin_excel' if normalizar_texto(datos_ubicacion.get('georeferenciado', '')) == 'SI' else 'sin_georreferencia')
                    tiene_georreferencia = excel_tiene_georreferencia(datos_excel)
                    debe_gestionar_georreferencia = tiene_georreferencia or (normalizar_texto(datos_ubicacion.get('georeferenciado', '')) == 'SI')
                    validar_coordenadas_excel = tiene_georreferencia
                    comparacion_ubicacion['fecha_captura'] = True if not debe_gestionar_georreferencia else (normalizar_texto(datos_ubicacion.get('fecha_captura', '')) == normalizar_texto(FECHA_CAPTURA_UBICACION))
                    comparacion_ubicacion['hora_captura'] = True if not debe_gestionar_georreferencia else (normalizar_texto(datos_ubicacion.get('hora_captura', '')) == normalizar_texto(HORA_CAPTURA_UBICACION))
                    comparacion_ubicacion['direccion_resumen'] = direccion_residencia_coincide(datos_excel.get('direccion_residencia', ''), datos_ubicacion.get('direccion_resumen', ''))
                    comparacion_ubicacion['grados_latitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')[0]) == normalizar_texto(datos_ubicacion.get('grados_latitud', '')))
                    comparacion_ubicacion['minutos_latitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')[1]) == normalizar_texto(datos_ubicacion.get('minutos_latitud', '')))
                    comparacion_ubicacion['segundos_latitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')[2]) == normalizar_texto(datos_ubicacion.get('segundos_latitud', '')))
                    comparacion_ubicacion['grados_longitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')[0]) == normalizar_texto(datos_ubicacion.get('grados_longitud', '')))
                    comparacion_ubicacion['minutos_longitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')[1]) == normalizar_texto(datos_ubicacion.get('minutos_longitud', '')))
                    comparacion_ubicacion['segundos_longitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')[2]) == normalizar_texto(datos_ubicacion.get('segundos_longitud', '')))
                    comparacion_ubicacion['georeferenciado'] = normalizar_texto(datos_ubicacion.get('georeferenciado', '')) == ('SI' if debe_gestionar_georreferencia else 'NO')
                    ubicacion_ok = all(comparacion_ubicacion.values()) if comparacion_ubicacion else False

                    if not ubicacion_ok:
                        print("  [*] Ubicación no coincide; corrigiendo...")
                        resultado_ubicacion = corregir_datos_ubicacion(datos_excel, es_alta_nueva=beneficiario_creado_desde_excel)
                        imprimir_campos_ubicacion()
                        datos_ubicacion = leer_datos_ubicacion()
                        comparacion_ubicacion = comparar_datos_ubicacion(datos_excel, datos_ubicacion)
                        modo_georreferencia = (resultado_ubicacion or {}).get('modo_georreferencia', modo_georreferencia)
                        tiene_georreferencia = (resultado_ubicacion or {}).get('tiene_georreferencia_excel', tiene_georreferencia)
                        debe_gestionar_georreferencia = (resultado_ubicacion or {}).get('debe_gestionar_georreferencia', debe_gestionar_georreferencia)
                        validar_coordenadas_excel = tiene_georreferencia
                        comparacion_ubicacion['fecha_captura'] = True if not debe_gestionar_georreferencia else (normalizar_texto(datos_ubicacion.get('fecha_captura', '')) == normalizar_texto(FECHA_CAPTURA_UBICACION))
                        comparacion_ubicacion['hora_captura'] = True if not debe_gestionar_georreferencia else (normalizar_texto(datos_ubicacion.get('hora_captura', '')) == normalizar_texto(HORA_CAPTURA_UBICACION))
                        comparacion_ubicacion['direccion_resumen'] = direccion_residencia_coincide(datos_excel.get('direccion_residencia', ''), datos_ubicacion.get('direccion_resumen', ''))
                        comparacion_ubicacion['grados_latitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')[0]) == normalizar_texto(datos_ubicacion.get('grados_latitud', '')))
                        comparacion_ubicacion['minutos_latitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')[1]) == normalizar_texto(datos_ubicacion.get('minutos_latitud', '')))
                        comparacion_ubicacion['segundos_latitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['latitud_excel'], 'latitud')[2]) == normalizar_texto(datos_ubicacion.get('segundos_latitud', '')))
                        comparacion_ubicacion['grados_longitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')[0]) == normalizar_texto(datos_ubicacion.get('grados_longitud', '')))
                        comparacion_ubicacion['minutos_longitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')[1]) == normalizar_texto(datos_ubicacion.get('minutos_longitud', '')))
                        comparacion_ubicacion['segundos_longitud'] = True if not validar_coordenadas_excel else (normalizar_texto(parsear_georreferencia(datos_excel['longitud_excel'], 'longitud')[2]) == normalizar_texto(datos_ubicacion.get('segundos_longitud', '')))
                        comparacion_ubicacion['georeferenciado'] = normalizar_texto(datos_ubicacion.get('georeferenciado', '')) == ('SI' if debe_gestionar_georreferencia else 'NO')
                        ubicacion_ok = all(comparacion_ubicacion.values()) if comparacion_ubicacion else False
                    else:
                        print("  [=] Ubicación ya correcta; pasando a siguiente pestaña")

                if guardado:
                    completar_pertenencia_etnica(datos_excel)

                grupo_familiar = {}
                grupo_familiar_ok = False
                guardado_final = False
                guardado_final_mensaje = "No se intentó el guardado final"
                siguiente_nino_listo = False
                if guardado:
                    try:
                        grupo_familiar = verificar_y_ajustar_grupo_familiar(datos_excel)
                        grupo_familiar_ok = grupo_familiar_esta_ok(grupo_familiar)
                    finally:
                        # CAMBIO: Siempre intentar guardar después de modificar grupo_familiar
                        # con auto-corrección de hasta 3 intentos ante errores conocidos
                        print("    [*] Guardando al finalizar el proceso de Grupo Familiar...")
                        MAX_INTENTOS_GUARDADO = 3
                        for _intento_guardado in range(MAX_INTENTOS_GUARDADO):
                            guardado_final, guardado_final_mensaje = guardar_formulario(exigir_mensaje_verde=True, contexto="final")
                            if guardado_final:
                                print(f"    [+] Guardado final: OK")
                                break
                            print(f"    [!] Guardado final NO exitoso (intento {_intento_guardado + 1}/{MAX_INTENTOS_GUARDADO}): {guardado_final_mensaje}")
                            if _intento_guardado < MAX_INTENTOS_GUARDADO - 1:
                                se_corrigio = autocorregir_desde_error(guardado_final_mensaje, datos_excel)
                                if not se_corrigio:
                                    print("    [!] No hay corrector para este error, se omiten reintentos.")
                                    break
                                # Volver a la pestaña datos básicos antes de reintentar guardar
                                abrir_tab_grupo_familiar()

                    if guardado_final:
                        siguiente_nino_listo = iniciar_nuevo_registro()
                        print(f"    Siguiente niño listo: {'OK' if siguiente_nino_listo else 'NO'}")
                else:
                    guardado_final_mensaje = "No se intentó el guardado final porque falló el guardado inicial"
                
                resultado = {
                    'documento': datos_excel['documento'],
                    'coincide': all(coincidencias.values()),
                    'foto_cargada': foto_cargada,
                    'guardado': guardado,
                    'guardado_mensaje': guardado_mensaje,
                    'guardado_final': guardado_final,
                    'guardado_final_mensaje': guardado_final_mensaje,
                    'siguiente_nino_listo': siguiente_nino_listo,
                    'ubicacion_ok': ubicacion_ok,
                    'grupo_familiar_ok': grupo_familiar_ok,
                    'excel_basico': {
                        'primer_nombre': datos_excel['primer_nombre'],
                        'segundo_nombre': datos_excel['segundo_nombre'],
                        'primer_apellido': datos_excel['primer_apellido'],
                        'segundo_apellido': datos_excel['segundo_apellido'],
                        'sexo': sexo_excel,
                    },
                    'formulario_basico': {
                        'primer_nombre': datos_formulario_actualizado['primer_nombre'],
                        'segundo_nombre': datos_formulario_actualizado['segundo_nombre'],
                        'primer_apellido': datos_formulario_actualizado['primer_apellido'],
                        'segundo_apellido': datos_formulario_actualizado['segundo_apellido'],
                        'sexo': normalizar_sexo(datos_formulario_actualizado['sexo']),
                    },
                    'beneficiario_creado_desde_excel': beneficiario_creado_desde_excel,
                    'excel_ubicacion': {
                        'pais_residencia': datos_excel['pais_residencia'],
                        'departamento_residencia': datos_excel['departamento_residencia'],
                        'municipio_residencia': datos_excel['municipio_residencia'],
                        'zona_residencia': datos_excel['zona_residencia'],
                        'direccion_resumen': datos_excel['direccion_residencia'],
                        'telefono': datos_excel['telefono'],
                        'telefono_original_excel': datos_excel.get('telefono_original_excel', datos_excel['telefono']),
                        'fecha_captura': FECHA_CAPTURA_UBICACION,
                        'hora_captura': HORA_CAPTURA_UBICACION,
                    },
                    'formulario_ubicacion': datos_ubicacion if guardado and comparacion_ubicacion else {},
                    'detalles_grupo_familiar': grupo_familiar,
                    'detalles_ubicacion': comparacion_ubicacion,
                    'detalles': coincidencias
                }
                resultados.append(resultado)
                formulario_inicializado = siguiente_nino_listo
            
            except Exception as e:
                if error_es_reintentable_registro(e) and reintentos_registro.get(datos_excel['documento'], 0) < 1:
                    reintentos_registro[datos_excel['documento']] = reintentos_registro.get(datos_excel['documento'], 0) + 1
                    print(f"    [~] Error transitorio durante el procesamiento ({str(e).splitlines()[0][:120]}). Se reintentará el registro completo...")
                    formulario_inicializado = False
                    TAB_ACTIVA = ""
                    registros_prueba.insert(idx, datos_excel)
                    time.sleep(1.2)
                    continue
                print(f"    [!] Error extrayendo datos del formulario: {str(e)}")
                formulario_inicializado = False
                resultado = {
                    'documento': datos_excel['documento'],
                    'coincide': False,
                    'guardado': False,
                    'guardado_mensaje': 'No se completó por error durante el procesamiento',
                    'guardado_final': False,
                    'guardado_final_mensaje': 'No se completó por error durante el procesamiento',
                    'error': str(e)
                }
                resultados.append(resultado)
        
        except Exception as e:
            if error_es_reintentable_registro(e) and reintentos_registro.get(datos_excel['documento'], 0) < 1:
                reintentos_registro[datos_excel['documento']] = reintentos_registro.get(datos_excel['documento'], 0) + 1
                print(f"    [~] Error transitorio durante la búsqueda ({str(e).splitlines()[0][:120]}). Se reintentará el registro completo...")
                formulario_inicializado = False
                TAB_ACTIVA = ""
                registros_prueba.insert(idx, datos_excel)
                time.sleep(1.2)
                continue
            print(f"    [!] Error en búsqueda: {str(e)}")
            formulario_inicializado = False
            resultado = {
                'documento': datos_excel['documento'],
                'coincide': False,
                'guardado': False,
                'guardado_mensaje': 'No se completó por error durante la búsqueda',
                'guardado_final': False,
                'guardado_final_mensaje': 'No se completó por error durante la búsqueda',
                'error': str(e)
            }
            resultados.append(resultado)
except KeyboardInterrupt:
    ejecucion_interrumpida = True
    print("\n[!] Ejecución interrumpida manualmente")
    if documento_interrumpido:
        resultados.append({
            'documento': documento_interrumpido,
            'coincide': False,
            'guardado': False,
            'guardado_mensaje': 'No se completó por interrupción manual',
            'guardado_final': False,
            'guardado_final_mensaje': 'No se completó por interrupción manual',
            'error': 'Ejecución interrumpida manualmente por el usuario'
        })

# ========== RESUMEN ==========
print(f"\n\n{'='*80}")
print("RESUMEN DE VERIFICACIÓN")
print(f"{'='*80}")

if ejecucion_interrumpida and documento_interrumpido:
    print(f"\nEjecución parcial. Último documento en proceso: {documento_interrumpido}")

exitosos = sum(1 for r in resultados if r['coincide'])
fallidos = len(resultados) - exitosos

print(f"\nTotal registros verificados: {len(resultados)}")
print(f"Coinciden: {exitosos}")
print(f"No coinciden: {fallidos}")

print("\n[DETALLES]:")
for r in resultados:
    estado = "✓" if r['coincide'] else "✗"
    print(f"  {estado} Documento {r['documento']}")
    if 'error' not in r:
        print(f"      Detalles: {r['detalles']}")
        if 'guardado' in r:
            print(f"      Guardado: {r['guardado']}")
        if 'guardado_final' in r:
            print(f"      Guardado final: {r['guardado_final']}")
        if 'siguiente_nino_listo' in r:
            print(f"      Siguiente niño listo: {r['siguiente_nino_listo']}")
        if 'ubicacion_ok' in r:
            print(f"      Ubicación OK: {r['ubicacion_ok']}")
        if r.get('detalles_ubicacion'):
            print(f"      Detalles ubicación: {r['detalles_ubicacion']}")
        if 'grupo_familiar_ok' in r:
            print(f"      Grupo Familiar OK: {r['grupo_familiar_ok']}")
        if r.get('detalles_grupo_familiar'):
            print(f"      Detalles grupo familiar: {r['detalles_grupo_familiar']}")
    else:
        print(f"      Error: {r['error']}")

ruta_reporte = generar_reporte_inconsistencias(resultados)

print("\n[*] El navegador quedará abierto para verificación manual]\n")
print("[+] Chrome permanece abierto; cierre la ventana manualmente cuando termine de revisar")
